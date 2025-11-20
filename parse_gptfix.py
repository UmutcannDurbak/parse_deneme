"""
Focused Donuk (DONDURMALAR) writer.
- Parses CSV and writes only the ice cream section of the Donuk workbook.
- Detects size columns (3,5 KG | 350 GR | 150 GR) and flavor rows.
- Clears numeric values in that block and writes aggregated quantities.

This module intentionally keeps only what's needed now.
"""
from __future__ import annotations

from typing import Dict, Iterable, Optional, Tuple
import os
import re
import unicodedata
import pandas as pd  # pyright: ignore[reportMissingImports]
import openpyxl   # pyright: ignore[reportMissingModuleSource]
from openpyxl import load_workbook, Workbook # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Alignment # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]

DATA_START_ROW = 3

# ----------------------------- Normalization -----------------------------

def normalize_text(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s)
    tr = str.maketrans({
        "ı": "i", "ğ": "g", "ş": "s", "ö": "o", "ç": "c", "ü": "u",
        "İ": "I", "Ğ": "G", "Ş": "S", "Ö": "O", "Ç": "C", "Ü": "U",
    })
    s = s.translate(tr)
    s = s.upper()
    s = s.replace("GOGUSLU", "GOGSU").replace("GOGSULU", "GOGSU")
    s = s.replace("HARMANDALI", "EFESUS")
    s = s.replace("AMASRA", "DADAYLI")
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def size_from_stock_or_unit(stock_name: str, unit_text: str) -> Optional[str]:
    # Preserve punctuation for size patterns like 1*3,50 and 350 GR
    def up_keep_punct(s: str) -> str:
        if s is None:
            return ""
        try:
            s = unicodedata.normalize("NFKD", str(s))
        except Exception:
            s = str(s)
        tr = str.maketrans({
            "ı": "i", "ğ": "g", "ş": "s", "ö": "o", "ç": "c", "ü": "u",
            "İ": "I", "Ğ": "G", "Ş": "S", "Ö": "O", "Ç": "C", "Ü": "U",
        })
        s = s.translate(tr).upper()
        s = s.replace(" ", " ")
        return s

    upn_raw = up_keep_punct(stock_name)
    upu_raw = up_keep_punct(unit_text)

    # Helper checks
    def is_35kg(s: str) -> bool:
        # 1*3,5 | 1x3,5 | 1*3.5 | 3,5 KG | 3,50 KG | KL_3,5_KG (do NOT match bare 350)
        if re.search(r"1\s*[\*Xx]\s*3[,\.]?5\b", s):
            return True
        if re.search(r"\b3[,\.]?5\s*KG\b", s):
            return True
        if re.search(r"\b3[,\.]50\s*KG\b", s):  # treat 3,50 KG as 3.5 KG, but not bare 350
            return True
        if re.search(r"KL[\s_\-]*3[,\.]?5", s):
            return True
        return False

    def is_350gr(s: str) -> bool:
        return re.search(r"\b350\s*(GR|G)\b", s) is not None

    def is_150gr(s: str) -> bool:
        return re.search(r"\b150\s*(GR|G)\b", s) is not None

    if is_35kg(upn_raw) or is_35kg(upu_raw):
        return "35KG"
    if is_350gr(upn_raw) or is_350gr(upu_raw):
        return "350GR"
    if is_150gr(upn_raw) or is_150gr(upu_raw):
        return "150GR"
    return None

# --------------------------- CSV Utilities ---------------------------

def read_csv(csv_path: str) -> pd.DataFrame:
    try:
        return pd.read_csv(csv_path, encoding="utf-8", delimiter=",", header=2)
    except Exception:
        return pd.read_csv(csv_path, encoding="utf-8", delimiter=",", header=0)


def find_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    cols_up = {normalize_text(c): c for c in df.columns}
    for c in candidates:
        cc = normalize_text(c)
        if cc in cols_up:
            return cols_up[cc]
    for c in candidates:
        cc = normalize_text(c)
        for up, orig in cols_up.items():
            if cc in up:
                return orig
    return None


def read_branch_from_file(csv_path: str) -> tuple[Optional[str], Optional[str]]:
    """Extract branch name from CSV file with primary and fallback options.
    
    Returns (primary, fallback) where:
    - primary: Inner part from "OUTER(INNER)" format - should be tried first
    - fallback: Outer part - use if primary doesn't match
    
    Example:
    - "AYDIN(KUSADASI)" -> returns ("KUSADASI", "AYDIN")
    - "MANISA(MEYDAN)" -> returns ("MEYDAN", "MANISA")
    
    The fuzzy matcher will try primary first, then fallback if no match found.
    
    Args:
        csv_path: Path to CSV file
    
    Returns:
        Tuple of (primary_branch, fallback_branch)
    """
    try:
        with open(csv_path, encoding="utf-8") as f:
            for line in f:
                up = normalize_text(line)
                if "SUBE" in up and ("KODU" in up or "ADI" in up):
                    # Extract part after colon
                    part = line.split(":", 1)[-1] if ":" in line else line
                    part = part.strip()
                    # Remove quotes if present
                    part = part.strip('"').strip("'").strip()
                    # Split by dash and take the part after it
                    if "-" in part:
                        part = part.split("-", 1)[-1]
                    part = part.strip()
                    
                    # If parens exist, return (inner, outer) for priority matching
                    m = re.search(r"^([^(]+)\(([^)]+)\)$", part)
                    if m:
                        outer = m.group(1).strip()
                        inner = m.group(2).strip()
                        # PRIMARY: inner (parantez içi) - try this first
                        # FALLBACK: outer (parantez dışı) - try if primary fails
                        return (inner, outer)
                    
                    # No parens - return single value for both
                    if part.upper().endswith(" DEPO"):
                        part = part[:-5].strip()
                    return (part, None)
    except Exception:
        pass
    return (None, None)

# --------------------------- Excel helpers ---------------------------

def master_cell(ws: openpyxl.worksheet.worksheet.Worksheet, r: int, c: int):
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mr.bounds
        if min_row <= r <= max_row and min_col <= c <= max_col:
            return ws.cell(row=min_row, column=min_col)
    return ws.cell(row=r, column=c)

def locate_donuk_products_block(ws: openpyxl.worksheet.worksheet.Worksheet, min_c: int, max_c: int, branch_name: str, debug: bool = False) -> Dict[str, Tuple[int, int, str]]:
    """Locate frozen ('DONUK') products within branch span and build map for each variant.
    Uses simple text matching to find products, without scoring or variant handling.
    
    Args:
        ws: Worksheet to scan
        min_c: Branch start column 
        max_c: Branch end column
        branch_name: Name of the branch to match
        debug: Enable debug logging
        
    Returns:
        Dict mapping normalized variant text -> (row_index, column_index, original_text)
    """
    res: Dict[str, Tuple[int, int, str]] = {}
    
    # Use the provided branch span (min_c to max_c) as branch columns
    # These are already determined by find_branch_span earlier
    if debug:
        print(f"[DEBUG] Using branch span cols {min_c}-{max_c} for '{branch_name}'")
        print(f"[DEBUG] Worksheet: '{ws.title}', max_row={ws.max_row}, max_column={ws.max_column}")

    # Sections to skip/ignore (keep these to avoid section confusion)
    skip_sections = {
        "MAKARON", "PASTA", "DONDURMA", "CHEESECAKE", "CATAL", "ÇATAL", 
        "BOREK", "BÖREK", "TATLI", "KUNEFE", "SERBET", "TRILECE"
    }
    
    # Define exact products to collect (order matters - check longer names first)
    target_products = [
        "CITIR MANTI",
        "MANTI",  # Plain MANTI - in different column (c=12) on same row as BOYOZ
        "CEVIZLI TAHINLI BAKLAVA",  # Must check before "SOGUK BAKLAVA"
        "SOGUK BAKLAVA",
        "BOYOZ",
        "PATATES",
        "HAMBURGER KOFTE",
        "HAMBURGER EKMEGI",  # Must be separate from HAMBURGER KOFTE
        "TAVUK BUT",
        "EKSI MAYALI TOST EKMEGI",
        "ZERDECALLI TOST EKMEGI"
    ]
    
    # CRITICAL FIX: Find DONUK section header first to avoid matching products in wrong sections
    donuk_header_row = None
    for r in range(1, min(ws.max_row + 1, 100)):
        for c in range(1, min(ws.max_column + 1, 5)):  # Check first few columns for section headers
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            up_v = normalize_text(v)
            # Look for DONUK section marker (DONUK, DONUK ÜRÜNLER, etc.)
            if "DONUK" in up_v and "DONDURMA" not in up_v:
                donuk_header_row = r
                if debug:
                    print(f"[DEBUG] Found DONUK section header at row {r}: '{v}'")
                break
        if donuk_header_row:
            break
    
    # If DONUK header not found, default to row 10 (avoid scanning header rows)
    if not donuk_header_row:
        donuk_header_row = 10
        if debug:
            print(f"[DEBUG] No DONUK header found, defaulting to row {donuk_header_row}")
    
    # ONLY scan rows BELOW the DONUK header (and within reasonable range)
    scan_start_row = donuk_header_row + 1
    scan_end_row = min(ws.max_row + 1, donuk_header_row + 50)  # Scan next 50 rows max
    
    if debug:
        print(f"[DEBUG] Scanning for DONUK products in rows {scan_start_row}-{scan_end_row}, cols {min_c}-{max_c}")
    
    # Look for products ONLY in branch columns (min_c to max_c)
    for r in range(scan_start_row, scan_end_row):
        for c in range(min_c, max_c + 1):
            try:
                # Safe cell value access - handle merged cells
                cell = ws.cell(row=r, column=c)
                v = cell.value
            except AttributeError:
                # MergedCell object - try to get master cell value
                try:
                    # Find master cell of this merged range
                    for mr in ws.merged_cells.ranges:
                        min_row, min_col, max_row, max_col = mr.bounds
                        if min_row <= r <= max_row and min_col <= c <= max_col:
                            v = ws.cell(row=min_row, column=min_col).value
                            break
                    else:
                        v = None
                except Exception:
                    v = None
            except Exception:
                v = None
                
            if not v or not isinstance(v, str):
                continue
            
            orig_text = str(v).strip()
            up = normalize_text(orig_text)
            if not up:
                continue
            
            # Remove trailing numbers to handle cells that already have quantities
            # e.g., "CITIR MANTI    1" -> "CITIR MANTI"
            up_clean = re.sub(r'\s+\d+\s*$', '', up).strip()

            # Skip section headers to avoid confusion
            skip_this = False
            for section in skip_sections:
                # Only skip if it's ONLY a section name, not part of product name
                if up_clean == section:
                    skip_this = True
                    break
            if skip_this:
                continue

            # Check for exact matches with target products (order matters!)
            matched = False
            for target in target_products:
                if up_clean == target:
                    # Skip if we already found this product (keep the first/leftmost occurrence)
                    if target in res:
                        continue
                    
                    # CRITICAL FIX: Skip horizontally-merged header cells
                    # Product labels should be in single cells or vertically-merged cells only.
                    # Horizontally-merged cells (spanning multiple columns) are typically headers.
                    merge = is_merged_at(ws, r, c)
                    if merge:
                        master_r, master_c, max_merge_r, max_merge_c = merge
                        # Check if this is a horizontal merge (spans multiple columns)
                        if max_merge_c > master_c:
                            # This is a horizontally-merged header cell, skip it
                            if debug:
                                print(f"[DEBUG] Skipping horizontally-merged header at r={r} c={c} for '{target}' (merge spans cols {master_c}-{max_merge_c})")
                            continue
                        # For vertically-merged cells: Keep the CURRENT row position (r)
                        # This ensures each product in a vertical list gets its own row number
                        # even if they share a merged cell horizontally
                        cell = ws.cell(row=r, column=c)
                    else:
                        cell = ws.cell(row=r, column=c)
                    
                    # Store original text as-is (cleaning will happen during write)
                    res[target] = (r, c, orig_text)
                    
                    if debug:
                        print(f"[DEBUG] Found DONUK product '{target}' at r={r} c={c} text='{orig_text}'")
                    
                    matched = True
                    break
            
            if matched:
                continue
    
    if debug:
        print(f"[DEBUG] locate_donuk_products_block for branch '{branch_name}' collected {len(res)} products: {list(res.keys())}")
    
    return res


def safe_write(ws: openpyxl.worksheet.worksheet.Worksheet, r: int, c: int, value) -> None:
    """Write a value into worksheet cell (r,c) safely handling merged cells.

    If the target is inside a merged range, write into the master cell.
    Does NOT unmerge cells to avoid corrupting the template structure.
    
    CRITICAL: Before writing, reads existing cell value and cleans any previously
    appended qty/unit using clean_text_from_quantities. This ensures we don't
    accumulate duplicates like "1 SPT. 1 SPT." on repeated writes.
    """
    # First attempt: direct write to the target cell
    cell = ws.cell(row=r, column=c)
    
    # Check if it's a MergedCell (read-only)
    from openpyxl.cell.cell import MergedCell
    if isinstance(cell, MergedCell):
        # Find the merged range and write to master cell
        merged = is_merged_at(ws, r, c)
        if merged:
            min_row, min_col, max_row, max_col = merged
            master = ws.cell(row=min_row, column=min_col)
            try:
                master.value = value
                return
            except Exception:
                # If master cell write fails, give up
                # Do NOT unmerge as it corrupts the template
                return
    else:
        # Regular cell, just write
        try:
            cell.value = value
            return
        except Exception:
            # Last attempt with explicit cell access
            try:
                ws.cell(row=r, column=c).value = value
                return
            except Exception:
                return


def find_branch_span(ws: openpyxl.worksheet.worksheet.Worksheet, branch_name: str) -> Optional[Tuple[int, int, int]]:
    if not branch_name:
        return None
    up = normalize_text(branch_name)
    
    # PASS 1: Prefer exact matches in merged ranges (avoids FOLKART matching FOLKART VEGA)
    exact_matches = []
    partial_matches = []
    
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mr.bounds
        v = ws.cell(row=min_row, column=min_col).value
        if not v:
            continue
        vv = normalize_text(v)
        if vv == up:  # Exact match
            exact_matches.append((min_col, max_col, min_row))
        elif up in vv or vv in up:  # Partial match
            partial_matches.append((min_col, max_col, min_row))
    
    # Return exact match if found
    if exact_matches:
        return exact_matches[0]
    
    # PASS 2: Scan early rows for exact matches
    exact_cells = []
    partial_cells = []
    
    for r in range(1, min(25, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            vv = normalize_text(v)
            if vv == up:  # Exact match
                # If inside a merge, return the full span
                for mr in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = mr.bounds
                    if min_row <= r <= max_row and min_col <= c <= max_col:
                        exact_cells.append((min_col, max_col, min_row))
                        break
                else:
                    exact_cells.append((c, c, r))
            elif up in vv or vv in up:  # Partial match
                for mr in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = mr.bounds
                    if min_row <= r <= max_row and min_col <= c <= max_col:
                        partial_cells.append((min_col, max_col, min_row))
                        break
                else:
                    partial_cells.append((c, c, r))
    
    # Return exact cell match if found
    if exact_cells:
        return exact_cells[0]
    
    # PASS 3: Fallback to partial matches (for backward compatibility)
    if partial_matches:
        return partial_matches[0]
    if partial_cells:
        return partial_cells[0]
    
    return None

def is_merged_at(ws: openpyxl.worksheet.worksheet.Worksheet, r: int, c: int) -> Optional[Tuple[int, int, int, int]]:
    for mr in ws.merged_cells.ranges:
        # bounds returns (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = mr.bounds
        if min_row <= r <= max_row and min_col <= c <= max_col:
            return (min_row, min_col, max_row, max_col)
    return None

def resolve_numeric_col(ws: openpyxl.worksheet.worksheet.Worksheet, r: int, c: int, min_c: int, max_c: int) -> int:
    # If target cell is part of a merged region whose master is not this column, shift right until outside merge
    merged = is_merged_at(ws, r, c)
    if merged:
        _, mcol, _, mmax = merged
        # If this column is the master, ok; otherwise move to the right edge of this merge (bounded by branch span)
        if mcol != c:
            c = min(mmax, max_c)
    # Ensure final cell is not part of a label merge; if still merged and master != c, step right
    while True:
        merged = is_merged_at(ws, r, c)
        if not merged:
            break
        _, mcol, _, _ = merged
        if mcol == c:
            break
        if c >= max_c:
            break
        c += 1
    return c


def format_text_with_qty(text: str, qty) -> str:
    """Return text with quantity appended or replaced consistently as 'LEFT = qty'.

    - If the original text contains '=', replace the RHS with the qty.
    - Otherwise, remove any trailing numeric tokens and units (previous appended quantities) and append ' = qty'.
    
    CRITICAL: This function MUST clean the input text to prevent accumulation.
    """
    try:
        fmt_qty = int(qty) if float(qty).is_integer() else qty
    except Exception:
        fmt_qty = qty
    t = str(text or "").strip()
    # Always clean existing qty/unit first
    t_clean = clean_text_from_quantities(t)
    if "=" in t_clean:
        left = t_clean.partition("=")[0].strip()
    else:
        left = t_clean
    if left == "":
        return f"= {fmt_qty}"
    return f"{left} = {fmt_qty}"


def clean_text_from_quantities(text: str) -> str:

    """Sadece sonuna eklenmiş miktar birimlerini (ör. '4 SPT.', '2 KL.', '5 TEPSI', '3 KL.') siler.
    Parantezli (+4) ve '=' gibi assignment marker'lar korunur."""
    t = str(text or "")
    if not t:
        return ""
    t = t.rstrip()

    # Sadece sonuna eklenmiş miktar birimi kalıplarını sil
    # Örnek: 'ÜRÜN ADI 4 SPT.' -> 'ÜRÜN ADI'
    qty_unit_pattern = re.compile(r"(\s*[0-9]+(?:[\.,][0-9]+)?\s*(?:SPT\.|KL\.|TEPSI|TEPSİ))+$", re.IGNORECASE)
    t = qty_unit_pattern.sub("", t).rstrip()

    return t.strip()


def append_text_with_space(text: str, qty, sep: str = "    ") -> str:
    """Append qty to original text using a separator (default: 4 spaces).

    Behavior:
    - Remove any previously appended numeric tokens and units (KL., SPT., TEPSİ) at the end.
    - Return "LEFT<sep>qty" where LEFT is the cleaned original text.
    
    CRITICAL: This function MUST clean the input text before appending to prevent
    accumulation like "PRODUCT 1 SPT. 1 SPT." on repeated writes.
    """
    try:
        fmt_qty = int(qty) if float(qty).is_integer() else qty
    except Exception:
        fmt_qty = qty
    # Always clean existing qty/unit from text before appending new value
    left = clean_text_from_quantities(str(text or ""))
    if left == "":
        return f"{sep}{fmt_qty}"
    return f"{left}{sep}{fmt_qty}"


def find_size_columns(ws: openpyxl.worksheet.worksheet.Worksheet, min_c: int, max_c: int, row_hint: int) -> Dict[str, Optional[int]]:
    sizes: Dict[str, Optional[int]] = {"35KG": None, "350GR": None, "150GR": None}
    def scan_rows(r1: int, r2: int):
        # Scan within branch span, plus small margin (max 4 columns) for size headers
        c_start = max(1, min_c)
        c_end = min(ws.max_column, max_c if max_c >= min_c else min_c)
        # Allow small expansion but respect max_c limit to avoid crossing into next branch
        # For single-column spans, add small margin (up to 4 cols) but cap at max_c + 4
        if c_start == c_end and c_end < ws.max_column:
            c_end = min(ws.max_column, c_end + 4, max_c + 4)  # Limit expansion
        for r in range(max(1, r1), min(ws.max_row, r2) + 1):
            for c in range(c_start, c_end + 1):
                v = ws.cell(row=r, column=c).value
                if not v:
                    continue
                s = normalize_text(v)
                def rightmost_if_merged(rr: int, cc: int) -> int:
                    for mr in ws.merged_cells.ranges:
                        min_row, min_col, max_row, max_col = mr.bounds
                        if min_row <= rr <= max_row and min_col <= cc <= max_col:
                            return max_col
                    return cc
                if ("3,5" in s or ("35" in s and "KG" in s)) and sizes["35KG"] is None:
                    sizes["35KG"] = rightmost_if_merged(r, c)
                if ("350" in s and ("GR" in s or "G" in s)) and sizes["350GR"] is None:
                    sizes["350GR"] = rightmost_if_merged(r, c)
                if ("150" in s and ("GR" in s or "G" in s)) and sizes["150GR"] is None:
                    sizes["150GR"] = rightmost_if_merged(r, c)
    if row_hint:
        scan_rows(row_hint, row_hint + 3)
    if not all(sizes.values()):
        scan_rows(1, 25)
    return sizes


def safe_cell_value(ws: openpyxl.worksheet.worksheet.Worksheet, r: int, c: int):
    """Safely get cell value, handling MergedCell objects."""
    try:
        return ws.cell(row=r, column=c).value
    except AttributeError:
        # MergedCell - find master cell
        try:
            for mr in ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = mr.bounds
                if min_row <= r <= max_row and min_col <= c <= max_col:
                    return ws.cell(row=min_row, column=min_col).value
        except Exception:
            pass
        return None
    except Exception:
        return None


def locate_dondurmalar_block(ws: openpyxl.worksheet.worksheet.Worksheet, min_c: int, max_c: int, debug: bool = False) -> Tuple[int, Dict[str, int]]:
    """Find the 'DONDURMALAR' header row and the size columns on that row within (min_c..max_c, plus small right margin).
    Returns: (header_row_index, size_columns dict with keys '35KG','350GR','150GR', and 'MONO','KUCUK','BUYUK').
    """
    header_row = None
    pasta_cols = {"MONO": None, "KUCUK": None, "BUYUK": None}
    all_pasta_rows = []  # Pasta başlıkları için tüm satırları tut

    for r in range(1, min(ws.max_row, 100) + 1):
        v = safe_cell_value(ws, r, 1)
        if not v:
            continue
        up = normalize_text(v)
        if debug:
            print(f"[DEBUG] Checking row {r} for DONDURMALAR: '{v}' -> '{up}'")
        if "DONDURMALAR" in up:
            if debug:
                print(f"[DEBUG] Found DONDURMALAR header at row {r}")
            header_row = r
            if ("PASTA" in up or "KROKAN" in up or "ORMAN" in up or "GANAJ" in up or "ANANAS" in up or "FISTIK" in up):
                all_pasta_rows.append(r)
                if debug:
                    print(f"[DEBUG] Found potential pasta row at {r}: '{v}'")

        if all_pasta_rows:
            # En üstteki pasta satırı
            first_pasta_row = min(all_pasta_rows)
            # Bu satırdan önceki 2 satırı kontrol et - MONO/KÜÇÜK/BÜYÜK başlıkları burada olmalı
            for r in range(max(1, first_pasta_row - 2), first_pasta_row + 1):
                for c in range(min_c, max_c + 1):
                    v = safe_cell_value(ws, r, c)
                    if not v:
                        continue
                    up = normalize_text(v)
                    if debug:
                        print(f"[DEBUG] Checking potential pasta header cell r={r} c={c}: '{v}' -> '{up}'")
                    # Daha esnek başlık eşleştirme
                    if "MONO" in up or "TEK" in up or "36" in up:
                        pasta_cols["MONO"] = c
                        if debug:
                            print(f"[DEBUG] Found MONO pasta column at r={r} c={c}")
                    elif ("KUCUK" in up or "KÜÇÜK" in up):
                        pasta_cols["KUCUK"] = c
                        if debug:
                            print(f"[DEBUG] Found KUCUK pasta column at r={r} c={c}")
                    elif ("BUYUK" in up or "BÜYÜK" in up):
                        pasta_cols["BUYUK"] = c
                        if debug:
                            print(f"[DEBUG] Found BUYUK pasta column at r={r} c={c}")
            # Pasta kolonları için özel kontrol: header row ve önceki 3 satır
            pasta_cols = {"MONO": None, "KUCUK": None, "BUYUK": None}
            check_rows = list(range(max(1, header_row - 3), header_row + 1))
            if debug:
                print(f"[DEBUG] Checking rows {check_rows} for pasta columns in cols {min_c}-{max_c}")
            for rr in check_rows:
                for c in range(min_c, max_c + 1):
                    v = safe_cell_value(ws, rr, c)
                    if not v:
                        continue
                    up = normalize_text(v)
                    if debug:
                        print(f"[DEBUG] Checking cell r={rr} c={c}: '{v}' -> '{up}'")
                    if ("MONO" in up or "TEK" in up) and "PASTA" in up:
                        pasta_cols["MONO"] = c
                        if debug:
                            print(f"[DEBUG] Found MONO pasta column at r={rr} c={c}")
                    elif ("KUCUK" in up or "KÜÇÜK" in up) and "PASTA" in up:
                        pasta_cols["KUCUK"] = c
                        if debug:
                            print(f"[DEBUG] Found KUCUK pasta column at r={rr} c={c}")
                    elif ("BUYUK" in up or "BÜYÜK" in up) and "PASTA" in up:
                        pasta_cols["BUYUK"] = c
                        if debug:
                            print(f"[DEBUG] Found BUYUK pasta column at r={rr} c={c}")
            break
    
    if header_row is None:
        # Fallback: use given row_hint area
        header_row = max(1, min(10, ws.max_row))
        if debug:
            print(f"[DEBUG] Using fallback header row {header_row}")
    
    sizes = find_size_columns(ws, min_c, max_c, row_hint=header_row)
    
    if debug:
        print(f"[DEBUG] Found size columns: {sizes}")
        print(f"[DEBUG] Found pasta columns: {pasta_cols}")

    # Combine size and pasta columns and return (always)
    all_cols = {k: v for k, v in sizes.items() if v}
    all_cols.update({k: v for k, v in pasta_cols.items() if v})
    return header_row, all_cols


def locate_makaron_block(ws: openpyxl.worksheet.worksheet.Worksheet, min_c: int, max_c: int, debug: bool = False) -> Dict[str, Tuple[int, int, str]]:
    """Locate the MAKARON header within the branch span and collect all variants beneath it.
    
    Args:
        ws: Worksheet to scan
        min_c: Branch start column
        max_c: Branch end column (expanded)
        debug: Enable debug logging
    
    Returns:
        Dict mapping normalized variant text -> (row_index, column_index, original_text).
        Original text is preserved for later writing format: "variant = qty".
    """
    res: Dict[str, Tuple[int, int, str]] = {}
    header_r = None
    header_c = None
    # find MAKARON header within branch span in first 100 rows
    for r in range(1, min(100, ws.max_row) + 1):
        for c in range(min_c, max_c + 1):  # Only scan within branch span
            v = safe_cell_value(ws, r, c)
            if not v:
                continue
            if normalize_text(v) == "MAKARON":
                header_r = r
                header_c = c
                if debug:
                    print(f"[DEBUG] Found MAKARON header at r={r} c={c} val='{v}' (within branch span {min_c}-{max_c})")
                break
        if header_r:
            break
    if not header_r:
        if debug:
            print("[DEBUG] No MAKARON header found")
        return res

    # For makaron, we want to look at the full 3x2 grid from branch start
    # The 2 columns start from the branch min_c
    variant_rows = [header_r + i for i in range(1, 4)]  # 3 rows beneath header
    variant_cols = list(range(min_c, min(max_c + 1, min_c + 4)))  # Look at up to 6 columns (wider than needed)
    
    if debug:
        print(f"[DEBUG] Will scan for variants in {len(variant_rows)} rows x {len(variant_cols)} cols grid starting at row {header_r+1} col {min_c}")
        
    # Get the full 3x2 grid beneath MAKARON
    seen = set()  # Track processed cells to avoid duplicates
    
    if debug:
        print(f"[DEBUG] Scanning for variants in rows {variant_rows} cols {variant_cols}")
    
    for r in variant_rows:
        for c in variant_cols:
            # If this cell is part of a merge, get its master
            merge = is_merged_at(ws, r, c)
            if merge:
                master_r, master_c, _, _ = merge
                if (master_r, master_c) in seen:
                    continue
                seen.add((master_r, master_c))
                cell = ws.cell(row=master_r, column=master_c)
            else:
                if (r, c) in seen:
                    continue
                seen.add((r, c))
                cell = ws.cell(row=r, column=c)
            
            if not cell.value:
                continue
            
            # Keep original text for format: "variant = qty" - clean any existing qty/unit
            orig_text = str(cell.value).strip()
            if not orig_text:
                continue
            
            up = normalize_text(orig_text)
            if not up:
                continue
            
            # Clean the original text from any accumulated quantities
            clean_orig_text = clean_text_from_quantities(orig_text)
            
            # Store row, column, and cleaned original text
            res[up] = (r, c, clean_orig_text)
            if debug:
                print(f"[DEBUG] MAKARON variant '{up}' at row={r} col={c} text='{orig_text}' cleaned='{clean_orig_text}'")
    

    if debug:
        print(f"[DEBUG] locate_makaron_block collected {len(res)} variants: {list(res.keys())}")
    return res

def find_pasta_rows(ws: openpyxl.worksheet.worksheet.Worksheet, base_col: int, start_row: int, debug: bool = False) -> Dict[str, Optional[int]]:
    """Find pasta product rows in worksheet by scanning the branch's first column (base_col).

    Scans downward from start_row+1 to locate rows that contain the pasta type labels
    (KROKAN, FISTIK, ORMAN/MEYVELI, GANAJ, ANANAS) in the given column.
    Returns a dict mapping pasta keys to the found row index (or None).
    """
    targets = {
        "KROKANLI": None,
        "FISTIKLI": None,
        "ORMAN": None,
        "GANAJ": None,
        "ANANAS": None
    }

    # scan a reasonable range below the header (e.g., next 30 rows) or until worksheet end
    max_scan = min(ws.max_row, start_row + 60)
    for r in range(start_row + 1, max_scan + 1):
        v = ws.cell(row=r, column=base_col).value
        if not v:
            continue
        up = normalize_text(v)
        if debug:
            print(f"[DEBUG] Checking row {r} col {base_col} for pasta: '{v}' -> '{up}'")

        if ("KROKANLI" in up or "KROKAN" in up) and targets["KROKANLI"] is None:
            targets["KROKANLI"] = r
            if debug:
                print(f"[DEBUG] Found KROKANLI pasta at row {r}")
        elif ("FISTIKLI" in up or "FISTIK" in up or "ANTEP" in up) and targets["FISTIKLI"] is None:
            targets["FISTIKLI"] = r
            if debug:
                print(f"[DEBUG] Found FISTIKLI pasta at row {r}")
        elif ("ORMAN" in up or ("MEYVELI" in up and "ORMAN" in up)) and targets["ORMAN"] is None:
            targets["ORMAN"] = r
            if debug:
                print(f"[DEBUG] Found ORMAN MEYVELI pasta at row {r}")
        elif ("GANAJ" in up or "GANAJLI" in up) and targets["GANAJ"] is None:
            targets["GANAJ"] = r
            if debug:
                print(f"[DEBUG] Found GANAJ pasta at row {r}")
        elif ("ANANAS" in up or "ANANASLI" in up) and targets["ANANAS"] is None:
            targets["ANANAS"] = r
            if debug:
                print(f"[DEBUG] Found ANANAS pasta at row {r}")

    if debug:
        print(f"[DEBUG] Found pasta rows: {targets}")
        for k, v in targets.items():
            if v is None:
                print(f"[DEBUG] WARNING: Could not find row for {k} pasta")

    return targets

def pasta_key_from_name(name_up: str) -> str:
    tests = [
        ("KROKANLI", ["KROKAN"]),
        ("FISTIKLI", ["FISTIK"]),
        ("ORMAN", ["ORMAN", "MEYVELI"]),
        ("GANAJ", ["GANAJ"]),
        ("ANANAS", ["ANANAS"])
    ]
    for key, needles in tests:
        if any(n in name_up for n in needles):
            return key
    return ""

# ----------------------- DONUK: GENERIC MATRIX BLOCKS -----------------------

def find_group_header_rows(ws: openpyxl.worksheet.worksheet.Worksheet, groups: Iterable[str]) -> Dict[str, int]:
    res: Dict[str, int] = {}
    wanted = {normalize_text(g): g for g in groups}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        up = normalize_text(v)
        if up in wanted and wanted[up] not in res:
            res[wanted[up]] = r
    return res


def scan_variant_columns(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, min_c: int, max_c: int) -> Tuple[Dict[str, int], int]:
    """Scan for variant headers on the given header row or the immediate next row.
    Returns (variants_map, row_used).
    """
    def scan_on_row(row_idx: int) -> Dict[str, int]:
        variants: Dict[str, int] = {}
        c_start = max(1, min_c)
        c_end = min(ws.max_column, max(max_c, min_c) + 12)
        for c in range(c_start, c_end + 1):
            v = ws.cell(row=row_idx, column=c).value
            if not v:
                continue
            up = normalize_text(v)
            if not up or any(k in up for k in ["3,5", "350", "150", "KG", "GR", "DONDURMALAR"]):
                # Skip size-like headers or section labels
                continue
            # Skip pure numbers or tokens without letters (e.g., '2')
            if not re.search(r"[A-Z]", up):
                continue
            if len(up) <= 2:
                continue
            # rightmost if merged
            def rightmost(rr: int, cc: int) -> int:
                for mr in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = mr.bounds
                    if min_row <= rr <= max_row and min_col <= cc <= max_col:
                        return max_col
                return cc
            col = rightmost(row_idx, c)
            variants[up] = col
        return variants

    # try header row first, then header_row+1, then header_row+2
    v_now = scan_on_row(header_row)
    if v_now:
        return v_now, header_row
    if header_row + 1 <= ws.max_row:
        v_next = scan_on_row(header_row + 1)
        if v_next:
            return v_next, header_row + 1
    if header_row + 2 <= ws.max_row:
        v_next2 = scan_on_row(header_row + 2)
        if v_next2:
            return v_next2, header_row + 2
    return {}, header_row


def scan_product_rows(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, stop_row: int) -> Dict[int, str]:
    rows: Dict[int, str] = {}
    for r in range(start_row, min(ws.max_row, stop_row)):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            # allow anonymous block row
            continue
        up = normalize_text(v)
        # Skip section headers (another group name)
        if up in {"DONDURMALAR", "TOST", "EKMEK", "CHEESECAKE", "CATAL BOREK", "ÇATAL BÖREK"}:
            break
        rows[r] = up
        # Heuristic: stop collecting after hitting an empty line following content
    if not rows:
        # Provide one anonymous row to write into if there is no explicit product row
        rows[start_row] = ""
    return rows


def build_blocks(ws: openpyxl.worksheet.worksheet.Worksheet, min_c: int, max_c: int) -> list:
    groups = ["TOST", "EKMEK", "CHEESECAKE", "ÇATAL BÖREK", "CATAL BOREK"]
    hdrs = find_group_header_rows(ws, groups)
    # Determine the next header row to bound product rows
    hdr_positions = sorted(hdrs.values()) + [ws.max_row + 1]
    blocks = []
    for gname, r in hdrs.items():
        # find next header below r
        next_r = None
        for p in hdr_positions:
            if p > r:
                next_r = p
                break
        variants, used_row = scan_variant_columns(ws, r, min_c, max_c)
        # start products below the variant header row (used_row); default r+1
        start_products = (used_row + 1) if variants else (r + 1)
        prows = scan_product_rows(ws, start_products, next_r if next_r is not None else ws.max_row + 1)
        if variants:
            blocks.append({
                "group": gname,
                "header_row": r,
                "variants": variants,  # up -> col
                "rows": prows,         # row_index -> up_name (may be empty)
                "write_row": start_products,
            })
    return blocks


def is_specific_group_product(name_up: str, group_up: str = "") -> bool:
    """Check if a product belongs to a specific group (TOST, EKMEK, etc.) that should be handled separately.
    
    Args:
        name_up: Normalized product name
        group_up: Normalized group name from CSV (optional)
    
    Returns:
        bool: True if product belongs to a specific other group
    """
    specific_groups = {
        "MUTFAK": ["SOSLU TAVUK", "KOFTE", "KÖFTE", "HAMBURGER"],
        "TOST": ["TOST"],
        "EKMEK": ["EKMEK"],
        "CHEESECAKE": ["CHEESECAKE", "TRILECE", "TRILEÇE"],
        "BOREK": ["BÖREK", "BOREK", "CATAL", "ÇATAL", "KOL BOREGI", "SU BOREGI"],
        "TATLI": ["BAKLAVA", "TATLI", "KÜNEFE", "KUNEFE"],
        "PASTA": ["PASTA"],
        "DONDURMA": ["DONDURMA", "ROKOKO"]
        }
    
        # If product belongs to a specific group by CSV group column
    if group_up in specific_groups:
        return True
    
    # Check product name against specific group keywords
    for group, keywords in specific_groups.items():
        if any(kw in name_up for kw in keywords):
            # Exception: if it's explicitly a DONUK/frozen product
            if "DONUK" in name_up or "CITIR MANTI" in name_up or "ÇITIR MANTI" in name_up or "BOYOZ" in name_up:
                return False
            return True
    return False

def route_group_for_name(name_up: str) -> Optional[str]:
    """Route CSV row name to a target block group by keywords."""
    if any(k in name_up for k in ["CHEESE", "CHEESECAKE", "TRILECE", "TRILECE"]):
        return "CHEESECAKE"
    if any(k in name_up for k in ["BOREK", "BÖREK", "CATAL", "ÇATAL", "KOL BOREGI", "SU BOREGI", "ISPANAKLI", "PATATESLI", "KIYMALI"]):
        return "CATAL BOREK"
    if "EKMEK" in name_up:
        return "EKMEK"
    if "TOST" in name_up:
        return "TOST"
    return None


def match_block_entry(name_up: str, blocks: list, desired_group: Optional[str] = None) -> Optional[Tuple[dict, int, int]]:
    # Try to find best (block, product_row, variant_col)
    best = None
    best_score = 0
    for b in blocks:
        # Restrict to routed desired group if provided
        if desired_group and normalize_text(b.get("group", "")) != normalize_text(desired_group):
            continue
        # match variant
        for v_up, col in b["variants"].items():
            if not v_up:
                continue
            # tokenized variant matching: any word length>=3 in csv name
            tokens = [t for t in re.split(r"\s+", v_up) if len(t) >= 3]
            if any(t in name_up for t in tokens):
                # match product row if any
                candidate_rows = list(b["rows"].items())
                chosen_row = None
                chosen_row_score = 0
                for r_idx, r_name in candidate_rows:
                    if not r_name:
                        # anonymous row baseline
                        if chosen_row is None:
                            chosen_row = r_idx
                            chosen_row_score = 1
                        continue
                    if r_name and r_name in name_up:
                        # prefer exact product mention
                        if 10 > chosen_row_score:
                            chosen_row = r_idx
                            chosen_row_score = 10
                score = 100 + chosen_row_score + len(v_up)
                if score > best_score and chosen_row is not None:
                    best_score = score
                    best = (b, chosen_row, col)
    return best




def find_dondurma_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Optional[int]]:
    targets = {"SUTLU": None, "KAKAOLU": None, "ANTEP": None, "KROKAN": None, "KARADUT": None,
               "LIMON": None, "DAMLA": None, "CILEK": None, "LIGHT": None, "BLUE": None,
               "CARK": None, "DOSIDO": None}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        up = normalize_text(v)
        if "SUTLU" in up and targets["SUTLU"] is None:
            targets["SUTLU"] = r
        elif "KAKAOLU" in up and targets["KAKAOLU"] is None:
            targets["KAKAOLU"] = r
        elif ("ANTEP" in up or "FISTIK" in up) and targets["ANTEP"] is None:
            targets["ANTEP"] = r
        elif "KROKAN" in up and targets["KROKAN"] is None:
            targets["KROKAN"] = r
        elif "KARADUT" in up and targets["KARADUT"] is None:
            targets["KARADUT"] = r
        elif "LIMON" in up and targets["LIMON"] is None:
            targets["LIMON"] = r
        elif ("DAMLA" in up or "SAKIZ" in up) and targets["DAMLA"] is None:
            targets["DAMLA"] = r
        elif "CILEK" in up and targets["CILEK"] is None:
            targets["CILEK"] = r
        elif ("LIGHT" in up) and targets["LIGHT"] is None:
            targets["LIGHT"] = r
        elif ("BLUE" in up or "SKY" in up) and targets["BLUE"] is None:
            targets["BLUE"] = r
        elif ("CARK" in up or "CARKIFELEK" in up) and targets["CARK"] is None:
            targets["CARK"] = r
        elif ("DOSIDO" in up or "DOSİDO" in up) and targets["DOSIDO"] is None:
            targets["DOSIDO"] = r
    return targets


def flavor_key_from_name(name_up: str) -> str:
    # Precedence matters: BLUE SKY overrides SADE; SADE maps to SUTLU (plain) unless LIGHT explicitly present
    if ("BLUE" in name_up or "SKY" in name_up):
        return "BLUE"
    if "LIGHT" in name_up:
        return "LIGHT"
    if "SADE" in name_up:
        return "SUTLU"
    tests = [
        ("SUTLU", ["SUTLU"]),
        ("KAKAOLU", ["KAKAOLU"]),
        ("ANTEP", ["ANTEP", "FISTIK"]),
        ("KROKAN", ["KROKAN"]),
        ("KARADUT", ["KARADUT"]),
        ("LIMON", ["LIMON"]),
        ("DAMLA", ["DAMLA", "SAKIZ"]),
        ("CILEK", ["CILEK"]),
        ("CARK", ["CARK", "CARKIFELEK"]),
        ("DOSIDO", ["DOSIDO", "DOSİDO"]),
    ]
    for key, needles in tests:
        if any(n in name_up for n in needles):
            return key
    return ""

# ----------------------- DONUK: DONDURMALAR ONLY -----------------------

def map_special_csv_names(name_up: str, debug: bool = False) -> str:
    """Special mapping for product name matches (supports substring matching).
    
    Args:
        name_up: Normalized CSV product name (may include quantity, codes, etc.)
        debug: Enable debug output
    Returns:
        String with mapped product name if match found, otherwise original name
    """
    # Define CSV pattern -> Excel name mappings (order matters - longer patterns first!)
    special_mappings = [
        # Baklava mappings - MUST check longer pattern first
        ("CEVIZLI TAHINLI SOGUK BAKLAVA", "CEVIZLI TAHINLI BAKLAVA"),
        # Don't map plain SOGUK BAKLAVA - it should stay as is
        
        # Tost bread mappings
        ("TOST EKMEGI TAM BUGDAY EKSI MAYALI", "EKSI MAYALI TOST EKMEGI"),
        ("ZERDECALLI EKMEK EKSI MAYALI", "ZERDECALLI TOST EKMEGI"),
    ]

    # Check for substring matches (order matters!)
    for csv_pattern, excel_name in special_mappings:
        if csv_pattern in name_up:
            # Replace the pattern with the Excel name
            result = name_up.replace(csv_pattern, excel_name)
            if debug:
                print(f"[SPECIAL MAPPING] {name_up} -> {result}")
            return result

    return name_up

def match_donuk_product(name_up: str, donuk_map: Dict[str, Tuple[int, int, str]], debug: bool = False) -> Optional[Tuple[str, int]]:
    """Match products with strict separation for MANTI, BAKLAVA variants and other critical products.
    
    Args:
        name_up: Normalized CSV product name
        donuk_map: Map of donuk products from Excel
        debug: Enable debug output
    Returns:
        Tuple of (matched_key, score) if match found, None otherwise
    """
    clean_up = re.sub(r"[\(\{\}\)]", "", name_up).strip()

    # Apply special name mapping first
    mapped_name = map_special_csv_names(clean_up, debug=debug)
    
    if debug:
        if mapped_name != clean_up:
            print(f"[MATCH_DONUK] After special mapping: {clean_up} => {mapped_name}")
        else:
            print(f"[MATCH_DONUK] Processing: {mapped_name}")

    # 1) Try exact match first (after special mapping)
    for excel_key in donuk_map.keys():
        excel_clean = re.sub(r"[\(\{\}\)]", "", excel_key).strip()
        
        # Exact matches for all products
        if mapped_name == excel_clean:
            if debug:
                print(f"  => EXACT match => {excel_key}")
            return (excel_key, 100)
            
        # Special handling for MANTI to keep variants separate
        if "MANTI" in mapped_name or "MANTI" in excel_clean:
            # Only match CITIR MANTI with CITIR MANTI
            if "CITIR" in mapped_name and "CITIR" in excel_clean:
                if mapped_name == excel_clean:
                    if debug:
                        print(f"  => EXACT match (CITIR MANTI) => {excel_key}")
                    return (excel_key, 100)
            # Only match plain MANTI with plain MANTI
            elif mapped_name == "MANTI" and excel_clean == "MANTI":
                if debug:
                    print(f"  => EXACT match (MANTI) => {excel_key}")
                return (excel_key, 100)
            # Skip other partial matches for MANTI
            continue
            
        # Special handling for BAKLAVA to keep variants separate
        if "BAKLAVA" in mapped_name or "BAKLAVA" in excel_clean:
            # Only exact matches for BAKLAVA variants
            if mapped_name == excel_clean:
                if debug:
                    print(f"  => EXACT match (BAKLAVA variant) => {excel_key}")
                return (excel_key, 100)
            # Skip fuzzy matches for BAKLAVA
            continue

    # 2) Try fuzzy matching (but skip MANTI and BAKLAVA variants - only exact match allowed)
    best_key = None
    best_score = 0
    
    for excel_key, (row, col, label) in donuk_map.items():
        excel_clean = re.sub(r"[\(\{\}\)]", "", excel_key).strip()

        # Skip fuzzy matching for MANTI and BAKLAVA variants - only exact match allowed
        if "MANTI" in mapped_name or "MANTI" in excel_clean:
            continue
        if "BAKLAVA" in mapped_name or "BAKLAVA" in excel_clean:
            continue

        # Simple substring matching for other products
        if mapped_name in excel_clean or excel_clean in mapped_name:
            if len(excel_clean) >= 5:
                score = 90
                if score > best_score:
                    best_score = score
                    best_key = excel_key

    # 3) Return best if found
    if best_key and best_score >= 80:
        if debug:
            print(f"  => SUBSTRING match => {best_key} (score={best_score})")
        return (best_key, best_score)
    else:
        if debug:
            if best_key:
                print(f"  => NO MATCH (best was '{best_key}' with score {best_score})")
            else:
                print(f"  => NO MATCH found at all")
        return None

def process_donuk_csv(csv_path: str, output_path: str = "sevkiyat_donuk.xlsx", sheet_name: Optional[str] = None, debug: bool = False, force_donuk: Optional[Iterable[str]] = None):
    df = read_csv(csv_path)
    stok_col = find_col(df, ["STOK KODU", "STOKKODU", "KOD"])
    miktar_col = find_col(df, ["MIKTAR", "MİKTAR", "ADET"])
    grup_col = find_col(df, ["GRUP", "KATEGORI", "KATEGORI ADI"])
    if not stok_col or not miktar_col:
        raise ValueError("CSV'de 'Stok Kodu' veya 'Miktar' sütunu bulunamadı.")

    # Extract branch name from CSV with priority: inner (primary) then outer (fallback)
    branch_primary_raw, branch_fallback_raw = read_branch_from_file(csv_path)
    
    # Apply branch name mapping (e.g., FORUMAVM → FORUM, HARMANDALI → EFESUS)
    from shipment_oop import BranchDecisionEngine
    branch_primary = BranchDecisionEngine._apply_branch_mapping(branch_primary_raw) if branch_primary_raw else None
    branch_fallback = BranchDecisionEngine._apply_branch_mapping(branch_fallback_raw) if branch_fallback_raw else None
    
    branch_guess = branch_primary or branch_fallback  # For display/logging
    branch_name = branch_guess  # Use for text updates
    if debug:
        print(f"[DEBUG] CSV Branch (after mapping) - Primary: '{branch_primary}' (raw: '{branch_primary_raw}'), Fallback: '{branch_fallback}' (raw: '{branch_fallback_raw}')")

    # Prepare force-donuk set (normalized) for trial runs
    force_set = set()
    if force_donuk:
        for item in force_donuk:
            if item:
                force_set.add(normalize_text(item))
    forced_hits = []  # collect entries that were forced to donuk handling for reporting

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()

    # Prepare force-donuk set (normalized) for trial runs
    force_set = set()
    if force_donuk:
        for item in force_donuk:
            if item:
                force_set.add(normalize_text(item))
    forced_hits = []  # collect entries that were forced to donuk handling for reporting

    # Select target worksheet with priority: try primary branch first, then fallback
    ws = None
    span = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Try primary branch first (inner part from parens)
        if branch_primary:
            for w in wb.worksheets:
                sp = find_branch_span(w, branch_primary)
                if sp:
                    ws = w
                    span = sp
                    if debug:
                        print(f"[DEBUG] Matched PRIMARY branch '{branch_primary}' in sheet '{w.title}'")
                    break
        
        # If primary failed, try fallback (outer part from parens)
        if ws is None and branch_fallback:
            for w in wb.worksheets:
                sp = find_branch_span(w, branch_fallback)
                if sp:
                    ws = w
                    span = sp
                    if debug:
                        print(f"[DEBUG] Matched FALLBACK branch '{branch_fallback}' in sheet '{w.title}'")
                    break
        
        # Last resort: use first sheet
        if ws is None:
            ws = wb.worksheets[0]
    
    # Get span if not already found
    if span is None and branch_guess:
        span = find_branch_span(ws, branch_guess)
    if span:
        min_c, max_c, branch_row = span
        
        # Find next branch column to avoid crossing boundaries
        next_branch_col = None
        for c in range(max_c + 1, min(ws.max_column + 1, max_c + 15)):
            val = ws.cell(row=branch_row, column=c).value
            if val and str(val).strip():
                # Check if it's a branch name (not a date/time header)
                if not any(x in str(val).upper() for x in ["TARIH", "SIPARIS", "TESLIM"]):
                    next_branch_col = c
                    break
        
        # Expand search area but respect next branch boundary
        if next_branch_col:
            # Next branch exists - stop before it (leave 1 col gap)
            max_c = min(max_c + 4, next_branch_col - 1)
        else:
            # Last branch - allow moderate expansion (max 8 cols)
            max_c = min(max_c + 8, min_c + 12, ws.max_column)
        
        if debug:
            print(f"[DEBUG] Expanded max_c to {max_c} (next_branch_col: {next_branch_col})")
    else:
        min_c, max_c, branch_row = 2, 14, 2  # Default olarak daha geniş bir aralık
    
    # Sheet-based override: in Kayseri-Sivas and Adana sheets, use KOLI for all non-dondurma groups
    force_koli_all = False
    try:
        ws_title_norm = normalize_text(ws.title if ws else "")
        if ("KAYSERI" in ws_title_norm and "SIVAS" in ws_title_norm) or ("ADANA" in ws_title_norm):
            force_koli_all = True
    except Exception:
        force_koli_all = False
    # Discover the actual DONDURMALAR header row and size columns near this branch span
    header_row, size_cols = locate_dondurmalar_block(ws, min_c, max_c, debug)

    # Determine pasta columns relative to the branch span (user provided mapping):
    # Mono: branch_col + 1, Küçük: branch_col + 2, Büyük: branch_col + 3
    pasta_cols = {
        "MONO": (min_c + 1) if (min_c + 1) <= ws.max_column else None,
        "KUCUK": (min_c + 2) if (min_c + 2) <= ws.max_column else None,
        "BUYUK": (min_c + 3) if (min_c + 3) <= ws.max_column else None,
    }
    if debug:
        print(f"[DEBUG] Pasta columns assigned from branch span: {pasta_cols} (min_c={min_c})")

    # Find pasta rows by scanning the first column of the branch (min_c)
    pasta_rows = find_pasta_rows(ws, min_c, header_row, debug)

    # Locate MAKARON block (3 rows x 2 columns) and build variant -> (row, col, text) map
    makaron_map = locate_makaron_block(ws, min_c, max_c, debug=debug)
    if debug:
        print(f"[DEBUG] Makaron map: {makaron_map}")

    # Locate DONUK products and build product -> (row, col, text) map
    donuk_map = locate_donuk_products_block(ws, min_c, max_c, branch_name=branch_name, debug=debug)
    if debug:
        print(f"[DEBUG] Donuk products map: {donuk_map}")

    # --- SIMPLE SPECIAL-ITEM PASS -------------------------------------------------
    # For the user's requested list we apply a very simple deterministic match:
    # - If CSV product normalized contains any of the canonical tokens below,
    #   we try to find an excel target by a simple substring search in donuk_map keys
    #   or elsewhere in the branch span. If found, we write the qty immediately
    #   (append to the found cell) and mark the CSV row as processed so later
    #   scoring/aggregation doesn't touch it.
    # Order matters: longer/more specific items first to avoid substring confusion
    special_items = [
        "CEVİZLİ TAHİNLİ BAKLAVA",  # Must come before "SOĞUK BAKLAVA"
        "SOĞUK BAKLAVA",
        "HAMBURGER EKMEĞİ",  # Must come before "HAMBURGER KÖFTE"
        "HAMBURGER KÖFTE",
        "ÇITIR MANTI",  # Must come before plain "MANTI"
        "EKŞİ MAYALI TOST EKMEĞİ",
        "ZERDEÇALLI TOST EKMEĞİ",
        "ACI-TATLI SOSLU TAVUK",
        "USTANIN KÖFTESİ",
        "KADAYIFLI ŞİNİTSEL",
        "MADALYON BONFİLE",
        "SPAGETTİ ET",
        "ÇITIR TAVUK",
        "TAVUK BUT",
        "DANA ASADO",
        "MANTI",  # Plain MANTI (separate from CITIR MANTI)
        "BOYOZ",
        "PATATES"
    ]
    # normalize tokens
    special_norm = [normalize_text(s) for s in special_items]

    # We'll collect a simple report of which CSV rows we wrote from this pass
    simple_pass_hits = []

    # Ensure processed_products exists (may be declared later in this function)
    try:
        processed_products
    except NameError:
        processed_products = set()

    # Simple pass: iterate CSV and handle rows matching any special token
    for idx, row in df.iterrows():
        try:
            name_raw = str(row[stok_col])
            qty_raw = row[miktar_col]
        except Exception:
            continue
        up = normalize_text(name_raw)
        clean_up = re.sub(r"[\(\{\}\)]", "", up).strip()

        # skip already-processed (from earlier runs)
        if clean_up in processed_products:
            continue

        # Apply special mappings FIRST to handle products like "CEVIZLI TAHINLI SOGUK BAKLAVA"
        # This maps it to "CEVIZLI TAHINLI BAKLAVA"
        mapped_clean = map_special_csv_names(clean_up, debug=False)
        
        # Check for exact matches first, then substring matches
        # This prevents "SOGUK BAKLAVA" from matching "CEVIZLI TAHINLI BAKLAVA"
        matched_token = None
        
        # First try exact match
        for tn in special_norm:
            if tn and mapped_clean == tn:
                matched_token = tn
                break
        
        # If no exact match, try substring match (but only for items not already processed)
        if not matched_token:
            for tn in special_norm:
                if tn and tn in mapped_clean:
                    matched_token = tn
                    break
        
        if not matched_token:
            continue

        # parse qty
        try:
            qty = float(str(qty_raw).replace(",", "."))
        except Exception:
            if debug:
                print(f"[DEBUG] SIMPLE PASS: could not parse qty for '{name_raw}'")
            continue

        # Try exact match in donuk_map first (normalize keys)
        target = None
        for dk, target_val in donuk_map.items():
            if not dk:
                continue
            # Use exact match to prevent confusion (e.g., "MANTI" vs "CITIR MANTI")
            if matched_token == normalize_text(dk):
                target = target_val
                if debug:
                    print(f"[DEBUG] SIMPLE PASS: matched token '{matched_token}' -> donuk_map key '{dk}'")
                break

        # If not found in donuk_map, scan the worksheet branch span for any textual cell
        if not target:
            found = None
            try:
                for rr in range(1, ws.max_row + 1):
                    for cc in range(max(1, min_c), min(ws.max_column, max_c) + 1):
                        val = ws.cell(row=rr, column=cc).value
                        if not val or not isinstance(val, str):
                            continue
                        upv = normalize_text(val)
                        # Use exact match for critical products to avoid confusion
                        if matched_token == upv:
                            found = (rr, cc, val)
                            break
                    if found:
                        break
            except Exception:
                found = None
            if found:
                target = (found[0], found[1], found[2])
                if debug:
                    print(f"[DEBUG] SIMPLE PASS: matched token '{matched_token}' -> worksheet cell r={found[0]} c={found[1]} val='{found[2]}'")

        # If we found a target, write appended text and mark processed
        if target:
            row_idx, col_idx, orig_text = target
            # Append unit for DONUK special items: default SEPET, but KAY/SIVAS & ADANA force KOLI
            try:
                fmt_qty = int(qty) if float(qty).is_integer() else qty
            except Exception:
                fmt_qty = qty
            # Force KL. for specified products regardless of sheet override
            force_always_kl_norm = {normalize_text(x) for x in [
                "HAMBURGER EKMEĞİ",
                "EKŞİ MAYALI TOST EKMEĞİ",
                "ZERDEÇALLI TOST EKMEĞİ",
            ]}
            unit_text = "KL." if (matched_token in force_always_kl_norm or force_koli_all) else "SPT."
            new_text = append_text_with_space(orig_text, f"{fmt_qty} {unit_text}")
            try:
                safe_write(ws, row_idx, col_idx, new_text)
                processed_products.add(clean_up)
                simple_pass_hits.append({"csv": name_raw, "matched": orig_text, "row": row_idx, "col": col_idx, "qty": qty, "excel_name": map_special_csv_names(clean_up)})
                if debug:
                    print(f"[DEBUG] SIMPLE PASS WRITE r={row_idx} c={col_idx} val='{new_text}' for CSV='{name_raw}' -> EXCEL='{map_special_csv_names(clean_up)}' qty={qty}")
            except Exception as e:
                if debug:
                    print(f"[DEBUG] SIMPLE PASS WRITE ERROR for '{name_raw}' -> {e}")

    if simple_pass_hits and debug:
        print("\n[SIMPLIFIED MATCH PASS] Summary of writes for provided special list:")
        for h in simple_pass_hits:
            print(f" - CSV: '{h['csv']}' | EXCEL: '{h['excel_name']}' -> wrote to r={h['row']} c={h['col']} (cell='{h['matched']}') qty={h['qty']}")
    # --- end simple pass --------------------------------------------------------

    # Ensure aggregator exists
    aggreg: Dict[Tuple[int, int], float] = {}
    matched = 0

    # Dictionary to collect DONUK product quantities
    donuk_aggreg: Dict[str, float] = {}
    processed_products = set()
    # Separate dictionary for MAKARON quantities (keep makaron logic independent from DONUK)
    makaron_aggreg: Dict[str, float] = {}

    # Helper: check whether a CSV product name has any reasonable match in the
    # Excel-side maps (donuk_map or makaron_map). If a product is classified as
    # a "specific group" we will only allow processing when the Excel template
    # actually contains a matching key.
    def product_matches_excel(name_up: str) -> bool:
        # tokenized significant words
        name_words = set(w for w in name_up.split() if len(w) > 2)

        # check donuk map
        for k in donuk_map.keys():
            if not k:
                continue
            k_clean = re.sub(r"[\(\{\}\)]", "", k).strip()
            if k_clean == name_up or name_up in k_clean or k_clean in name_up:
                return True
            k_words = set(w for w in k_clean.split() if len(w) > 2)
            if name_words & k_words:
                return True

        # check makaron map
        for k in makaron_map.keys():
            if not k:
                continue
            if k == name_up or name_up in k or k in name_up:
                return True
            k_words = set(w for w in k.split() if len(w) > 2)
            if name_words & k_words:
                return True

        # scan worksheet in the branch span for any textual cell that matches
        # This allows specific-group products to be accepted if they appear
        # elsewhere in the template (pasta rows, matrix blocks, etc.).
        try:
            for rr in range(1, ws.max_row + 1):
                for cc in range(max(1, min_c), min(ws.max_column, max_c) + 1):
                    val = ws.cell(row=rr, column=cc).value
                    if not val or not isinstance(val, str):
                        continue
                    upv = normalize_text(val)
                    if not upv:
                        continue
                    if upv == name_up or name_up in upv or upv in name_up:
                        return True
                    upv_words = set(w for w in upv.split() if len(w) > 2)
                    if name_words & upv_words:
                        return True
        except Exception:
            # If anything goes wrong scanning the sheet, fall back to False
            pass

        return False

    # Aggregate DONUK entries from CSV
    if donuk_map:
        for _, r in df.iterrows():
            try:
                name_raw = str(r[stok_col])
                group_val = normalize_text(str(r.get(grup_col, ""))) if grup_col else ""
                miktar_val = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                continue

            # Normalize name
            up = normalize_text(name_raw)
            clean_up = re.sub(r"[\(\{\}\)]", "", up).strip()

            # If this product matches any forced token, treat as DONUK candidate
            forced_flag = False
            if force_set:
                for f in force_set:
                    if f and f in clean_up:
                        forced_flag = True
                        break

            # Apply special mappings with debug info
            mapped_name = map_special_csv_names(clean_up, debug=debug)
            if mapped_name != clean_up and debug:
                print(f"\n[SPECIAL PRODUCT FOUND] Processing:")
                print(f"Original CSV name: {name_raw}")
                print(f"Normalized: {clean_up}")
                print(f"Mapped to: {mapped_name}")
                print(f"Quantity: {miktar_val}\n")

            # Skip if already processed
            if clean_up in processed_products:
                continue

            # Check if product belongs to another specific group. Only skip
            # it when the Excel template does NOT contain a matching product/key.
            if is_specific_group_product(clean_up, group_val):
                if not product_matches_excel(clean_up):
                    if debug:
                        print(f"[DEBUG] Skipping '{name_raw}' - classified as specific group '{group_val}' and no Excel match found")
                    continue
                else:
                    if debug:
                        print(f"[DEBUG] Specific-group product '{name_raw}' has an Excel match; continuing processing")

            # Only process if it's a DONUK product or in DONUK/BOREK group
            is_donuk_product = ("CITIR MANTI" in clean_up or "ÇITIR MANTI" in clean_up or 
                              "BOYOZ" in clean_up)
            is_donuk_group = group_val in ("DONUK", "BOREK")

            # Force to DONUK if present in force_set
            if forced_flag:
                is_donuk_product = True
                if debug:
                    print(f"[DEBUG] FORCED DONUK candidate detected: '{name_raw}' (force tokens matched)")

            if not (is_donuk_product or is_donuk_group):
                if debug:
                    print(f"[DEBUG] Skipping '{name_raw}' - not a DONUK product/group")
                continue

            # Try to find best matching product
            best_match = None
            best_score = 0
            
            for excel_key in donuk_map.keys():
                # Compare names after removing special characters
                excel_clean = re.sub(r"[\(\{\}\)]", "", excel_key).strip()
                
                # Exact match first
                if up == excel_clean:
                    best_match = excel_key
                    break

                # Special case matches
                if ("HAMBURGER" in up and "HAMBURGER" in excel_clean) or \
                   ("KOFTE" in up and "KOFTE" in excel_clean) or \
                   ("KÖFTE" in up and "KÖFTE" in excel_clean) or \
                   ("SOS" in up and "SOS" in excel_clean) or \
                   ("TAVUK" in up and "TAVUK" in excel_clean) or \
                   ("MANTI" in up and "MANTI" in excel_clean) or \
                   ("BOYOZ" in up and "BOYOZ" in excel_clean):
                    score = 100 + len(set(up.split()) & set(excel_clean.split()))
                    if score > best_score:
                        best_match = excel_key
                        best_score = score
                        continue

                # Partial match - look for shared significant words
                name_words = set(word for word in up.split() if len(word) > 2)
                excel_words = set(word for word in excel_clean.split() if len(word) > 2)
                common_words = name_words & excel_words
                
                if common_words:
                    # Enhanced scoring system
                    base_score = len(common_words) * 10
                    length_score = sum(len(word) for word in common_words)
                    position_score = 5 if any(up.startswith(word) and excel_clean.startswith(word) for word in common_words) else 0
                    
                    # Special bonus for key product words
                    key_word_bonus = 0
                    key_words = ["CITIR", "ÇITIR", "DONUK", "MANTI", "BOYOZ", "SOS", "TAVUK", "KOFTE", "KÖFTE"]
                    for word in common_words:
                        if any(key in word for key in key_words):
                            key_word_bonus += 15
                    
                    total_score = base_score + length_score + position_score + key_word_bonus
                    
                    if total_score > best_score:
                        best_match = excel_key
                        best_score = total_score

            # If no best_match found but this row was forced, try a looser match using match_donuk_product
            if not best_match and forced_flag:
                mm = match_donuk_product(clean_up, donuk_map, debug=debug)
                if mm:
                    best_match = mm[0]
                    if debug:
                        print(f"[DEBUG] FORCED MATCH: found excel_key='{best_match}' via match_donuk_product for '{name_raw}'")

            if best_match:
                try:
                    qty = float(str(r[miktar_col]).replace(",", "."))
                except Exception:
                    if debug:
                        print(f"[DEBUG] DONUK WARNING: could not parse qty for '{name_raw}'")
                    continue

                if "MAKARON" in up:
                    # Extract variant from parentheses or main text
                    m = re.search(r"\(([^)]+)\)", name_raw)
                    variant = normalize_text(m.group(1) if m else name_raw)
                    # Normalize variant names
                    variant_map = {
                        "FRAMBUAZ": "FRAMBUAZLI",
                        "HINDCEVIZ": "HINDCEVIZLI",
                        "ANTEP": "ANTEPLI",
                        "KARAMEL": "KARAMEL",
                        "MERSIN": "MERSINLI"
                    }
                    # Find best matching variant key
                    best_variant = None
                    for k in variant_map:
                        if k in variant:
                            best_variant = variant_map[k]
                            break
                    if best_variant:
                        # Only aggregate macaron quantities into the dedicated makaron_aggreg
                        # and only if the Excel template actually contains that variant.
                        if best_variant in makaron_map:
                            makaron_aggreg[best_variant] = makaron_aggreg.get(best_variant, 0.0) + qty
                        else:
                            if debug:
                                print(f"[DEBUG] MAKARON WARNING: Excel has no cell for variant '{best_variant}' (source='{name_raw}')")
                            # don't treat makaron as donuk; skip
                            continue
                elif "HAMBURGER" in clean_up and "KOFTE" in clean_up and "MUTFAK" == group_val:
                    donuk_key = "HAMBURGER KÖFTE"
                    donuk_aggreg[donuk_key] = donuk_aggreg.get(donuk_key, 0.0) + qty
                elif "SOSLU" in clean_up and "TAVUK" in clean_up and "MUTFAK" == group_val:
                    donuk_key = "ACI-TATLI SOSLU TAVUK"
                    donuk_aggreg[donuk_key] = donuk_aggreg.get(donuk_key, 0.0) + qty
                elif best_match:
                    donuk_aggreg[best_match] = donuk_aggreg.get(best_match, 0.0) + qty
                matched += 1
                if debug:
                    print(f"[DEBUG] DONUK + {name_raw} -> product='{best_match}' qty={qty}")
                # record forced hit for reporting (if this row was in force set)
                if forced_flag:
                    forced_hits.append({
                        "csv_name": name_raw,
                        "normalized": clean_up,
                        "matched_key": best_match,
                        "qty": qty,
                    })

        # After collecting quantities, update cells with space-separated format
        for product_key, total_qty in donuk_aggreg.items():
            target = None
            
            # First try donuk_map
            target = donuk_map.get(product_key)
            
            # Then try makaron_map if not found
            if not target:
                target = makaron_map.get(product_key)
                
            if not target and debug:
                print(f"[DEBUG] DONUK WARNING: no excel cell found for collected product '{product_key}'")
                continue

            row_idx, col_idx, orig_text = target
            fmt_qty = int(total_qty) if float(total_qty).is_integer() else total_qty

            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                # DONUK text cells: append qty and unit (SEPET by default)
                unit_text = "KL." if force_koli_all else "SPT."
                new_text = append_text_with_space(orig_text, f"{fmt_qty} {unit_text}")
                
                # Handle merged cells (use safe_write to reliably handle merges)
                try:
                    safe_write(ws, row_idx, col_idx, new_text)
                except Exception:
                    # best-effort: fall back to direct cell assignment
                    try:
                        ws.cell(row=row_idx, column=col_idx).value = new_text
                    except Exception:
                        pass

                if debug:
                    print(f"[DEBUG] DONUK TEXT WRITE r={row_idx} c={col_idx} val='{new_text}'")

            except Exception as e:
                if debug:
                    print(f"[DEBUG] DONUK WRITE ERROR r={row_idx} c={col_idx} error={e}")

    # Aggregate MAKARON entries from CSV: CSV usually has variant in parentheses (Makaron (ÇİKOLATALI))
    if makaron_map:
        for _, r in df.iterrows():
            try:
                name_raw = str(r[stok_col])
            except Exception:
                continue
            up = normalize_text(name_raw)
            grp_val = normalize_text(str(r.get(grup_col, ""))) if grup_col else ""
            # skip rows already processed in the donuk pass
            clean_up = re.sub(r"[\(\{\}\)]", "", up).strip()
            if clean_up in processed_products:
                continue
            # detect makaron rows: name or group mentions MAKARON
            if "MAKARON" not in up and "MAKARON" not in grp_val:
                continue
            
            # Prepare candidate key from CSV name/group
            csv_variant = None
            # First try extracting from parentheses
            m = re.search(r"\(([^)]+)\)", name_raw)

            
            # Normalize FISTIK spelling in parentheses
            if m:
                csv_variant = normalize_text(m.group(1))
                
                # Normalize FISTIK spelling (check for None first)
                if csv_variant and ("FISTIK" in csv_variant or "FISTIKLI" in csv_variant):
                    csv_variant = csv_variant.replace("FISTIKLI", "ANTEPLI").replace("FISTIK", "ANTEP")

            # Try to find best matching variant
            best_match = None
            best_score = 0
            
            for excel_key, target in makaron_map.items():
                if not excel_key:
                    continue
                    
                # Compare CSV variant (if found) with Excel key
                if csv_variant:
                    # Exact match
                    if csv_variant == excel_key:
                        best_match = excel_key
                        break
                    # One contains the other (ignoring -LI suffix)
                    csv_base = csv_variant.replace("LI", "").strip()
                    excel_base = excel_key.replace("LI", "").strip()
                    if csv_base in excel_base or excel_base in csv_base:
                        if len(excel_base) > best_score:
                            best_match = excel_key
                            best_score = len(excel_base)
                
                # If no match yet, try normalized CSV name
                if not best_match:
                    # Remove common words and suffixes
                    clean_name = up.replace("MAKARON", "").replace("LI", "").replace("LU", "").strip()
                    clean_key = excel_key.replace("LI", "").replace("LU", "").strip()
                    if clean_key in clean_name or clean_name in clean_key:
                        if len(clean_key) > best_score:
                            best_match = excel_key
                            best_score = len(clean_key)
            
            variant_key = best_match if best_match else ""
            
            if not variant_key:
                if debug:
                    print(f"[DEBUG] MAKARON WARNING: could not determine variant for '{name_raw}'")
                continue
            
            try:
                qty = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                if debug:
                    print(f"[DEBUG] DONUK WARNING: could not parse qty for '{name_raw}'")
                continue
            
            # Update macaron aggregation (separate from donuk aggregation)
            makaron_aggreg[best_match] = makaron_aggreg.get(best_match, 0.0) + qty
            processed_products.add(clean_up)
            matched += 1
            if debug:
                print(f"[DEBUG] MAKARON + {name_raw} -> variant='{best_match}' qty={qty}")

        # After collecting all quantities, format and write text cells
        for variant_key, total_qty in makaron_aggreg.items():
            target = makaron_map.get(variant_key)
            if not target:
                if debug:
                    print(f"[DEBUG] MAKARON WARNING: no excel cell found for collected variant '{variant_key}'")
                continue
            
            row_idx, col_idx, orig_text = target
            fmt_qty = int(total_qty) if float(total_qty).is_integer() else total_qty

            # Use the existing cell's alignment for right-aligned text
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                '''if not cell.alignment:
                    cell.alignment = Alignment(horizontal='right')
                '''
                # Format consistently: preserve original text and append qty with space (user preference)
                # MAKARON cells: always KOLI; override also KOLI
                unit_text = "KL."
                new_text = append_text_with_space(orig_text, f"{fmt_qty} {unit_text}")
                
                # Try master cell first if merged
                master = None
                merge = is_merged_at(ws, row_idx, col_idx)
                if merge:
                    master_r, master_c, _, _ = merge
                    master = ws.cell(row=master_r, column=master_c)
                
                target_cell = master if master else cell
                '''if target_cell.alignment:
                    target_cell.alignment = Alignment(horizontal='right')'''
                # Use safe_write so merged cells are handled consistently
                try:
                    safe_write(ws, row_idx, col_idx, new_text)
                except Exception:
                    try:
                        ws.cell(row=row_idx, column=col_idx).value = new_text
                    except Exception:
                        pass
                
                if debug:
                    print(f"[DEBUG] MAKARON TEXT WRITE r={row_idx} c={col_idx} val='{new_text}' aligned=right")
                
            except Exception as e:
                if debug:
                    print(f"[DEBUG] MAKARON WRITE ERROR r={row_idx} c={col_idx} val='{new_text}' error={e}")

    # Aggregate pasta entries from CSV into aggreg by mapping pasta type -> pasta_rows and size -> pasta_cols
    for _, r in df.iterrows():
        name = str(r[stok_col])
        up = normalize_text(name)

        pasta_key = pasta_key_from_name(up)
        if not pasta_key:
            continue

        unit_text = str(r.get("Birim", r.get("BIRIM", "")))

        if debug:
            print(f"[DEBUG] Found pasta CSV row: {name} -> pasta_key={pasta_key}")

        # choose column by size tokens in CSV name/unit
        col_idx = None
        if "MONO" in up or "TEK" in up or re.search(r"\b36\b", up):
            col_idx = pasta_cols.get("MONO")
            if debug:
                print(f"[DEBUG] MONO pasta detected, column={col_idx}")
        elif "KUCUK" in up or "KÜÇÜK" in up:
            col_idx = pasta_cols.get("KUCUK")
            if debug:
                print(f"[DEBUG] KUCUK pasta detected, column={col_idx}")
        elif "BUYUK" in up or "BÜYÜK" in up:
            col_idx = pasta_cols.get("BUYUK")
            if debug:
                print(f"[DEBUG] BUYUK pasta detected, column={col_idx}")

        if debug:
            print(f"[DEBUG] Pasta rows: {pasta_rows}")

        if not col_idx:
            if debug:
                print(f"[DEBUG] WARNING: No column match for size in: {name}")
            continue

        try:
            qty = float(str(r[miktar_col]).replace(",", "."))
        except Exception:
            continue

        row_idx = pasta_rows.get(pasta_key)
        if not row_idx:
            if debug:
                print(f"[DEBUG] WARNING: No pasta row found for key {pasta_key} (source='{name}')")
            continue

        key = (row_idx, col_idx)
        aggreg[key] = aggreg.get(key, 0.0) + qty
        matched += 1

    if debug:
        print(f"[DEBUG] Branch span min_c={min_c} max_c={max_c} row={branch_row}")
        print(f"[DEBUG] DONDURMALAR header at row={header_row}")
        print(f"[DEBUG] Size columns: {size_cols}")

    flavor_rows = find_dondurma_rows(ws)
    if debug:
        print(f"[DEBUG] Flavor rows: {flavor_rows}")

    # 'aggreg' and 'matched' may already be populated by earlier pasta aggregation.
    # Keep block_aggreg and other counters here.
    block_aggreg: Dict[Tuple[int, int], float] = {}
    rokoko_total: float = 0.0

    blocks = build_blocks(ws, min_c, max_c)
    if debug:
        print(f"[DEBUG] Blocks: {[{'group': b['group'], 'header_row': b['header_row'], 'variants': list(b['variants'].keys())} for b in blocks]}")

    # Explicit 4-subcolumn layout under this branch
    # Determine the 4 numeric subcolumns of this branch
    sub_cols = [min_c + i for i in range(4) if (min_c + i) <= ws.max_column]

    # CRITICAL FIX: Clamp size columns to the branch's 4-column layout to prevent spillover
    # Layout per branch group:
    #   [0] 3,5 KG (col 1)
    #   [1] 3,5 KG (col 2)
    #   [2] 350 GR
    #   [3] 150 GR
    # If auto-detection picked headers outside this branch (e.g., next branch), override here.
    if len(sub_cols) >= 4:
        # Ensure 35KG points inside the first two columns of this branch
        if size_cols.get("35KG") not in (sub_cols[0], sub_cols[1]):
            size_cols["35KG"] = sub_cols[0]
        # Force granular sizes to the 3rd and 4th subcolumns of THIS branch
        size_cols["350GR"] = sub_cols[2]
        size_cols["150GR"] = sub_cols[3]
    # Helper: find row index by label(s) in column A
    def find_row_by_label(keywords: Iterable[str]) -> Optional[int]:
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(row=rr, column=1).value
            if not v:
                continue
            upv = normalize_text(v)
            if all(k in upv for k in keywords):
                return rr
        return None

    row_tost = find_row_by_label(["TOST"]) if sub_cols else None
    row_ekmek = find_row_by_label(["EKMEK"]) if sub_cols else None
    row_cheese = find_row_by_label(["CHEESECAKE"]) if sub_cols else None
    row_catal = find_row_by_label(["CATAL", "BOREK"]) if sub_cols else None
    
    # Aggregators for explicit layout
    matrix_aggreg: Dict[Tuple[int, int], float] = {}
    kunefe_total_sum: float = 0.0
    trilece_total_sum: float = 0.0
    ekler_total_sum: float = 0.0
    for _, r in df.iterrows():
        name = str(r[stok_col])
        grp = str(r[grup_col]) if grup_col and grup_col in df.columns else ""
        up = normalize_text(name)
        g_up = normalize_text(grp)
        

        if "EKLER" in up:
            try:
                qty_r = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                qty_r = 0.0
            ekler_total_sum += qty_r
            matched += 1
            if debug:
                print(f"[DEBUG] EKLER +{qty_r} now {ekler_total_sum}")
            continue
        # Collect MEYVELI ROKOKO totals for a dedicated text cell update later
        if ("ROKOKO" in up and "MEYVELI" in up) or ("MEYVELI" in up and "ROKOKO" in up):
            try:
                qty_r = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                qty_r = 0.0
            rokoko_total += qty_r
            if debug:
                print(f"[DEBUG] ROKOKO +{qty_r} now {rokoko_total}")
            continue  # ROKOKO is handled as a text cell, not in the DONDURMALAR grid
        # New explicit mapping path for TATLI/BOREK items
        if "KUNEFE" in up:
            try:
                qty_r = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                qty_r = 0.0
            kunefe_total_sum += qty_r
            matched += 1
            if debug:
                print(f"[DEBUG] KUNEFE +{qty_r} now {kunefe_total_sum}")
            continue
        if "TRILECE" in up:
            try:
                qty_r = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                qty_r = 0.0
            trilece_total_sum += qty_r
            matched += 1
            if debug:
                print(f"[DEBUG] TRILECE +{qty_r} now {trilece_total_sum}")
            continue
        
        if g_up in ("TATLI", "BOREK") and sub_cols:
            # parse quantity
            try:
                qty = float(str(r[miktar_col]).replace(",", "."))
            except Exception:
                qty = None
            if qty is not None and qty != 0:
                def add(rr: Optional[int], cols: Iterable[int], val: float):
                    if not rr:
                        return
                    for c_ in cols:
                        key = (rr, c_)
                        matrix_aggreg[key] = matrix_aggreg.get(key, 0.0) + val
                        if debug:
                            print(f"[DEBUG] MATRIX + r={rr} c={c_} val={val} src='{name}'")

                # Specials (apply before groups)
                

                # TOST
                if row_tost and "TOST" in up:
                    if "KASAR" in up:
                        add(row_tost, [sub_cols[0]], qty)
                        matched += 1
                        continue
                    if "KEPEK" in up and len(sub_cols) >= 2:
                        add(row_tost, [sub_cols[1]], qty)
                        matched += 1
                        continue
                    if "KARISIK" in up:
                        cols = sub_cols[2:4] if len(sub_cols) >= 4 else sub_cols[2:3]
                        if cols:
                            add(row_tost, cols, qty)
                            matched += 1
                            continue

                # EKMEK
                if row_ekmek:
                    if ("EKMEK" in up) and ("BEYAZ" in up):
                        add(row_ekmek, [sub_cols[0]], qty)
                        matched += 1
                        continue
                    if ("EKMEK" in up) and ("ESMER" in up) and len(sub_cols) >= 2:
                        add(row_ekmek, [sub_cols[1]], qty)
                        matched += 1
                        continue
                    if ("KIYMALI" in up or "KIYMA" in up) and ("BORE" in up):
                        cols = sub_cols[2:4] if len(sub_cols) >= 4 else sub_cols[2:3]
                        if cols:
                            add(row_ekmek, cols, qty)
                            matched += 1
                            continue

                # CHEESECAKE
                if row_cheese:
                    if "SEBASTIAN" in up:
                        add(row_cheese, sub_cols[:2], qty)
                        matched += 1
                        continue
                    if "FRAMBUAZ" in up:
                        cols = sub_cols[2:4] if len(sub_cols) >= 3 else []
                        if cols:
                            add(row_cheese, cols, qty)
                            matched += 1
                            continue

                # ÇATAL BÖREK
                if row_catal:
                    if "PATATES" in up:
                        add(row_catal, [sub_cols[0]], qty)
                        matched += 1
                        continue
                    if "ISPANAK" in up and len(sub_cols) >= 2:
                        add(row_catal, [sub_cols[1]], qty)
                        matched += 1
                        continue
                    if ("SU" in up) and ("BORE" in up):
                        cols = sub_cols[2:4] if len(sub_cols) >= 3 else []
                        if cols:
                            add(row_catal, cols, qty)
                            matched += 1
                            continue
        # Else try DONDURMALAR grid
        fkey = flavor_key_from_name(up)
        if not fkey:
            continue
        row_idx = flavor_rows.get(fkey)
        if not row_idx:
            continue
        unit_text = str(r.get("Birim", r.get("BIRIM", "")))
        sz = size_from_stock_or_unit(name, unit_text)
        if not sz:
            # Special-case DOSIDO: no explicit size, write into 3,5 KG by convention
            if fkey == "DOSIDO":
                sz = "35KG"
            else:
                continue
        col_idx = size_cols.get("35KG") if sz == "35KG" else size_cols.get("350GR") if sz == "350GR" else size_cols.get("150GR")
        # If size columns couldn't be detected for this branch, skip to avoid writing into wrong areas
        if not col_idx:
            continue
        try:
            qty = float(str(r[miktar_col]).replace(",", "."))
        except Exception:
            continue
        key = (row_idx, col_idx)
        aggreg[key] = aggreg.get(key, 0.0) + qty
        matched += 1
        if debug:
            print(f"[DEBUG] + {name} -> fkey={fkey} size={sz} row={row_idx} col={col_idx} qty={qty}")

    rows_to_clear = [r for r in flavor_rows.values() if r and r > header_row]
    cols_to_clear_raw = [c for c in [size_cols.get("35KG"), size_cols.get("350GR"), size_cols.get("150GR")] if c]
    for rr in rows_to_clear:
        for cc_raw in cols_to_clear_raw:
            cc = resolve_numeric_col(ws, rr, cc_raw, min_c, max_c)
            cell = ws.cell(row=rr, column=cc)
            val = cell.value
            if isinstance(val, (int, float)):
                cell.value = None
            # Clean text values that contain qty/unit patterns
            elif isinstance(val, str) and re.search(r"\d+\s*(?:SPT\.|KL\.|TEPSI|TEPSİ)", val, re.IGNORECASE):
                cleaned = clean_text_from_quantities(val)
                if cleaned:
                    cell.value = cleaned
                else:
                    cell.value = None
            else:
                try:
                    fv = float(str(val).replace(",", ".")) if (val not in (None, "") and str(val).strip() != "") else None
                    if fv is not None:
                        cell.value = None
                except Exception:
                    pass

    # Clear numeric cells in Dondurmalar block
    for (r_, c_), v in aggreg.items():
        cc = resolve_numeric_col(ws, r_, c_, min_c, max_c)
        # If this cell belongs to PASTA area, write as text with unit instead of numeric
        try:
            pasta_row_set = set([rv for rv in (pasta_rows.values() if 'pasta_rows' in locals() else []) if rv])
            pasta_col_set = set([cv for cv in ((pasta_cols.values() if 'pasta_cols' in locals() else [])) if cv])
        except Exception:
            pasta_row_set, pasta_col_set = set(), set()

        is_pasta_cell = (r_ in pasta_row_set) and (cc in pasta_col_set)

        try:
            if is_pasta_cell:
                fmt_v = int(v) if float(v).is_integer() else v
                unit_text = "KL." if force_koli_all else "SPT."
                out_text = f"{fmt_v} {unit_text}"
                safe_write(ws, r_, cc, out_text)
                if debug:
                    print(f"[DEBUG] PASTA TEXT WRITE r={r_} c={cc} val='{out_text}'")
            else:
                safe_write(ws, r_, cc, v)
                if debug:
                    print(f"[DEBUG] WRITE r={r_} c={cc} val={v}")
        except Exception as e:
            if debug:
                print(f"[DEBUG] WRITE ERROR r={r_} c={cc} val={'{:.2f}'.format(v) if isinstance(v,(int,float)) else v} error={e}")
            # Try to write to master cell if merged
            try:
                if is_pasta_cell:
                    fmt_v = int(v) if float(v).is_integer() else v
                    unit_text = "KL." if force_koli_all else "SPT."
                    out_text = f"{fmt_v} {unit_text}"
                    safe_write(ws, r_, cc, out_text)
                    if debug:
                        print(f"[DEBUG] PASTA MASTER WRITE r={r_} c={cc} val='{out_text}'")
                else:
                    # Use safe_write to attempt writing into the master cell of the merge
                    safe_write(ws, r_, cc, v)
                    if debug:
                        print(f"[DEBUG] MASTER WRITE r={r_} c={cc} val={v}")
            except Exception as e2:
                if debug:
                    print(f"[DEBUG] MASTER WRITE ERROR r={r_} c={cc} val={v} error={e2}")

    # Explicit matrix writes: clear touched cells, write aggregated, mirror ŞERBET per column
    if matrix_aggreg:
        touched: Dict[int, set] = {}
        for (rr, cc) in matrix_aggreg.keys():
            touched.setdefault(rr, set()).add(cc)
        # clear - also clean text values with qty/unit patterns
        for rr, cols in touched.items():
            for cc in sorted(cols):
                try:
                    cell = ws.cell(row=rr, column=cc)
                    val = cell.value
                    # Clear numeric values
                    if isinstance(val, (int, float)) or (isinstance(val, str) and val.strip() and re.match(r"^[0-9,\.]+$", val.strip())):
                        cell.value = None
                    # Clean text values that contain qty/unit patterns (e.g., "2 KL.", "4 SPT.")
                    elif isinstance(val, str) and re.search(r"\d+\s*(?:SPT\.|KL\.|TEPSI|TEPSİ)", val, re.IGNORECASE):
                        # Clean the text but keep the product name part
                        cleaned = clean_text_from_quantities(val)
                        if cleaned:
                            cell.value = cleaned
                        else:
                            cell.value = None
                except Exception:
                    pass
        # write (as text with unit KOLI for TOST/EKMEK/CHEESECAKE/ÇATAL BÖREK)
        for (rr, cc), v in matrix_aggreg.items():
            try:
                fmt_v = int(v) if float(v).is_integer() else v
                unit_text = "KL."  # always KOLI for these groups; special sheets also KOLI
                out_text = f"{fmt_v} {unit_text}"
                safe_write(ws, rr, cc, out_text)
                if debug:
                    print(f"[DEBUG] MATRIX WRITE r={rr} c={cc} val='{out_text}'")
            except Exception as e:
                if debug:
                    print(f"[DEBUG] MATRIX WRITE ERROR r={rr} c={cc} val='{v}' error={e}")
                # Try to write to master cell if merged
                try:
                    safe_write(ws, rr, cc, out_text)
                    if debug:
                        print(f"[DEBUG] MATRIX MASTER WRITE r={rr} c={cc} val='{out_text}'")
                except Exception as e2:
                    if debug:
                        print(f"[DEBUG] MATRIX MASTER WRITE ERROR r={rr} c={cc} val='{out_text}' error={e2}")
        # no direct numeric mirroring for ŞERBET; text update below

    # Update KÜNEFE/ŞERBET/DONUK KAR. TRİLEÇE using MEYVELI ROKOKO text update logic
    # These should update the text in the cell, not write to separate numeric cells

    if kunefe_total_sum and ws is not None and min_c and max_c:
        fmt_qty = int(kunefe_total_sum) if float(kunefe_total_sum).is_integer() else kunefe_total_sum
        
        # Search for KÜNEFE only in branch span columns (min_c to max_c)
        for r in range(1, ws.max_row + 1):
            for c in range(min_c, max_c + 1):
                val = ws.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                upv = normalize_text(val)
                if "KUNEFE" in upv or "KÜNEFE" in upv:
                    # CRITICAL: Clean existing cell value to remove old qty/unit before appending new value
                    text_clean = clean_text_from_quantities(val)
                    unit_text = "KL." if force_koli_all else "SPT."
                    new_text = append_text_with_space(text_clean, f"{fmt_qty} {unit_text}")
                    try:
                        safe_write(ws, r, c, new_text)
                    except Exception:
                        try:
                            ws.cell(row=r, column=c).value = new_text
                        except Exception:
                            pass
                    if debug:
                        print(f"[DEBUG] KUNEFE TEXT WRITE r={r} c={c} val='{new_text}' (branch column)")
                    # Update only the first match
                    break
            else:
                continue
            break

    if trilece_total_sum and ws is not None and min_c and max_c:
        fmt_qty = int(trilece_total_sum) if float(trilece_total_sum).is_integer() else trilece_total_sum
        
        # Search for DONUK KAR. TRİLEÇE only in branch span columns (min_c to max_c)
        for r in range(1, ws.max_row + 1):
            for c in range(min_c, max_c + 1):
                    val = ws.cell(row=r, column=c).value
                    if not isinstance(val, str):
                        continue
                    upv = normalize_text(val)
                    if "DONUK" in upv and ("TRILECE" in upv or "TRİLEÇE" in upv):
                        # CRITICAL: Clean existing cell value to remove old qty/unit before appending new value
                        text_clean = clean_text_from_quantities(val)
                        unit_text = "KL." if force_koli_all else "SPT."
                        new_text = append_text_with_space(text_clean, f"{fmt_qty} {unit_text}")
                        try:
                            safe_write(ws, r, c, new_text)
                        except Exception:
                            try:
                                ws.cell(row=r, column=c).value = new_text
                            except Exception:
                                pass
                        if debug:
                            print(f"[DEBUG] TRİLEÇE TEXT WRITE r={r} c={c} val='{new_text}' (branch column)")
                        # Update only the first match
                        break
            else:
                continue
            break

    # ŞERBET should mirror KÜNEFE quantity using text update logic
    if kunefe_total_sum and ws is not None and min_c and max_c:
        fmt_qty = int(kunefe_total_sum) if float(kunefe_total_sum).is_integer() else kunefe_total_sum
        
        # Search for ŞERBET only in branch span columns (min_c to max_c)
        for r in range(1, ws.max_row + 1):
            for c in range(min_c, max_c + 1):
                    val = ws.cell(row=r, column=c).value
                    if not isinstance(val, str):
                        continue
                    upv = normalize_text(val)
                    if "SERBET" in upv or "ŞERBET" in upv:
                        # CRITICAL: Clean existing cell value to remove old qty/unit before appending new value
                        text_clean = clean_text_from_quantities(val)
                        unit_text = "KL." if force_koli_all else "SPT."
                        new_text = append_text_with_space(text_clean, f"{fmt_qty} {unit_text}")
                        try:
                            safe_write(ws, r, c, new_text)
                        except Exception:
                            try:
                                ws.cell(row=r, column=c).value = new_text
                            except Exception:
                                pass
                        if debug:
                            print(f"[DEBUG] ŞERBET TEXT WRITE r={r} c={c} val='{new_text}' (branch column)")
                        # Update only the first match
                        break
            else:
                continue
            break    

    # Update MEYVELI ROKOKO text cell if present and total > 0
    if rokoko_total and ws is not None and min_c and max_c:
        fmt_qty = int(rokoko_total) if float(rokoko_total).is_integer() else rokoko_total
        
        # Search for ROKOKO only in branch span columns (min_c to max_c)
        for r in range(1, ws.max_row + 1):
            for c in range(min_c, max_c + 1):
                    val = ws.cell(row=r, column=c).value
                    if not isinstance(val, str):
                        continue
                    upv = normalize_text(val)
                    if "ROKOKO" in upv:
                        # CRITICAL: Clean existing cell value to remove old qty/unit before formatting
                        text_clean = clean_text_from_quantities(val)
                        unit_text = "KL." if force_koli_all else "SPT."
                        new_text = format_text_with_qty(text_clean, f"{fmt_qty} {unit_text}")
                        try:
                            safe_write(ws, r, c, new_text)
                        except Exception:
                            try:
                                ws.cell(row=r, column=c).value = new_text
                            except Exception:
                                pass
                        if debug:
                            print(f"[DEBUG] MEYVELI ROKOKO TEXT WRITE r={r} c={c} val='{new_text}' (branch column)")
                        # Update only the first match
                        break
            else:
                continue
            break

    if ekler_total_sum and ws is not None and min_c and max_c:
        fmt_qty = int(ekler_total_sum) if float(ekler_total_sum).is_integer() else ekler_total_sum
        
        # Search for EKLER only in branch span columns (min_c to max_c)
        for r in range(1, ws.max_row + 1):
            for c in range(min_c, max_c + 1):
                    val = ws.cell(row=r, column=c).value
                    if not isinstance(val, str):
                        continue
                    upv = normalize_text(val)
                    if "EKLER" in upv:
                        # CRITICAL: Clean existing cell value to remove old qty/unit before formatting
                        text_clean = clean_text_from_quantities(val)
                        unit_text = "KL." if force_koli_all else "SPT."
                        new_text = format_text_with_qty(text_clean, f"{fmt_qty} {unit_text}")
                        try:
                            safe_write(ws, r, c, new_text)
                        except Exception:
                            try:
                                ws.cell(row=r, column=c).value = new_text
                            except Exception:
                                pass
                        if debug:
                            print(f"[DEBUG] EKLER TEXT WRITE r={r} c={c} val='{new_text}' (branch column)")
                        # Update only the first match
                        break
            else:
                continue
            break

    wb.save(output_path)
    # If forced hits were collected, print a concise report for the trial
    if forced_hits:
        print("\n[FORCED-DONUK REPORT] The following CSV rows were treated as forced donuk candidates and matched:")
        for fh in forced_hits:
            print(f" - CSV: '{fh.get('csv_name')}' -> Matched Excel Key: '{fh.get('matched_key')}' Qty: {fh.get('qty')}")

    print(f"[DONUK] Dondurmalar yazıldı: {len(aggreg)} hücre, {matched} kalem -> {output_path}")
    return (matched, 0)

# Minimal placeholder for compatibility

def process_csv(csv_path: str, output_path: str = "sevkiyat_tatlı.xlsx", sheet_name: Optional[str] = None) -> Tuple[int, int]:
    # Restore legacy Tatlı writer: reads CSV 'TATLI' items and writes to sevkiyat_tatlı.xlsx template
    # using the branch columns (TEPSI/ADET) and product rows discovered at runtime.
    # CRITICAL: sheet_name parameter for multi-day branch support
    # Helpers specific to Tatlı flow
    def normalize_text_strict(s):
        return normalize_text(s).replace(" ", "")

    def normalize_variant(v, product_name=""):
        if not v:
            return ""
        v = str(v).upper()
        v = v.replace("*", "X")
        v = v.replace("İ", "I").replace("Ş", "S").replace("Ğ", "G").replace("Ü", "U").replace("Ö", "O").replace("Ç", "C")
        v = v.replace("BÜYÜK", "BUYUK").replace("EKONOMIK PAKET", "BUYUK").replace("EKONOMIKPAKET", "BUYUK")
        v = v.replace("TEKLIPAKET", "PAKET").replace("TEKLI PAKET", "PAKET")
        if product_name in ("EKMEK KADAYIFI", "SEKERPARE"):
            return "ADET"
        if "TEPSI" in v or "TEPSİ" in v or re.search(r"\b42\b", v) or re.search(r"1X?42", v):
            return "TEPSI"
        if "BUYUK" in v:
            return "BUYUK"
        if "KASE" in v:
            return "KASE"
        if "TEKLI" in v:
            return "TEKLI"
        if "PAKET" in v:
            return "PAKET"
        if re.search(r"\d+X\d+", v) or "ADET" in v or re.search(r"PK", v) or "GR" in v:
            return "ADET"
        return re.sub(r"[^\w\s]", "", v).strip()

    def split_tatli_and_variant(s):
        s = str(s or "")
        s = unicodedata.normalize("NFKD", s).upper()
        s = s.replace("İ", "I").replace("Ş", "S").replace("Ğ", "G").replace("Ü", "U").replace("Ö", "O").replace("Ç", "C")
        s = re.sub(r"\{.*?\}", "", s)
        s = re.sub(r"[^\w\s\*\(\)]", "", s)
        s = re.sub(r"\s+", " ", s)
        s = s.strip()
        m = re.match(r"^(.*?)\s*\(([^)]+)\)", s)
        if m:
            ana_ad = m.group(1).strip()
            varyant = m.group(2).strip()
        else:
            ana_ad = s
            varyant = ""
        varyant_kelimeler = ["TEKLI PAKET", "TEKLI", "PAKET", "KASE", "BUYUK", "ADET", "TEPSI"]
        for kelime in varyant_kelimeler:
            if ana_ad.endswith(" " + kelime):
                ana_ad = ana_ad[:-(len(kelime)+1)].strip()
                varyant = kelime
            elif ana_ad.endswith(kelime):
                ana_ad = ana_ad[:-(len(kelime))].strip()
                varyant = kelime
        if re.search(r"\bTEPS[Iİ]\b", ana_ad) or re.search(r"42 ?L[Iİ]?", ana_ad) or re.search(r"1\*?42", ana_ad):
            ana_ad = re.sub(r"\bTEPS[Iİ]\b", "", ana_ad)
            ana_ad = re.sub(r"42 ?L[Iİ]?", "", ana_ad)
            ana_ad = re.sub(r"1\*?42", "", ana_ad)
            ana_ad = ana_ad.strip()
            varyant = "TEPSI"
        return ana_ad.strip(), varyant.strip()

    def tatli_eslesir(excel_ad, csv_ad):
        if excel_ad == csv_ad:
            return True
        ex_strict = normalize_text_strict(excel_ad)
        csv_strict = normalize_text_strict(csv_ad)
        
        # CRITICAL: TAVUK GÖĞSÜ vs TAVUK GÖĞSÜLÜ KAZANDİBİ distinction
        # These products must NEVER match each other - they are completely different items
        if "TAVUKGOGSU" in ex_strict and "TAVUKGOGSU" in csv_strict:
            # Both contain TAVUKGOGSU - check if one is plain and other is KAZANDIBI
            ex_has_kaz = "KAZANDIBI" in ex_strict or "KAZ" in ex_strict
            csv_has_kaz = "KAZANDIBI" in csv_strict or "KAZ" in csv_strict
            
            # If mismatch (one has KAZ, other doesn't), DO NOT match
            if ex_has_kaz != csv_has_kaz:
                return False
            
            # If both have KAZ, allow match
            if ex_has_kaz and csv_has_kaz:
                return True
            
            # If neither has KAZ (both plain TAVUK GÖĞSÜ), allow match only if CSV doesn't have KAZANDIBI/KAZ
            # This prevents "TAVUK GÖĞSÜ" Excel from matching "TAVUK GÖĞSÜLÜ KAZANDİBİ" CSV
            if not ex_has_kaz and not csv_has_kaz:
                return True
        
        # Original special case for TAVUK GÖĞSÜLÜ KAZANDİBİ matching
        if "TAVUKGOGSUKAZ" in ex_strict and "TAVUKGOGSUKAZANDIBI" in csv_strict:
            return True
        
        # General substring matching - but exclude TAVUKGOGSU to prevent cross-matching
        if (ex_strict in csv_strict or csv_strict in ex_strict) and ("KAZ" not in ex_strict and "KAZANDIBI" not in ex_strict):
            # Additional safety: if Excel has TAVUKGOGSU, CSV must not have KAZANDIBI
            if "TAVUKGOGSU" in ex_strict and ("KAZANDIBI" in csv_strict or "KAZ" in csv_strict):
                return False
            return True
        return False

    def varyant_eslesir(excel_v, csv_v):
        if excel_v == csv_v:
            return True
        if excel_v == "TEPSI":
            return csv_v == "TEPSI"
        if excel_v == "ADET":
            return csv_v in ("ADET", "")
        if not excel_v:
            return csv_v in ("", "ADET")
        return excel_v == csv_v

    # CSV oku (eski davranışla uyumlu header toleransı)
    try:
        df = pd.read_csv(csv_path, encoding="utf-8", delimiter=",", header=2)
    except Exception:
        df = pd.read_csv(csv_path, encoding="utf-8", delimiter=",", header=0)

    stok_col = find_col(df, ["STOK KODU", "STOKKODU", "KOD"]) or "STOK KODU"
    miktar_col = find_col(df, ["MIKTAR", "MİKTAR", "ADET"]) or "MIKTAR"
    grup_col = find_col(df, ["GRUP", "KATEGORI", "KATEGORI ADI"]) or "GRUP"

    # Extract branch name from CSV with priority: inner (primary) then outer (fallback)
    sube_primary_raw, sube_fallback_raw = read_branch_from_file(csv_path)
    if not sube_primary_raw and not sube_fallback_raw:
        raise ValueError("CSV'den şube adı (ŞUBE KODU/ADI) tespit edilemedi.")
    
    # Apply branch name mapping (e.g., FORUMAVM → FORUM, HARMANDALI → EFESUS)
    from shipment_oop import BranchDecisionEngine
    sube_primary = BranchDecisionEngine._apply_branch_mapping(sube_primary_raw) if sube_primary_raw else None
    sube_fallback = BranchDecisionEngine._apply_branch_mapping(sube_fallback_raw) if sube_fallback_raw else None
    
    sube_primary_norm = normalize_text(sube_primary) if sube_primary else None
    sube_fallback_norm = normalize_text(sube_fallback) if sube_fallback else None

    # Tatlı şablonu/yazım dosyası
    if not os.path.exists(output_path):
        raise FileNotFoundError(f"Tatlı şablonu bulunamadı: {output_path}")
    wb = load_workbook(output_path)

    # Doğru sheet'i ve şube sütunlarını bul (2. satırda şube başlıkları)
    # PRIORITY: Try primary branch first (parantez içi), then fallback (parantez dışı)
    ws = None
    col_tepsi = None
    col_adet = None
    matched_branch = None
    
    # Determine which sheets to search
    sheets_to_search = []
    if sheet_name:
        # User specified a sheet - use ONLY that sheet
        sheet_name_norm = normalize_text(sheet_name)
        for w in wb.worksheets:
            if normalize_text(w.title) == sheet_name_norm or normalize_text(w.title) in sheet_name_norm or sheet_name_norm in normalize_text(w.title):
                sheets_to_search.append(w)
                break
        if not sheets_to_search:
            # Fallback: if exact match failed, use first sheet
            sheets_to_search = [wb.worksheets[0]]
    else:
        # No sheet specified - search all sheets
        sheets_to_search = wb.worksheets
    
    # Helper function for branch matching with exact match priority
    def find_branch_columns(branch_norm: str, branch_display: str, sheets: list):
        """Find branch columns with exact match priority (avoids FOLKART matching FOLKART VEGA)"""
        if not branch_norm:
            return None, None, None, None
        
        # PASS 1: Exact matches only
        for w in sheets:
            subeler = {}
            for cell in w[2][1:]:  # row=2, columns after first
                if cell.value:
                    sname = normalize_text(cell.value)
                    subeler[sname] = {
                        "tepsi": cell.column,
                        "tepsi_2": cell.column + 1,
                        "adet": cell.column + 2,
                        "adet_2": cell.column + 3,
                    }
            # Exact match pass
            for sname, cols in subeler.items():
                if sname == branch_norm:  # Exact match only
                    return w, cols["tepsi"], cols["adet"], branch_display
        
        # PASS 2: Partial matches (backward compatibility)
        for w in sheets:
            subeler = {}
            for cell in w[2][1:]:
                if cell.value:
                    sname = normalize_text(cell.value)
                    subeler[sname] = {
                        "tepsi": cell.column,
                        "tepsi_2": cell.column + 1,
                        "adet": cell.column + 2,
                        "adet_2": cell.column + 3,
                    }
            # Partial match pass
            for sname, cols in subeler.items():
                if branch_norm in sname or sname in branch_norm:
                    return w, cols["tepsi"], cols["adet"], branch_display
        
        return None, None, None, None
    
    # First pass: try PRIMARY branch (parantez içi)
    ws, col_tepsi, col_adet, matched_branch = find_branch_columns(
        sube_primary_norm, sube_primary, sheets_to_search
    )
    
    # Second pass: if PRIMARY failed, try FALLBACK branch (parantez dışı)
    if ws is None and sube_fallback_norm:
        ws, col_tepsi, col_adet, matched_branch = find_branch_columns(
            sube_fallback_norm, sube_fallback, sheets_to_search
        )

    if ws is None or not col_tepsi or not col_adet:
        raise ValueError(f"Şube '{sube_primary or sube_fallback}' için hedef sayfa/sütunlar bulunamadı.")

    # Tarih yaz (B1 benzeri: row=2, col=1)
    from datetime import datetime
    ws.cell(row=2, column=1).value = datetime.today().strftime('%d.%m.%Y')

    # Tüm merge'leri geçici olarak aç
    original_merged = [str(r) for r in ws.merged_cells.ranges]
    for r in original_merged:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass

    # Excel tarafındaki tatlı ürünleri ve hedef hücreleri indexle
    tatli_cells: Dict[Tuple[str, str], Tuple[int, int]] = {}
    skip_keywords = ["SIPARIS TARIHI", "SIPARIS ALAN", "TESLIM TARIHI", "TEYID EDEN"]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=1):
        ana_cell = row[0]
        if not ana_cell.value:
            continue
        ana_ad, varyant = split_tatli_and_variant(ana_cell.value)
        ana_ad_norm = normalize_text(ana_ad)
        if not ana_ad_norm:
            continue
        if any(ana_ad_norm.startswith(k) or ana_ad_norm == k for k in skip_keywords):
            continue
        varyant_norm = normalize_variant(varyant)
        if "MIKTAR" in ana_ad_norm or "ADET" in ana_ad_norm or "TEPSI" in ana_ad_norm:
            continue
        if ana_cell.row <= 7:
            tatli_cells[(ana_ad_norm, "TEPSI")] = (ana_cell.row, col_tepsi)
            tatli_cells[(ana_ad_norm, "ADET")] = (ana_cell.row, col_adet)
        else:
            if varyant_norm == "TEPSI":
                tatli_cells[(ana_ad_norm, "TEPSI")] = (ana_cell.row, col_tepsi)
            else:
                tatli_cells[(ana_ad_norm, varyant_norm or "ADET")] = (ana_cell.row, col_tepsi)

    # CSV'den TATLI grubu ürünleri topla: {ana_ad_norm: [(varyant_norm, miktar), ...]}
    csv_index: Dict[str, list] = {}
    for _, r in df.iterrows():
        try:
            g = normalize_text(str(r[grup_col]))
        except Exception:
            g = ""
        if g != "TATLI":
            continue
        stok_name = r[stok_col]
        ana_ad, varyant = split_tatli_and_variant(stok_name)
        ana_ad_norm = normalize_text(ana_ad)
        varyant_norm = normalize_variant(varyant)
        try:
            mikt = float(str(r[miktar_col]).replace(",", "."))
        except Exception:
            mikt = r[miktar_col]
        csv_index.setdefault(ana_ad_norm, []).append((varyant_norm, mikt))

    # Önce hedef hücreleri temizle (eski alışkanlıkla: "-")
    for (a, v), (rr, cc) in tatli_cells.items():
        try:
            ws.cell(row=rr, column=cc).value = "-"
        except Exception:
            # nadir merge-hata güvenliği
            try:
                ws.unmerge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc)
                ws.cell(row=rr, column=cc).value = "-"
            except Exception:
                pass

    matched = 0
    # Yazma: önce doğrudan ad, sonra gevşek eşleşme
    for (excel_ad, excel_var), (rr, cc) in tatli_cells.items():
        yazildi = False
        if excel_ad in csv_index:
            for csv_var, csv_miktar in csv_index[excel_ad]:
                if varyant_eslesir(excel_var, csv_var):
                    # EKMEK KADAYIFI ve ŞEKERPARE için "TEPSİ" ekle
                    if excel_ad in ("EKMEKKADAYIFI", "SEKERPARE"):
                        try:
                            fmt_qty = int(csv_miktar) if float(csv_miktar).is_integer() else csv_miktar
                        except:
                            fmt_qty = csv_miktar
                        ws.cell(row=rr, column=cc).value = f"{fmt_qty} TEPSİ"
                    else:
                        ws.cell(row=rr, column=cc).value = csv_miktar
                    matched += 1
                    yazildi = True
                    break
                if excel_var == "ADET" and csv_var == "TEPSI" and excel_ad in ("EKMEK KADAYIFI", "SEKERPARE"):
                    # EKMEK KADAYIFI ve ŞEKERPARE için "TEPSİ" ekle
                    try:
                        fmt_qty = int(csv_miktar) if float(csv_miktar).is_integer() else csv_miktar
                    except:
                        fmt_qty = csv_miktar
                    ws.cell(row=rr, column=cc).value = f"{fmt_qty} TEPSİ"
                    matched += 1
                    yazildi = True
                    break
        if not yazildi:
            for csv_name, entries in csv_index.items():
                if tatli_eslesir(excel_ad, csv_name):
                    for csv_var, csv_miktar in entries:
                        if varyant_eslesir(excel_var, csv_var):
                            # EKMEK KADAYIFI ve ŞEKERPARE için "TEPSİ" ekle
                            if excel_ad in ("EKMEKKADAYIFI", "SEKERPARE"):
                                try:
                                    fmt_qty = int(csv_miktar) if float(csv_miktar).is_integer() else csv_miktar
                                except:
                                    fmt_qty = csv_miktar
                                ws.cell(row=rr, column=cc).value = f"{fmt_qty} TEPSİ"
                            else:
                                ws.cell(row=rr, column=cc).value = csv_miktar
                            matched += 1
                            yazildi = True
                            break
                        if excel_var == "ADET" and csv_var == "TEPSI" and excel_ad in ("EKMEK KADAYIFI", "SEKERPARE"):
                            # EKMEK KADAYIFI ve ŞEKERPARE için "TEPSİ" ekle
                            try:
                                fmt_qty = int(csv_miktar) if float(csv_miktar).is_integer() else csv_miktar
                            except:
                                fmt_qty = csv_miktar
                            ws.cell(row=rr, column=cc).value = f"{fmt_qty} TEPSİ"
                            matched += 1
                            yazildi = True
                            break
                        if excel_ad == "KAYMAK" and tatli_eslesir(excel_ad, csv_name):
                            ws.cell(row=rr, column=cc).value = csv_miktar
                            matched += 1
                            yazildi = True
                            break
                    if yazildi:
                        break

    # Merge'leri eski haline getir
    for r in original_merged:
        try:
            ws.merge_cells(r)
        except Exception:
            pass

    # ==================== SEPET HESAPLAMA ====================
    # Calculate basket (sepet) count for each branch based on product quantities
    # Handle basket calculation for all sheets
    
    import math
    
    sheet_name_up = normalize_text(ws.title) if ws.title else ""
    
    # Check if this is a KAYSERİ-SİVAS or ADANA sheet (different division values)
    is_special_sheet = any(keyword in sheet_name_up for keyword in ["KAYSERISIVAS", "ADANA", "KAYSERI", "SIVAS"])
    
    # Select appropriate division values based on sheet type
    if is_special_sheet:
        # KAYSERİ-SİVAS and ADANA sheets use different division values
        product_divisions = {
            "KAZANDIBI": 60,
            "ANTEPFISTIKLIKAZANDIBI": 60,
            "TAVUKGOGSU": 60,
            "TAVUKGOGSUKAZANDIBI": 60,
            "TAVUKGOGSUKAZ": 60,
            "SAKIZLIMUHALLEBI": 60,
            "LIGHTKAZANDIBI": 60,
            "KESKULKASE": 32,
            "KESKULTEKLI": 60,
            "KESKULPAKET": 60,
            "KREMSOKOLAKASE": 32,
            "KREMSOKOLATEKLI": 60,
            "KREMCOKOLATEKLI": 60,
            "KREMCOKOLATAKASE": 32,
            "FIRINSUTLACKASE": 32,
            "FIRINSUTLACTEKLI": 48,
            "FIRINSUTLACBUYUK": 24,
            "ASUREKASE": 32,
            "ASURETEKLI": 60,
            "ASUREPAKET": 60,
            "ASUREBUYUK": 24,
            "PROFITEROLKASE": 40,
            "PROFITEROLTEKLI": 60,
            "PROFITEROLBUYUK": 24,
            "KREMKARAMEL": 60,
            "EKMEKKADAYIFI": 8,
            "SEKERPARE": 8,
            "KAYMAK": 18,
            "KAYMAKTAVA": 18,
        }
    else:
        # Standard sheets (all other locations)
        product_divisions = {
            "KAZANDIBI": 45,
            "ANTEPFISTIKLIKAZANDIBI": 45,
            "TAVUKGOGSU": 45,
            "TAVUKGOGSUKAZANDIBI": 45,
            "TAVUKGOGSUKAZ": 45,  # Short variant
            "SAKIZLIMUHALLEBI": 45,
            "LIGHTKAZANDIBI": 45,
            "KESKULKASE": 24,
            "KESKULTEKLI": 45,
            "KESKULPAKET": 45,  # Paket variant (TEKLİ PAKET)
            "KREMSOKOLAKASE": 24,
            "KREMSOKOLATEKLI": 45,
            "KREMCOKOLATEKLI": 45,
            "KREMCOKOLATAKASE": 24,
            "FIRINSUTLACKASE": 24,
            "FIRINSUTLACTEKLI": 36,
            "FIRINSUTLACBUYUK": 18,
            "ASUREKASE": 24,
            "ASURETEKLI": 45,
            "ASUREPAKET": 45,  # Paket variant (TEKLİ PAKET)
            "ASUREBUYUK": 18,
            "PROFITEROLKASE": 30,
            "PROFITEROLTEKLI": 45,
            "PROFITEROLBUYUK": 18,
            "KREMKARAMEL": 30,
            "EKMEKKADAYIFI": 6,
            "SEKERPARE": 6,
            "KAYMAK": 18,
            "KAYMAKTAVA": 18,  # Tava variant
        }
    
    # First 5 products have both TEPSI (basket) and ADET (piece) columns
    # TEPSI is already a basket count (integer), ADET needs to be divided
    # This applies to all sheets (both standard and special)
    first_five_products = [
        "KAZANDIBI",
        "ANTEPFISTIKLIKAZANDIBI", 
        "TAVUKGOGSU",
        "TAVUKGOGSUKAZANDIBI",
        "TAVUKGOGSUKAZ",  # Short variant of TAVUKGOGSUKAZANDIBI
        "SAKIZLIMUHALLEBI"
    ]
    
    # Get all branch columns from row 2
    # Row 2 format: [Date, Branch1, ..., BranchN]
    # Each branch has 4 columns: TEPSI, TEPSI_2, ADET, ADET_2
    subeler = {}
    for cell in ws[2][1:]:  # Skip first column (date)
        if cell.value and str(cell.value).strip():
            sname = normalize_text(cell.value)
            has_digit = any(char.isdigit() for char in str(cell.value)[:8])
            # Only process if it looks like a branch name (not a date)
            if sname and not has_digit:
                subeler[sname] = {
                    "tepsi_col": cell.column,
                    "adet_col": cell.column + 2,  # ADET is 2 columns after TEPSI
                    "sepet_row": 1,  # Write basket count to row 1 (header row)
                    "sepet_col": cell.column  # Write to branch column (merged cell master)
                }
    
    # Products to exclude from basket calculation
    excluded_products = ["EKMEKKADAYIFI", "SEKERPARE"]
    
    # Calculate basket count for each branch
    for sube_name, cols in subeler.items():
        total_baskets = 0.0  # Use float for decimal calculations
        
        # Iterate through all product rows
        for row_idx in range(3, ws.max_row + 1):
            ana_cell = ws.cell(row_idx, 1)
            if not ana_cell.value:
                continue
            
            # Get product name and variant
            ana_ad, varyant = split_tatli_and_variant(ana_cell.value)
            ana_ad_norm = normalize_text(ana_ad)
            varyant_norm = normalize_variant(varyant)
            
            # Skip non-product rows
            if not ana_ad_norm:
                continue
            skip_keywords = ["SIPARIS TARIHI", "SIPARIS ALAN", "TESLIM TARIHI", "TEYID EDEN", "MIKTAR", "ADET", "TEPSI"]
            if any(k in ana_ad_norm for k in skip_keywords):
                continue
            
            # Build full product key (product + variant)
            # Remove all spaces from product key to match dictionary keys
            if varyant_norm and varyant_norm not in ["ADET", "TEPSI"]:
                product_key = (ana_ad_norm + varyant_norm).replace(" ", "")
            else:
                product_key = ana_ad_norm.replace(" ", "")
            
            # Skip excluded products (EKMEK KADAYIFI, ŞEKERPARE)
            if product_key in excluded_products:
                continue
            
            # Get division value for this product
            division_value = product_divisions.get(product_key)
            if not division_value:
                continue  # Skip products without division value
            
            # Check if this is one of the first 5 products (has TEPSI column)
            # Use the normalized product key (no spaces) for comparison
            is_first_five = product_key in first_five_products
            
            if is_first_five:
                # For first 5 products: ONLY use ADET column (ignore TEPSI)
                # Get ADET value and divide
                adet_cell = ws.cell(row=row_idx, column=cols["adet_col"])
                adet_val = adet_cell.value
                if adet_val not in (None, "", "-"):
                    try:
                        adet_count = float(str(adet_val).replace(",", "."))
                        basket_from_adet = adet_count / division_value
                        total_baskets += basket_from_adet
                    except:
                        pass
            else:
                # Other products: only ADET column (at TEPSI position)
                adet_cell = ws.cell(row=row_idx, column=cols["tepsi_col"])
                adet_val = adet_cell.value
                if adet_val not in (None, "", "-"):
                    try:
                        adet_count = float(str(adet_val).replace(",", "."))
                        basket_from_adet = adet_count / division_value
                        total_baskets += basket_from_adet
                    except:
                        pass
        
        # Round up to nearest integer (ceiling)
        final_basket_count = math.ceil(total_baskets)
        
        # Write basket count using safe_write (handles merged cells)
        # Only write if there's a positive basket count
        if final_basket_count > 0:
            safe_write(ws, cols["sepet_row"], cols["sepet_col"], final_basket_count)

    wb.save(output_path)
    return matched, 0


def process_all(csv_path: str,
                tatli_output: str = "sevkiyat_tatlı.xlsx",
                donuk_output: str = "sevkiyat_donuk.xlsx",
                lojistik_output: str = "sevkiyat_lojistik.xlsx",
                debug: bool = False,
                force_donuk: Optional[Iterable[str]] = None):
    # process_donuk_csv now accepts an optional iterable of forced donuk candidate names
    result = process_donuk_csv(csv_path, output_path=donuk_output, debug=debug, force_donuk=force_donuk)
    # process_donuk_csv returns (matched, 0) by default; keep compatibility
    if isinstance(result, tuple) or isinstance(result, list):
        matched = result[0]
    else:
        matched = 0
    return {"donuk": {"matched": matched, "unmatched": 0, "file": donuk_output}}

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Kullanım: python parse_gptfix.py <csv_yolu> [--debug]")
        raise SystemExit(1)
    csv_path = sys.argv[1]
    debug_flag = "--debug" in sys.argv or "-d" in sys.argv
    # Parse optional forced donuk list: --force-donuk "MANTI, PATATES, ..."
    force_list = None
    if "--force-donuk" in sys.argv:
        try:
            idx = sys.argv.index("--force-donuk")
            if idx + 1 < len(sys.argv):
                raw = sys.argv[idx + 1]
                # split by comma
                force_list = [s.strip() for s in raw.split(",") if s.strip()]
        except Exception:
            force_list = None
    summary = process_all(csv_path, debug=debug_flag, force_donuk=force_list)
    print(summary)
