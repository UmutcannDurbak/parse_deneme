
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import threading
import os
from copy import copy
import sys
import requests  # pyright: ignore[reportMissingModuleSource]
import json
import zipfile
import shutil
from pathlib import Path
import subprocess
from copy import copy
import sys
import requests  # pyright: ignore[reportMissingModuleSource]
import json
import zipfile
import shutil
from pathlib import Path

DATA_START_ROW = 3  # Verilerin başladığı satır (1-indexed)

# tkinterdnd2 desteği varsa import et
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # pyright: ignore[reportMissingImports]
    TK_DND_AVAILABLE = True
except ImportError:
    TK_DND_AVAILABLE = False

# PyInstaller ile build ederken .ico dosyasını eklemeyi unutmayın!
ICON_PATH = "appicon.ico"
VERSION = "v1.3.47"
DEVELOPER = "Developer U.D"

# Güncelleme ayarları
GITHUB_REPO = "UmutcannDurbak/parse_deneme"  # GitHub repository (owner/repo)
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"  # GitHub API endpoint
UPDATE_CHECK_INTERVAL = 24 * 60 * 60  # 24 saat (saniye cinsinden)
# Eğer güncelleme bulunduğunda otomatik indirme başlatılsın mı? (False = kullanıcı "İndir" butonuna basmalı)
# Otomatik indirmenin varsayılan davranışı: eğer uygulama PyInstaller ile paketlenmişse otomatik indir
"""AUTO_START_DOWNLOAD:
If True, the app will automatically start downloading an available update when it detects a newer release.
We enable this for testing/automation so the app immediately downloads the selected asset.
In production, you may prefer to enable this only when running a packaged exe (frozen).
"""
AUTO_START_DOWNLOAD = True

# Tercih edilen asset uzantı sıralaması — önce .exe, sonra .zip
PREFERRED_ASSET_EXTENSIONS = ['.exe', '.zip']

def resource_path(relative_path):
    import sys, os
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# Güncelleme fonksiyonları
def get_latest_version():
    """GitHub'dan en son sürümü kontrol eder"""
    try:
        response = requests.get(GITHUB_API_URL, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data.get('tag_name', ''), data.get('html_url', '')
        return None, None
    except Exception as e:
        print(f"Güncelleme kontrolü hatası: {e}")
        return None, None

def download_github_update(download_url, progress_callback=None):
    """Stream-download a GitHub asset to 'update.zip'. Returns True on success."""
    try:
        resp = requests.get(download_url, stream=True, timeout=60)
        resp.raise_for_status()
        total = int(resp.headers.get('content-length', 0) or 0)
        downloaded = 0
        with open('update.zip', 'wb') as fh:
            for chunk in resp.iter_content(chunk_size=8192):
                if not chunk:
                    continue
                fh.write(chunk)
                downloaded += len(chunk)
                if progress_callback and total:
                    try:
                        progress_callback((downloaded / total) * 100)
                    except Exception:
                        pass
        return True
    except Exception as e:
        print(f"İndirme hatası: {e}")
        return False


def install_update():
    """Simple and reliable update: Use VBScript (UTF-16 + 8.3 paths) for silent updates.
    Returns (success: bool, updater_path: str|None, message: str)
    """
    try:
        frozen = getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS')

        if not os.path.exists('update.zip'):
            return False, None, 'update.zip bulunamadı'

        with zipfile.ZipFile('update.zip', 'r') as z:
            # Find any exe file in the zip
            found_exe = None
            for name in z.namelist():
                if name.lower().endswith('.exe'):
                    found_exe = name
                    break

            if not found_exe:
                files_in_zip = ', '.join(z.namelist())
                return False, None, f'ZIP içinde exe bulunamadı. İçerik: {files_in_zip}'

            if frozen:
                import tempfile
                import ctypes
                from ctypes import wintypes

                def get_short_path(path: str) -> str:
                    try:
                        GetShortPathNameW = ctypes.windll.kernel32.GetShortPathNameW
                        GetShortPathNameW.argtypes = [wintypes.LPCWSTR, wintypes.LPWSTR, wintypes.DWORD]
                        GetShortPathNameW.restype = wintypes.DWORD
                        buf = ctypes.create_unicode_buffer(260)
                        r = GetShortPathNameW(path, buf, 260)
                        if r == 0:
                            return path
                        return buf.value
                    except Exception:
                        return path

                tmpdir = tempfile.mkdtemp()

                # Extract the new exe
                z.extract(found_exe, tmpdir)
                new_exe_path = os.path.join(tmpdir, found_exe)

                if not os.path.exists(new_exe_path):
                    return False, None, 'Çıkarılan exe bulunamadı'

                # Get paths
                current_exe = sys.executable
                # Create runnable timestamped backup alongside the exe (ends with .exe)
                import time as _upd_time
                exe_dir, exe_base = os.path.split(current_exe)
                _name, _ext = os.path.splitext(exe_base)
                backup_exe = os.path.join(exe_dir, f"{_name}_backup_{_upd_time.strftime('%Y%m%d_%H%M%S')}{_ext}")
                current_pid = os.getpid()

                # Compute 8.3 short paths to avoid Unicode issues in WSH/FSO
                sp_new = get_short_path(new_exe_path)
                sp_target = get_short_path(current_exe)
                sp_backup = get_short_path(backup_exe)
                sp_tmp = get_short_path(tmpdir)

                # Create VBScript updater (no dependencies, silent, reliable)
                vbs_script = os.path.join(tmpdir, 'updater.vbs')
                vbs_code = (
                    'Option Explicit\r\n'
                    'On Error Resume Next\r\n'
                    'Dim objFSO, objShell, objWMI\r\n'
                    'Dim strNewExe, strTargetExe, strBackupExe, strTmpDir\r\n'
                    'Dim intPID, intCounter, bProcessExists\r\n\r\n'
                    'strNewExe = "' + sp_new.replace('\\', '\\\\') + '"\r\n'
                    'strTargetExe = "' + sp_target.replace('\\', '\\\\') + '"\r\n'
                    'strBackupExe = "' + sp_backup.replace('\\', '\\\\') + '"\r\n'
                    'strTmpDir = "' + sp_tmp.replace('\\', '\\\\') + '"\r\n'
                    f'intPID = {current_pid}\r\n\r\n'
                    'Set objFSO = CreateObject("Scripting.FileSystemObject")\r\n'
                    'Set objShell = CreateObject("WScript.Shell")\r\n'
                    'Set objWMI = GetObject("winmgmts:{impersonationLevel=impersonate}!\\\\.\\root\\cimv2")\r\n\r\n'
                    "For intCounter = 1 To 50\r\n"
                    "    bProcessExists = False\r\n"
                    "    Dim colProcesses, objProcess\r\n"
                    "    Set colProcesses = objWMI.ExecQuery(\"SELECT * FROM Win32_Process WHERE ProcessId = \" & intPID)\r\n"
                    "    For Each objProcess In colProcesses\r\n"
                    "        bProcessExists = True\r\n"
                    "        Exit For\r\n"
                    "    Next\r\n"
                    "    If Not bProcessExists Then Exit For\r\n"
                    "    WScript.Sleep 100\r\n"
                    "Next\r\n\r\n"
                    "WScript.Sleep 1000\r\n\r\n"
                    'If objFSO.FileExists(strTargetExe) Then objFSO.CopyFile strTargetExe, strBackupExe, True\r\n'
                    'Err.Clear\r\n'
                    'objFSO.CopyFile strNewExe, strTargetExe, True\r\n'
                    'If Err.Number <> 0 Then WScript.Quit 1\r\n'
                    'Err.Clear\r\n'
                    'WScript.Sleep 2500\r\n'
                    'Err.Clear\r\n'
                    'objShell.ShellExecute strTargetExe, "", "", "open", 1\r\n'
                    'If Err.Number <> 0 Then\r\n'
                    '    Dim dt, st, cmdCreate\r\n'
                    '    dt = Now + TimeSerial(0,0,10)\r\n'
                    '    st = Right("0" & Hour(dt),2) & ":" & Right("0" & Minute(dt),2) & ":" & Right("0" & Second(dt),2)\r\n'
                    '    cmdCreate = "schtasks /Create /F /SC ONCE /ST " & st & " /TN TatliSiparisAutoRun /TR " & Chr(34) & strTargetExe & Chr(34)\r\n'
                    '    objShell.Run cmdCreate, 0, True\r\n'
                    '    objShell.Run "schtasks /Run /TN TatliSiparisAutoRun", 0, False\r\n'
                    'End If\r\n'
                    'WScript.Sleep 2000\r\n'
                    'objFSO.DeleteFolder strTmpDir, True\r\n'
                    'WScript.Quit 0\r\n'
                )

                # Write as UTF-16 LE with BOM (WSH Unicode-safe)
                with open(vbs_script, 'w', encoding='utf-16') as f:
                    f.write(vbs_code)

                # Clean up the update.zip so it doesn't get re-used
                try:
                    os.remove('update.zip')
                except Exception:
                    pass

                return True, vbs_script, 'VBScript updater hazır'

            else:
                # Non-frozen: direct update
                current_dir = os.getcwd()
                backup_path = os.path.join(current_dir, found_exe + '.backup')
                target_path = os.path.join(current_dir, found_exe)

                if os.path.exists(target_path):
                    shutil.copy2(target_path, backup_path)

                z.extract(found_exe, current_dir)

                try:
                    os.remove('update.zip')
                except Exception:
                    pass

                return True, None, 'Güncelleme tamamlandı'

    except Exception as e:
        return False, None, f"Kurulum hatası: {e}"
def check_for_updates(silent=False):
    """Background check for updates. If AUTO_START_DOWNLOAD is True, will auto-download and install.
    Returns True if an update was applied/launched, False otherwise.
    """
    try:
        latest_version, _ = get_latest_version()
        if not latest_version:
            return False
        if not is_newer_version(latest_version, VERSION):
            return False
        # fetch release details
        r = requests.get(GITHUB_API_URL, timeout=10)
        if r.status_code != 200:
            return False
        data = r.json()
        assets = data.get('assets', [])
        best = select_best_asset(assets)
        if not best:
            return False
        dl = best.get('browser_download_url')
        if not dl:
            return False
        if silent and not AUTO_START_DOWNLOAD:
            return False
        ok = download_github_update(dl)
        if not ok:
            return False
        return install_update()
    except Exception:
        return False

def select_best_asset(assets: list):
    """Verilen asset listesi içinden en uygun (tercih edilen uzantıya göre) asset'i döndürür.
    Döndürür: asset dict veya None
    """
    if not assets:
        return None
    # normalize names
    assets_sorted = list(assets)
    # try preferred extensions in order
    for ext in PREFERRED_ASSET_EXTENSIONS:
        for a in assets_sorted:
            name = (a.get('name') or '').lower()
            if name.endswith(ext) or ext.strip('.') in name:
                return a
    # fallback: return first asset
    return assets_sorted[0]


def is_newer_version(latest_version, current_version):
    """Basit semantik versiyon karşılaştırması. 'v' önekini kaldırır ve noktalı int'leri karşılaştırır."""
    try:
        def to_tuple(v):
            v = str(v).lstrip('vV')
            parts = [int(p) for p in v.split('.') if p.isdigit() or p.isnumeric()]
            return tuple(parts)
        return to_tuple(latest_version) > to_tuple(current_version)
    except Exception:
        return False

# Yeni OOP koordinatör (eski fonksiyonlar geriye dönük uyum için içeride kullanılacak)
from shipment_oop import (
    ShipmentCoordinator, 
    clear_workbook_values, 
    format_today_in_workbook, 
    IZMIR_BRANCHES,
    BranchDecisionEngine,
    MULTI_DAY_BRANCHES,
    SHEET_NAME_MAPPING,
    clear_tatli_values,
    clear_donuk_values,
    clear_lojistik_values
)
'''
# Hücre formatını bozmadan sadece ana/master hücreye değer silen fonksiyon
def clear_cell_value_preserve_format(ws, row, col, clear_formulas=False):
    """
    Hücreyi içindeki değeri temizler ancak hücre biçimini/merge yapısını bozmaz.
    - Eğer (row,col) bir merged-range içindeyse, merged aralığın master hücresini temizler.
    - clear_formulas=False ise formülleri silmez (korur).
    Döner: True (bir değer temizlendi), False (zaten boş veya formül korundu).
    """
    # merged-range içinde mi bak
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mr.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            master = ws.cell(row=min_row, column=min_col)
            # formül koruması
            if not clear_formulas and isinstance(master.value, str) and str(master.value).startswith('='):
                return False
            if master.value not in (None, ""):
                master.value = None
                return True
            return False

    # merged değilse direkt hücreyi temizle
    cell = ws.cell(row=row, column=col)
    if not clear_formulas and isinstance(cell.value, str) and str(cell.value).startswith('='):
        return False
    if cell.value not in (None, ""):
        cell.value = None
        return True
    return False
'''
from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]
import datetime

def clear_all_records(status_label, log_widget):
    """Clear tatlı file using the new clear_tatli_values function"""
    confirm = messagebox.askyesno("Onay", "Tatlı dosyasındaki tüm kayıtları (sepet değerleri dahil) temizlemek istediğinize emin misiniz?")
    if not confirm:
        status_label.config(text="İşlem iptal edildi.")
        return
    try:
        output_path = "sevkiyat_tatlı.xlsx"
        if not os.path.exists(output_path):
            status_label.config(text="❌ Önce bir sevkiyat dosyası oluşturulmalı!")
            messagebox.showerror("Hata", "Önce bir sevkiyat dosyası oluşturulmalı!")
            return
        
        cleared = clear_tatli_values(output_path)
        
        status_label.config(text=f"✅ Tatlı dosyası temizlendi! ({cleared} hücre)")
        safe_log_insert(log_widget, f"✅ Tatlı dosyası temizlendi! ({cleared} hücre - sepet değerleri dahil)\n")
    except Exception as e:
        status_label.config(text=f"❌ Hata: {e}")
        messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")

def _clear_cell_preserve_merge(ws, row, col, clear_formulas=False):
    """
    Tek bir hücreyi clear eder. Eğer hücre merged-range içindeyse:
    - Eğer merged master DATA_START_ROW'dan küçükse -> header master, silme (return False)
    - Aksi halde o range'i geçici unmerge et, hedef hücreyi temizle, sonra range'i merge edip master'ı restore et.
    Döner: True eğer bir hücre temizlendiyse, False aksi halde.
    """

    cell = ws.cell(row=row, column=col)

    if row == 2:
        return False
    if not clear_formulas and isinstance(cell.value, str) and cell.value.startswith('='):
        return False
    if cell.value not in (None, ""):
        cell.value = None
        return True
    return False

def show_day_selection_dialog(branch_name, possible_sheets):
    """Show a modal dialog for day selection with radio buttons.
    
    Returns:
        Selected sheet name or None if cancelled
    """
    dialog = tk.Toplevel()
    dialog.title("Gün Seçimi Gerekli")
    dialog.geometry("450x300")
    dialog.resizable(False, False)
    dialog.grab_set()  # Modal dialog
    
    # Center the dialog on screen
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
    y = (dialog.winfo_screenheight() // 2) - (300 // 2)
    dialog.geometry(f"+{x}+{y}")
    
    selected_sheet = tk.StringVar(value="")
    
    # Header with icon and message
    header_frame = tk.Frame(dialog, bg="#FFF3CD", pady=15)
    header_frame.pack(fill="x", padx=0, pady=0)
    
    icon_label = tk.Label(header_frame, text="⚠️", font=("Arial", 24), bg="#FFF3CD")
    icon_label.pack(side="left", padx=15)
    
    msg_text = f"'{branch_name}' şubesi birden fazla sevkiyat gününde bulunuyor.\nLütfen hangi gün için işlem yapmak istediğinizi seçin:"
    msg_label = tk.Label(header_frame, text=msg_text, font=("Arial", 10), bg="#FFF3CD", justify="left")
    msg_label.pack(side="left", padx=5)
    
    # Radio button frame
    radio_frame = tk.Frame(dialog, pady=20)
    radio_frame.pack(fill="both", expand=True, padx=30)
    
    tk.Label(radio_frame, text="Sevkiyat Günü:", font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 10))
    
    for sheet in possible_sheets:
        rb = tk.Radiobutton(
            radio_frame, 
            text=sheet, 
            variable=selected_sheet, 
            value=sheet,
            font=("Arial", 10),
            anchor="w"
        )
        rb.pack(anchor="w", pady=5, padx=20)
    
    # Button frame
    btn_frame = tk.Frame(dialog, pady=15)
    btn_frame.pack(fill="x", padx=30, side="bottom")
    
    result = [None]  # Use list to store result (closure workaround)
    
    def on_ok():
        if selected_sheet.get():
            result[0] = selected_sheet.get()
            dialog.destroy()
        else:
            messagebox.showwarning("Uyarı", "Lütfen bir gün seçin!", parent=dialog)
    
    def on_cancel():
        result[0] = None
        dialog.destroy()
    
    ok_btn = tk.Button(btn_frame, text="✓ Tamam", command=on_ok, width=12, bg="#28A745", fg="white", font=("Arial", 10, "bold"))
    ok_btn.pack(side="right", padx=5)
    
    cancel_btn = tk.Button(btn_frame, text="✗ İptal", command=on_cancel, width=12, font=("Arial", 10))
    cancel_btn.pack(side="right", padx=5)
    
    # Set focus to first radio button
    dialog.focus_set()
    
    # Wait for dialog to close
    dialog.wait_window()
    
    return result[0]


def run_process(csv_path, status_label, log_widget, izmir_day_var=None, show_popup=True):
    try:
        log_lines = []
        def custom_print(*args, **kwargs):
            msg = ' '.join(str(a) for a in args)
            log_lines.append(msg)
            def append_log():
                log_widget.config(state='normal')
                log_widget.insert(tk.END, msg + '\n')
                log_widget.see(tk.END)
                log_widget.config(state='disabled')
                log_widget.update_idletasks()
            log_widget.after(0, append_log)
        
        # Koordinatörü kullanarak üç sevkiyat dosyasını oluştur
        coord = ShipmentCoordinator()
        sheet_hint = izmir_day_var.get() if izmir_day_var else None
        sheet_hint = sheet_hint if sheet_hint not in ("", "Seçim yok") else None
        
        # Initialize branch_name and csv_filename early for use in messages
        branch_name = None
        csv_filename = os.path.basename(csv_path)
        
        # Check if branch requires day selection
        from shipment_oop import BranchDecisionEngine, CsvOrderReader, SHEET_NAME_MAPPING
        try:
            reader = CsvOrderReader(csv_path)
            reader.load()
            branch_name = reader.get_branch_name()
            
            # Also try read_branch_from_file for better accuracy
            if not branch_name:
                import re
                with open(csv_path, encoding="utf-8") as f:
                    for line in f:
                        from shipment_oop import TextNormalizer
                        up = TextNormalizer.up(line)
                        if "SUBE" in up and ("KODU" in up or "ADI" in up):
                            part = line.split(":", 1)[-1] if ":" in line else line
                            part = part.strip().strip('"').strip("'").strip()
                            if "-" in part:
                                part = part.split("-", 1)[-1]
                            part = part.strip()
                            m = re.search(r"\(([^)]+)\)", part)
                            if m:
                                branch_name = m.group(1).strip()
                            elif part:
                                branch_name = part
                            break
            
            if branch_name:
                engine = BranchDecisionEngine(branch_name)
                if engine.requires_day_selection() and not sheet_hint:
                    possible_sheets = engine.get_possible_sheets()
                    safe_log_insert(log_widget, f"[INFO] '{branch_name}' şubesi için gün seçimi gerekli.\n")
                    
                    # Show modal dialog for day selection
                    selected_day = show_day_selection_dialog(branch_name, possible_sheets)
                    
                    if not selected_day:
                        status_label.config(text=f"❌ İşlem iptal edildi{branch_info} (gün seçilmedi)")
                        safe_log_insert(log_widget, f"[INFO] Kullanıcı gün seçimini iptal etti{branch_info}.\n")
                        return
                    
                    sheet_hint = selected_day
                    safe_log_insert(log_widget, f"[INFO] Seçilen gün: {selected_day}\n")
                    
                if sheet_hint:
                    # Map user-friendly name to actual sheet name
                    sheet_hint = SHEET_NAME_MAPPING.get(sheet_hint, sheet_hint)
        except Exception as e:
            safe_log_insert(log_widget, f"[WARN] Branch kontrolü başarısız: {e}\n")
        
        # Prepare branch info message for display
        branch_info = f" [{branch_name}]" if branch_name else ""
        
        status_label.config(text=f"⏳ Başladı{branch_info}: CSV okunuyor...")
        safe_log_insert(log_widget, f"[INFO] İşlem başladı{branch_info}\n")
        safe_log_insert(log_widget, f"[INFO] CSV Dosyası: {csv_filename}\n")
        safe_log_insert(log_widget, f"[INFO] CSV okunuyor ve eşleştirilecek.\n")
        # Aşama: Çalıştır
        try:
            safe_log_insert(log_widget, f"[STEP] Tatlı eşleştirme başlıyor{branch_info}...\n")
            t_match, t_unmatch = coord.process_tatli(csv_path, output_path="sevkiyat_tatlı.xlsx", sheet_hint=sheet_hint)
            status_label.config(text=f"⏳ Tatlı tamamlandı{branch_info}: {t_match} yazıldı. Donuk hazırlanıyor...")
            safe_log_insert(log_widget, f"[STEP] Donuk eşleştirme başlıyor{branch_info}...\n")
            d_match, d_unmatch = coord.process_donuk(csv_path, output_path="sevkiyat_donuk.xlsx", sheet_hint=sheet_hint)
            status_label.config(text=f"⏳ Donuk tamamlandı{branch_info}: {d_match} yazıldı. Lojistik hazırlanıyor...")
            safe_log_insert(log_widget, f"[STEP] Lojistik eşleştirme başlıyor{branch_info}...\n")
            l_match, l_unmatch = coord.process_lojistik(csv_path, output_path="sevkiyat_lojistik.xlsx", sheet_hint=sheet_hint)
            summary = {
                "tatli": {"matched": t_match, "unmatched": t_unmatch, "file": "sevkiyat_tatlı.xlsx"},
                "donuk": {"matched": d_match, "unmatched": d_unmatch, "file": "sevkiyat_donuk.xlsx"},
                "lojistik": {"matched": l_match, "unmatched": l_unmatch, "file": "sevkiyat_lojistik.xlsx"},
            }
        except Exception as e:
            safe_log_insert(log_widget, f"[ERR-E1] run_all aşamasında hata: {e}\n")
            status_label.config(text="❌ Hata: [E1] Koordinatör çalıştırma başarısız")
            raise
        # Tarih hücresini sadece Tatlı dosyasında güncelle
        try:
            format_today_in_workbook(summary["tatli"]["file"])
        except Exception as e:
            safe_log_insert(log_widget, f"[WARN-W1] Tarih yazılamadı ({summary['tatli']['file']}): {e}\n")
        status_label.config(text=(
            f"✅ İşlem tamamlandı!{branch_info}\n"
            f"Tatlı: {summary['tatli']['matched']}/{summary['tatli']['file']}  "
            f"Donuk: {summary['donuk']['matched']}/{summary['donuk']['file']}  "
            f"Lojistik: {summary['lojistik']['matched']}/{summary['lojistik']['file']}"
        ))
        safe_log_insert(log_widget, f"[DONE] Tüm eşleştirmeler tamamlandı{branch_info} ve dosyalar kaydedildi.\n")
        
        # Enhanced success message with branch info and csv filename
        success_msg = f"Tüm sevkiyat dosyaları başarıyla oluşturuldu!"
        if branch_name:
            success_msg = f"'{branch_name}' şubesi için tüm sevkiyat dosyaları başarıyla oluşturuldu!"
        
        success_msg += f"\n\n📄 CSV Dosyası: {csv_filename}"
        if branch_name:
            success_msg += f"\n🏢 Şube: {branch_name}"
        
        success_msg += f"\n\n📊 Yazılan Ürün Sayısı:\n"
        success_msg += f"• Tatlı: {summary['tatli']['matched']} ürün\n"
        success_msg += f"• Donuk: {summary['donuk']['matched']} ürün\n"
        success_msg += f"• Lojistik: {summary['lojistik']['matched']} ürün"
        
        total_items = summary['tatli']['matched'] + summary['donuk']['matched'] + summary['lojistik']['matched']
        success_msg += f"\n\n✅ Toplam: {total_items} ürün işlendi"
        
        # Only show popup if requested (for single file processing)
        if show_popup:
            messagebox.showinfo("İşlem Başarılı", success_msg)
    except Exception as e:
        status_label.config(text=f"❌ Hata{branch_info}: {e}")
        safe_log_insert(log_widget, f"[ERR-E0] Genel hata: {e}\n")
        
        # Enhanced error message with context
        error_msg = f"Bir hata oluştu:\n\n{e}"
        try:
            if 'csv_filename' in locals():
                error_msg += f"\n\n📄 CSV Dosyası: {csv_filename}"
            if 'branch_name' in locals() and branch_name:
                error_msg += f"\n🏢 Şube: {branch_name}"
        except:
            pass
        
        # Only show error popup if requested (for single file processing)
        if show_popup:
            messagebox.showerror("Hata", error_msg)
        else:
            # For batch processing, just re-raise to be caught by run_multiple_processes
            raise

def safe_log_insert(log_widget, message):
    """Safely insert message into log widget (handles disabled state)"""
    log_widget.config(state='normal')
    log_widget.insert(tk.END, message)
    log_widget.see(tk.END)
    log_widget.config(state='disabled')

def select_file(status_label, log_widget, izmir_day_var=None):
    # Allow multiple file selection
    file_paths = filedialog.askopenfilenames(filetypes=[("CSV Dosyası", "*.csv")])
    if file_paths:
        # Convert tuple to list
        file_list = list(file_paths)
        status_label.config(text=f"İşleniyor... ({len(file_list)} dosya)")
        log_widget.config(state='normal')
        log_widget.delete(1.0, tk.END)
        log_widget.config(state='disabled')
        # Process multiple files sequentially
        threading.Thread(target=run_multiple_processes, args=(file_list, status_label, log_widget, izmir_day_var)).start()

def run_multiple_processes(file_paths, status_label, log_widget, izmir_day_var=None):
    """Process multiple CSV files sequentially"""
    total_files = len(file_paths)
    successful = 0
    failed = 0
    failed_files = []
    
    safe_log_insert(log_widget, f"{'='*80}\n")
    safe_log_insert(log_widget, f"[INFO] Toplu İşlem Başladı: {total_files} dosya işlenecek\n")
    safe_log_insert(log_widget, f"{'='*80}\n\n")
    
    for idx, file_path in enumerate(file_paths, 1):
        filename = os.path.basename(file_path)
        
        safe_log_insert(log_widget, f"\n{'─'*80}\n")
        safe_log_insert(log_widget, f"[{idx}/{total_files}] İşleniyor: {filename}\n")
        safe_log_insert(log_widget, f"{'─'*80}\n")
        
        status_label.config(text=f"⏳ İşleniyor [{idx}/{total_files}]: {filename}")
        
        try:
            # Process this file (without showing individual popup)
            run_process(file_path, status_label, log_widget, izmir_day_var, show_popup=False)
            successful += 1
            safe_log_insert(log_widget, f"[{idx}/{total_files}] ✅ Başarılı: {filename}\n\n")
        except Exception as e:
            failed += 1
            failed_files.append((filename, str(e)))
            safe_log_insert(log_widget, f"[{idx}/{total_files}] ❌ Hata: {filename} - {e}\n\n")
    
    # Final summary
    safe_log_insert(log_widget, f"\n{'='*80}\n")
    safe_log_insert(log_widget, f"[ÖZET] Toplu İşlem Tamamlandı\n")
    safe_log_insert(log_widget, f"{'='*80}\n")
    safe_log_insert(log_widget, f"✅ Başarılı: {successful}/{total_files} dosya\n")
    if failed > 0:
        safe_log_insert(log_widget, f"❌ Başarısız: {failed}/{total_files} dosya\n")
        for fname, error in failed_files:
            safe_log_insert(log_widget, f"   - {fname}: {error}\n")
    
    status_label.config(text=f"✅ Toplu İşlem Tamamlandı: {successful}/{total_files} başarılı")
    
    # Show summary popup
    summary_msg = f"Toplu işlem tamamlandı!\n\n"
    summary_msg += f"📊 Özet:\n"
    summary_msg += f"• Toplam: {total_files} dosya\n"
    summary_msg += f"• Başarılı: {successful} dosya\n"
    if failed > 0:
        summary_msg += f"• Başarısız: {failed} dosya\n\n"
        summary_msg += "❌ Başarısız dosyalar:\n"
        for fname, _ in failed_files[:5]:  # Show first 5
            summary_msg += f"  - {fname}\n"
        if len(failed_files) > 5:
            summary_msg += f"  ... ve {len(failed_files)-5} dosya daha"
    
    if failed == 0:
        messagebox.showinfo("Toplu İşlem Başarılı", summary_msg)
    else:
        messagebox.showwarning("Toplu İşlem Tamamlandı (Bazı Hatalar)", summary_msg)

def on_drop(event, status_label, log_widget):
    # Handle multiple dropped files (tkinterdnd2 can send multiple files)
    raw_data = event.data.strip()
    
    # Parse multiple file paths (can be space-separated or with {})
    file_paths = []
    if '{' in raw_data:
        # Handle {file1} {file2} format
        import re
        file_paths = re.findall(r'\{([^}]+)\}', raw_data)
    else:
        # Handle space-separated or single file
        # Split by spaces but handle paths with spaces
        parts = raw_data.split()
        current_path = []
        for part in parts:
            current_path.append(part)
            potential_path = ' '.join(current_path)
            if os.path.exists(potential_path) and potential_path.lower().endswith('.csv'):
                file_paths.append(potential_path)
                current_path = []
        # If nothing was found, try as single path
        if not file_paths and raw_data:
            file_paths = [raw_data]
    
    # Filter only CSV files
    csv_files = [f for f in file_paths if f.lower().endswith('.csv') and os.path.exists(f)]
    
    if csv_files:
        if len(csv_files) == 1:
            # Single file - use original behavior
            status_label.config(text="İşleniyor...")
            log_widget.config(state='normal')
            log_widget.delete(1.0, tk.END)
            log_widget.config(state='disabled')
            threading.Thread(target=run_process, args=(csv_files[0], status_label, log_widget, None)).start()
        else:
            # Multiple files - use batch processing
            status_label.config(text=f"İşleniyor... ({len(csv_files)} dosya)")
            log_widget.config(state='normal')
            log_widget.delete(1.0, tk.END)
            log_widget.config(state='disabled')
            threading.Thread(target=run_multiple_processes, args=(csv_files, status_label, log_widget, None)).start()
    else:
        messagebox.showerror("Hata", "Lütfen geçerli CSV dosyası/dosyaları bırakın.")


def open_file(path: str):
    try:
        if os.path.exists(path):
            os.startfile(path)  # Windows
        else:
            messagebox.showerror("Hata", f"Dosya bulunamadı: {path}")
    except Exception as e:
        messagebox.showerror("Hata", str(e))

def show_update_window(parent=None):
    """Güncelleme penceresini gösterir"""
    update_window = tk.Toplevel(parent)
    update_window.title("Güncelleme Kontrolü")
    update_window.geometry("500x400")
    update_window.resizable(False, False)
    
    # Ana frame
    main_frame = tk.Frame(update_window)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    # Başlık
    title_label = tk.Label(main_frame, text="Güncelleme Kontrolü", font=("Arial", 16, "bold"))
    title_label.pack(pady=(0, 20))
    
    # Mevcut sürüm bilgisi
    current_version_frame = tk.Frame(main_frame)
    current_version_frame.pack(fill=tk.X, pady=(0, 10))
    tk.Label(current_version_frame, text="Mevcut Sürüm:", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
    tk.Label(current_version_frame, text=VERSION, font=("Arial", 10)).pack(side=tk.LEFT, padx=(10, 0))
    
    # Durum etiketi
    status_label = tk.Label(main_frame, text="Kontrol ediliyor...", fg="blue")
    status_label.pack(pady=(0, 10))
    
    # Progress bar
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(main_frame, variable=progress_var, maximum=100)
    progress_bar.pack(fill=tk.X, pady=(0, 10))
    
    # Log alanı
    log_text = scrolledtext.ScrolledText(main_frame, height=15, width=50)
    log_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
    
    # Butonlar
    button_frame = tk.Frame(main_frame)
    button_frame.pack(fill=tk.X, pady=(10, 0))
    
    check_button = tk.Button(button_frame, text="Kontrol Et", width=15)
    download_button = tk.Button(button_frame, text="İndir", width=15, state=tk.DISABLED)
    install_button = tk.Button(button_frame, text="Kur", width=15, state=tk.DISABLED)
    close_button = tk.Button(button_frame, text="Kapat", width=15)
    
    check_button.pack(side=tk.LEFT, padx=(0, 5))
    download_button.pack(side=tk.LEFT, padx=5)
    install_button.pack(side=tk.LEFT, padx=5)
    close_button.pack(side=tk.RIGHT)
    
    # Değişkenler
    update_info = {"latest_version": None, "download_url": None, "download_path": None}
    
    def log_message(message):
        log_text.insert(tk.END, f"{message}\n")
        log_text.see(tk.END)
        update_window.update()
    
    def check_updates():
        status_label.config(text="Kontrol ediliyor...", fg="blue")
        log_text.delete(1.0, tk.END)
        check_button.config(state=tk.DISABLED)
        
        def check_thread():
            try:
                log_message("GitHub'a bağlanılıyor...")
                latest_version, release_url = get_latest_version()
                
                if not latest_version:
                    status_label.config(text="❌ Bağlantı hatası", fg="red")
                    log_message("❌ GitHub'a bağlanılamadı!")
                    check_button.config(state=tk.NORMAL)
                    return
                
                log_message(f"✅ En son sürüm bulundu: {latest_version}")
                
                if is_newer_version(latest_version, VERSION):
                    status_label.config(text=f"🔄 Yeni sürüm mevcut: {latest_version}", fg="orange")
                    log_message(f"🔄 Yeni sürüm mevcut!")
                    log_message(f"   Mevcut: {VERSION}")
                    log_message(f"   Yeni: {latest_version}")
                    
                    # Download URL'ini al
                    try:
                        response = requests.get(GITHUB_API_URL, timeout=10)
                        if response.status_code == 200:
                            data = response.json()
                            assets = data.get('assets', [])
                            if assets:
                                best = select_best_asset(assets)
                                if best is not None:
                                    download_url = best.get('browser_download_url')
                                    update_info["latest_version"] = latest_version
                                    update_info["download_url"] = download_url
                                    download_button.config(state=tk.NORMAL)
                                    log_message(f"✅ İndirme hazır! (Seçilen: {best.get('name')})")
                                    # Eğer otomatik indirme ayarlıysa indir butonunu tetikle
                                    if AUTO_START_DOWNLOAD:
                                        try:
                                            # Güvenli GUI çağrısı: ana iş parçacığında invoke et
                                            update_window.after(0, download_button.invoke)
                                            log_message("🔁 Otomatik indirme başlatıldı...")
                                        except Exception:
                                            log_message("❗ Otomatik indirme başlatılamadı; lütfen 'İndir' butonuna basın.")
                                else:
                                    log_message("❌ İndirme dosyası bulunamadı!")
                            else:
                                log_message("❌ İndirme dosyası bulunamadı!")
                        else:
                            log_message("❌ Release bilgileri alınamadı!")
                    except Exception as e:
                        log_message(f"❌ Hata: {e}")
                else:
                    status_label.config(text="✅ Güncel sürüm", fg="green")
                    log_message("✅ Uygulamanız güncel!")
                
                check_button.config(state=tk.NORMAL)
                
            except Exception as e:
                status_label.config(text="❌ Hata oluştu", fg="red")
                log_message(f"❌ Hata: {e}")
                check_button.config(state=tk.NORMAL)
        
        threading.Thread(target=check_thread, daemon=True).start()
    
    def download_update():
        if not update_info["download_url"]:
            return
        
        download_button.config(state=tk.DISABLED)
        status_label.config(text="İndiriliyor...", fg="blue")
        log_message("📥 Güncelleme indiriliyor...")
        
        def download_thread():
            try:
                def progress_callback(progress):
                    progress_var.set(progress)
                    update_window.update()
                
                success = download_github_update(update_info["download_url"], progress_callback)
                
                if success:
                    status_label.config(text="✅ İndirme tamamlandı", fg="green")
                    log_message("✅ İndirme tamamlandı!")
                    install_button.config(state=tk.NORMAL)
                    update_info["download_path"] = "update.zip"
                    
                    # Otomatik kurulum başlat (eğer AUTO_START_DOWNLOAD aktifse)
                    if AUTO_START_DOWNLOAD:
                        log_message("🔁 Otomatik kurulum başlatılıyor...")
                        # Kısa bir gecikme ile kurulum butonunu tetikle
                        update_window.after(1000, install_button.invoke)
                else:
                    status_label.config(text="❌ İndirme hatası", fg="red")
                    log_message("❌ İndirme başarısız!")
                
                download_button.config(state=tk.NORMAL)
                
            except Exception as e:
                status_label.config(text="❌ İndirme hatası", fg="red")
                log_message(f"❌ Hata: {e}")
                download_button.config(state=tk.NORMAL)
        
        threading.Thread(target=download_thread, daemon=True).start()
    
    def install_update_now():
        if not update_info["download_path"] or not os.path.exists(update_info["download_path"]):
            return
        
        # Sadece manuel kurulumda onay iste (otomatik kurulumda direkt başlat)
        if not AUTO_START_DOWNLOAD:
            result = messagebox.askyesno(
                "Güncelleme Kurulumu", 
                "Güncelleme kurulacak ve uygulama yeniden başlatılacak.\n\nDevam etmek istiyor musunuz?"
            )
            if not result:
                return
        else:
            # Otomatik kurulumda bilgilendirme mesajı
            log_message("⚠️ Güncelleme 2 saniye içinde başlatılacak...")
            status_label.config(text="⏳ Güncelleme başlatılıyor...", fg="orange")
            update_window.update()
        
        install_button.config(state=tk.DISABLED)
        status_label.config(text="Kuruluyor...", fg="blue")
        log_message("🔧 Güncelleme kuruluyor...")
        
        def install_thread():
            try:
                # Kısa gecikme (kullanıcı mesajları okusun)
                import time
                if AUTO_START_DOWNLOAD:
                    time.sleep(1)
                
                success, bat_path, message = install_update()
                
                if not success:
                    status_label.config(text="❌ Kurulum hatası", fg="red")
                    log_message(f"❌ Kurulum hatası: {message}")
                    install_button.config(state=tk.NORMAL)
                    return
                
                frozen = getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS')
                
                if frozen and bat_path:
                    # Launch the VBScript updater (bat_path is actually vbs_path)
                    try:
                        log_message("🔄 Güncelleme scripti başlatılıyor...")
                        log_message("✅ Uygulama kapatılıyor ve güncelleme uygulanıyor...")
                        status_label.config(text="✅ Güncelleme başlatıldı", fg="green")
                        
                        # UI'yi güncelle
                        update_window.update_idletasks()
                        update_window.update()
                        
                        # Start VBScript updater silently with wscript
                        # VBScript is built into Windows, no dependencies needed
                        subprocess.Popen(
                            ['wscript.exe', bat_path],  # bat_path is actually vbs_path
                            shell=False,
                            creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS,
                            close_fds=True
                        )
                        
                        # Short delay to let updater start
                        time.sleep(0.5)
                        
                        # Force application termination
                        log_message("⏹️ Uygulama kapatılıyor (PID: {})...".format(os.getpid()))
                        
                        # Direct exit to release file handles
                        os._exit(0)
                        
                    except Exception as e:
                        log_message(f"❌ Updater başlatılamadı: {e}")
                        install_button.config(state=tk.NORMAL)
                        return
                else:
                    # Non-frozen case
                    status_label.config(text="✅ Kurulum tamamlandı", fg="green")
                    log_message("✅ Kurulum tamamlandı.")
                    log_message("ℹ️ Lütfen uygulamayı yeniden başlatın.")
                    install_button.config(state=tk.NORMAL)
                    
            except Exception as e:
                status_label.config(text="❌ Kurulum hatası", fg="red")
                log_message(f"❌ Hata: {e}")
                install_button.config(state=tk.NORMAL)

        threading.Thread(target=install_thread, daemon=True).start()
    
    def close_window():
        update_window.destroy()
    
    # Event handlers
    check_button.config(command=check_updates)
    download_button.config(command=download_update)
    install_button.config(command=install_update_now)
    close_button.config(command=close_window)
    
    # Otomatik kontrol başlat
    check_updates()


def main():
    # tkinterdnd2 varsa onun root'unu kullan
    if TK_DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    root.title("Bayi Sipariş -> Sevkiyat (Tatlı / Donuk / Lojistik)")
    root.geometry("800x600")
    root.minsize(600, 400)
    import sys
    if sys.platform == "win32":
        try:
            import ctypes
            myappid = u'bhu.tatli.sevkiyat.1.2.0'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass
    try:
        root.iconbitmap(resource_path(ICON_PATH))
    except Exception:
        pass

    root.grid_rowconfigure(3, weight=1)
    root.grid_columnconfigure(0, weight=1)

    info = (
        "1) Şubeden gelen CSV'yi seçin veya sürükleyip bırakın.\n"
        "2) Uygulama üç dosyayı üretir: sevkiyat_tatlı.xlsx, sevkiyat_donuk.xlsx, sevkiyat_lojistik.xlsx.\n"
        "3) İzmir/Kuşadası şubeleri için gün seçimi yapabilirsiniz.\n"
        "4) Aşağıdaki kısayollardan açabilir veya temizleyebilirsiniz."
    )
    label = tk.Label(root, text=info, wraplength=700, justify="left")
    label.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))

    # Gün seçimi artık otomatik modal dialog ile yapılıyor - eski dropdown kaldırıldı
    izmir_day_var = None  # Compatibility için None olarak bırak

    # Ana üst butonlar için frame - sadece 3 buton
    top_btn_frame = tk.Frame(root, pady=15)
    top_btn_frame.grid(row=1, column=0)
    
    # CSV Seç butonu
    select_btn = tk.Button(
        top_btn_frame, 
        text="📄 CSV Seç", 
        width=20, 
        height=2,
        font=("Arial", 11, "bold"),
        bg="#007BFF",
        fg="white",
        command=lambda: select_file(status_label, log_widget, izmir_day_var)
    )
    select_btn.grid(row=0, column=0, padx=10)
    
    # Tüm Dosyaları Temizle butonu
    def clear_all_files():
        confirm = messagebox.askyesno("Onay", "Tüm sevkiyat dosyalarını (Tatlı, Donuk, Lojistik) temizlemek istediğinize emin misiniz?")
        if not confirm:
            status_label.config(text="İşlem iptal edildi.")
            return
        try:
            cleared_total = 0
            results = []
            
            # Clear tatlı file
            if os.path.exists("sevkiyat_tatlı.xlsx"):
                cleared = clear_tatli_values("sevkiyat_tatlı.xlsx")
                cleared_total += cleared
                results.append(f"Tatlı: {cleared} hücre")
            
            # Clear donuk file
            if os.path.exists("sevkiyat_donuk.xlsx"):
                cleared = clear_donuk_values("sevkiyat_donuk.xlsx")
                cleared_total += cleared
                results.append(f"Donuk: {cleared} hücre")
            
            # Clear lojistik file
            if os.path.exists("sevkiyat_lojistik.xlsx"):
                cleared = clear_lojistik_values("sevkiyat_lojistik.xlsx")
                cleared_total += cleared
                results.append(f"Lojistik: {cleared} hücre")
            
            status_label.config(text=f"✅ Tüm dosyalar temizlendi! ({cleared_total} hücre)")
            safe_log_insert(log_widget, f"✅ Tüm dosyalar temizlendi!\n")
            for result in results:
                safe_log_insert(log_widget, f"   - {result}\n")
        except Exception as e:
            status_label.config(text=f"❌ Hata: {e}")
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")
    
    clear_all_btn = tk.Button(
        top_btn_frame, 
        text="🧹 Tümünü Temizle", 
        width=20, 
        height=2,
        font=("Arial", 11, "bold"),
        bg="#FFC107",
        fg="black",
        command=clear_all_files
    )
    clear_all_btn.grid(row=0, column=1, padx=10)
    
    # Güncelleme butonu
    update_btn = tk.Button(
        top_btn_frame, 
        text="🔄 Güncelleme Kontrol", 
        width=20, 
        height=2,
        font=("Arial", 11, "bold"),
        bg="#28A745",
        fg="white",
        command=lambda: show_update_window(root)
    )
    update_btn.grid(row=0, column=2, padx=10)
    
    # Open and Clear buttons frame - organized vertically for better UI consistency
    files_frame = tk.Frame(root)
    files_frame.grid(row=4, column=0, pady=(4, 8))
    
    # Helper function to create clear button for specific file
    def clear_donuk_file():
        confirm = messagebox.askyesno("Onay", "Donuk sevkiyat dosyasını temizlemek istediğinize emin misiniz?")
        if not confirm:
            return
        try:
            cleared = clear_donuk_values("sevkiyat_donuk.xlsx")
            status_label.config(text=f"✅ Donuk dosyası temizlendi! ({cleared} hücre)")
            safe_log_insert(log_widget, f"✅ Donuk dosyası temizlendi! ({cleared} hücre)\n")
        except Exception as e:
            status_label.config(text=f"❌ Hata: {e}")
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")
    
    def clear_lojistik_file():
        confirm = messagebox.askyesno("Onay", "Lojistik sevkiyat dosyasını temizlemek istediğinize emin misiniz?")
        if not confirm:
            return
        try:
            cleared = clear_lojistik_values("sevkiyat_lojistik.xlsx")
            status_label.config(text=f"✅ Lojistik dosyası temizlendi! ({cleared} hücre)")
            safe_log_insert(log_widget, f"✅ Lojistik dosyası temizlendi! ({cleared} hücre)\n")
        except Exception as e:
            status_label.config(text=f"❌ Hata: {e}")
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")
    
    # Tatlı column
    tk.Button(files_frame, text="Tatlı Dosyasını Aç", width=22, command=lambda: open_file("sevkiyat_tatlı.xlsx")).grid(row=0, column=0, padx=5, pady=2)
    tk.Button(files_frame, text="Tatlı Dosyasını Temizle", width=22, command=lambda: clear_all_records(status_label, log_widget)).grid(row=1, column=0, padx=5, pady=2)
    
    # Donuk column
    tk.Button(files_frame, text="Donuk Dosyasını Aç", width=22, command=lambda: open_file("sevkiyat_donuk.xlsx")).grid(row=0, column=1, padx=5, pady=2)
    tk.Button(files_frame, text="Donuk Dosyasını Temizle", width=22, command=clear_donuk_file).grid(row=1, column=1, padx=5, pady=2)
    
    # Lojistik column
    tk.Button(files_frame, text="Lojistik Dosyasını Aç", width=22, command=lambda: open_file("sevkiyat_lojistik.xlsx")).grid(row=0, column=2, padx=5, pady=2)
    tk.Button(files_frame, text="Lojistik Dosyasını Temizle", width=22, command=clear_lojistik_file).grid(row=1, column=2, padx=5, pady=2)

    status_label = tk.Label(root, text="", fg="blue", anchor="w")
    status_label.grid(row=2, column=0, sticky="ew", padx=10, pady=5)

    # Log widget - disabled state to prevent user editing
    log_widget = scrolledtext.ScrolledText(
        root, 
        state='disabled',
        wrap=tk.WORD,
        bg="#F8F9FA",
        fg="#212529",
        font=("Consolas", 9),
        relief=tk.FLAT,
        borderwidth=2
    )
    log_widget.grid(row=3, column=0, sticky="nsew", padx=10, pady=10)
    log_widget.update_idletasks()

    # Sürükle-bırak desteği (tkinterdnd2 ile)
    if TK_DND_AVAILABLE:
        root.drop_target_register(DND_FILES)
        def drop_event_handler(e):
            # TkinterDnD bazen event.data'yı tuple olarak gönderebilir, string'e çevir
            file_path = e.data if isinstance(e.data, str) else str(e.data)
            file_path = file_path.strip('{}')
            if file_path.lower().endswith('.csv'):
                status_label.config(text="İşleniyor...")
                log_widget.delete(1.0, tk.END)
                threading.Thread(target=run_process, args=(file_path, status_label, log_widget, izmir_day_var)).start()
            else:
                messagebox.showerror("Hata", "Lütfen bir CSV dosyası bırakın.")
        root.dnd_bind('<<Drop>>', drop_event_handler)

    footer = tk.Label(root, text=f"{DEVELOPER} | {VERSION}", fg="gray")
    footer.grid(row=6, column=0, columnspan=2, sticky="ew", pady=5)
    
    # Otomatik güncelleme kontrolü (arka planda)
    def auto_check_updates():
        try:
            # Son kontrol zamanını kontrol et
            last_check_file = "last_update_check.txt"
            should_check = True
            
            if os.path.exists(last_check_file):
                try:
                    with open(last_check_file, 'r') as f:
                        last_check_time = float(f.read().strip())
                    current_time = os.path.getmtime(__file__)  # Dosya değişiklik zamanı
                    if current_time - last_check_time < UPDATE_CHECK_INTERVAL:
                        should_check = False
                except:
                    pass
            
            if should_check:
                # Arka planda kontrol et
                threading.Thread(target=lambda: check_for_updates(silent=True), daemon=True).start()
                
                # Son kontrol zamanını kaydet
                try:
                    with open(last_check_file, 'w') as f:
                        f.write(str(os.path.getmtime(__file__)))
                except:
                    pass
        except:
            pass
    
    # 2 saniye sonra otomatik kontrol başlat
    root.after(2000, auto_check_updates)

    root.mainloop()

if __name__ == "__main__":
    main()