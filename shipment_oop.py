import os
import sys
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd  # pyright: ignore[reportMissingImports]
import openpyxl  # pyright: ignore[reportMissingModuleSource]
from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from datetime import datetime

# ------------------ Normalization ------------------
class TextNormalizer:
    @staticmethod
    def up(s: Optional[str]) -> str:
        if s is None:
            return ""
        s = str(s)
        # First apply Turkish character mapping BEFORE normalization
        tr_map = str.maketrans({
            "ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c",
            "İ": "I", "Ğ": "G", "Ü": "U", "Ş": "S", "Ö": "O", "Ç": "C",
        })
        s = s.translate(tr_map)
        s = s.upper()
        # Then normalize unicode
        try:
            import unicodedata
            s = unicodedata.normalize("NFKD", s)
        except Exception:
            pass
        s = s.strip()
        return s

# Constants
DATA_START_ROW = 3

# İzmir Bayi Listesi (kullanıcıdan)
IZMIR_BRANCHES = [
    "Mavibahçe", "Forum", "Point", "Folkart", "Efesus", "Gaziemir", "Balçova", "Hatay", "Folkart Vega", "İstasyon"
]
IZMIR_BRANCH_HINTS = [TextNormalizer.up(b) for b in IZMIR_BRANCHES]
KUŞADASI_HINTS = ["KUSADASI", "KUŞADASI", "AYDIN"]

# CSV'den gelen özel şube isimleri → Excel'deki standart isimler
# Bu mapping sayesinde HARMANDALI→EFESUS, FORUMAVM→FORUM gibi dönüşümler yapılır
BRANCH_NAME_MAPPING = {
    "HARMANDALI": "EFESUS",      # CSV: IZMIR(HARMANDALI) → Excel: EFESUS
    "FORUMAVM": "FORUM",          # CSV: IZMIR(FORUMAVM) → Excel: FORUM
    "FORUM AVM": "FORUM",         # CSV: IZMIR(FORUM AVM) → Excel: FORUM (space variant)
    "FOLKARTVEGA": "FOLKART VEGA", # CSV: IZMIR(FOLKARTVEGA) → Excel: FOLKART VEGA
    "ELYSIUM": "ELAZIG",          # CSV: ELYSIUM → Excel: ELAZIG
    "MEYDAN AVM": "MEYDAN",       # CSV: MEYDAN AVM → Excel: MEYDAN
    "SEYHAN": "ADANA",            # CSV: ADANA(SEYHAN) → Excel: ADANA
}

# Birden fazla sevkiyat günü olan şubeler ve hangi Excel sayfalarında bulundukları
# Format: {branch_normalized: [list of possible sheet names]}
MULTI_DAY_BRANCHES = {
    "MAVIBAHCE": ["SALI KARŞIYAKA", "KSK CUMARTESİ"],
    "FORUM": ["SALI KARŞIYAKA", "GÜZELBAHÇE", "KSK CUMARTESİ"],
    "FOLKART": ["SALI KARŞIYAKA", "CUMA İZMİR"],
    "EFESUS": ["SALI KARŞIYAKA", "CUMA İZMİR"],
    "ISTASYON": ["SALI İZMİR", "KSK CUMARTESİ"],
    "GAZIEMIR": ["SALI İZMİR", "CUMA İZMİR"],
    "BALCOVA": ["SALI İZMİR", "CUMA İZMİR"],
    "HATAY": ["SALI İZMİR", "CUMA İZMİR"],
    "FOLKART VEGA": ["SALI İZMİR", "CUMA İZMİR"],
    "KUSADASI": ["KUŞADASI-AYDIN", "KUŞADASI CMERT"],
}

# Sheet name mapping for user-friendly display
SHEET_NAME_MAPPING = {
    "Salı Karşıyaka": "SALI KARŞIYAKA",
    "Salı İzmir": "SALI İZMİR",
    "Cuma İzmir": "CUMA İZMİR",
    "Cumartesi KSK": "KSK CUMARTESİ",
    "Güzelbahçe": "GÜZELBAHÇE",
    "Kuşadası-Aydın": "KUŞADASI-AYDIN",
    "Kuşadası Çmert": "KUŞADASI CMERT",
}


# ------------------ CSV Reader ------------------
@dataclass
class OrderRow:
    stok_kodu: str
    miktar: float
    grup: str


class CsvOrderReader:
    def __init__(self, csv_path: str):
        self.csv_path = csv_path
        self.df = None  # type: Optional[pd.DataFrame]

    def load(self) -> None:
        # Most inputs start with two header lines, so header=2; fall back to 0
        try:
            self.df = pd.read_csv(self.csv_path, encoding="utf-8", delimiter=",", header=2)
        except Exception:
            self.df = pd.read_csv(self.csv_path, encoding="utf-8", delimiter=",", header=0)

    def _find_col(self, poss: Iterable[str]) -> Optional[str]:
        assert self.df is not None
        cols_up = {TextNormalizer.up(c): c for c in self.df.columns}
        for p in poss:
            up = TextNormalizer.up(p)
            if up in cols_up:
                return cols_up[up]
        # try partial
        for p in poss:
            up = TextNormalizer.up(p)
            for cu, orig in cols_up.items():
                if up in cu:
                    return orig
        return None

    def get_branch_name(self) -> Optional[str]:
        assert self.df is not None
        # Possible header names for branch in many CSV exports
        candidates = [
            "SUBE", "SUBE ADI", "SUBEADI", "BAYI", "BAYI ADI", "FIRMA ADI",
        ]
        for c in candidates:
            if c in self.df.columns:
                val = str(self.df[c].iloc[0]) if len(self.df[c]) else ""
                return val
        # Sometimes present in a separate first header row; best-effort parse
        return None

    def iter_rows(self) -> Iterable[OrderRow]:
        assert self.df is not None
        stok = self._find_col(["STOK KODU", "STOKKODU", "STOK KOD", "KOD"])
        miktar = self._find_col(["MIKTAR", "MİKTAR", "ADET"])
        grup = self._find_col(["GRUP", "GRUP ADI", "KATEGORI", "KATEGORI ADI"])
        if not stok or not miktar:
            raise ValueError("CSV'de 'Stok Kodu' veya 'Miktar' sütunu bulunamadı.")
        if not grup:
            # default group if unavailable
            grup = "TATLI"
        for _, r in self.df.iterrows():
            try:
                mikt = float(str(r[miktar]).replace(",", "."))
            except Exception:
                try:
                    mikt = float(r[miktar])
                except Exception:
                    continue
            if pd.isna(mikt) or mikt == 0:
                continue
            yield OrderRow(stok_kodu=str(r[stok]), miktar=mikt, grup=str(r[grup]) if grup in r else "TATLI")


# ------------------ Branch Decision Engine ------------------
class BranchDecisionEngine:
    def __init__(self, branch_name: Optional[str]):
        # Apply branch name mapping for special cases
        self.branch_name = self._apply_branch_mapping(branch_name or "")
        self.branch_up = TextNormalizer.up(self.branch_name)
    
    @staticmethod
    def _apply_branch_mapping(branch_name: str) -> str:
        """Apply special branch name mappings (e.g., HARMANDALI→EFESUS)."""
        if not branch_name:
            return branch_name
        
        branch_up = TextNormalizer.up(branch_name)
        # Normalized version without spaces/punctuation for robust comparisons
        import re
        def norm(s: str) -> str:
            return re.sub(r"[^A-Z0-9]", "", TextNormalizer.up(s))
        branch_norm = norm(branch_up)
        
        # Check exact matches first
        if branch_up in BRANCH_NAME_MAPPING:
            return BRANCH_NAME_MAPPING[branch_up]
        # Check normalized equality against mapping keys
        for k, v in BRANCH_NAME_MAPPING.items():
            if norm(k) == branch_norm:
                return v
        
        # Check if any mapping key is contained in the branch name
        for csv_name, excel_name in BRANCH_NAME_MAPPING.items():
            if csv_name in branch_up:
                return excel_name
            # normalized containment (handles FORUM AVM vs FORUMAVM etc.)
            if norm(csv_name) in branch_norm:
                return excel_name

        # Common suffix cleanup (e.g., " AVM") then retry exact/mapping
        if branch_up.endswith(" AVM"):
            trimmed = branch_up[:-4].strip()
            if trimmed in BRANCH_NAME_MAPPING:
                return BRANCH_NAME_MAPPING[trimmed]
            for k, v in BRANCH_NAME_MAPPING.items():
                if norm(k) == norm(trimmed):
                    return v
            # If trimming AVM leads to a known multi-day branch, return trimmed
            if trimmed in MULTI_DAY_BRANCHES:
                return trimmed
        
        return branch_name

    def segment(self) -> str:
        up = self.branch_up
        if any(h in up for h in KUŞADASI_HINTS):
            return "KUSADASI"
        if any(h in up for h in IZMIR_BRANCH_HINTS):
            return "IZMIR"
        return "GENEL"
    
    def requires_day_selection(self) -> bool:
        """Check if this branch appears in multiple sheets (needs day selection)."""
        return self.branch_up in MULTI_DAY_BRANCHES
    
    def get_possible_sheets(self) -> List[str]:
        """Get list of possible sheet names for this branch."""
        return MULTI_DAY_BRANCHES.get(self.branch_up, [])


# ------------------ Writers ------------------
class BaseExcelWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb: Optional[openpyxl.Workbook] = None
        self.ws: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

    def load(self) -> None:
        if os.path.exists(self.output_path):
            self.wb = openpyxl.load_workbook(self.output_path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            # basic header
            self.ws.cell(row=1, column=1, value="Sevkiyat Tarihi")
            self.ws.cell(row=2, column=1, value=datetime.today().strftime('%d.%m.%Y'))
            self.ws.cell(row=3, column=1, value="Stok Kodu")
            self.ws.cell(row=3, column=2, value="Miktar")
            self.ws.cell(row=3, column=3, value="Grup")
        
    def save(self) -> None:
        assert self.wb is not None
        self.wb.save(self.output_path)

    def clear_values(self) -> int:
        # Generic clear: from DATA_START_ROW onward, all sheets, values only (keep formulas and headers)
        assert self.wb is not None
        cleared = 0
        for ws in self.wb.worksheets:
            for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        continue
                    if cell.value not in (None, ""):
                        cell.value = None
                        cleared += 1
        return cleared


class SimpleListWriter(BaseExcelWriter):
    """Fallback writer for formats we don't have a template for.
    Appends rows under headers: Stok Kodu | Miktar | Grup
    """

    def append_rows(self, items: Iterable[OrderRow]) -> int:
        assert self.ws is not None
        # Find first empty row after header
        start_row = max(DATA_START_ROW, 4)
        row = start_row
        written = 0
        for it in items:
            # Direct write without safe_write to avoid unmerge-remerge issues
            # Note: We get the cell object first, then set value
            # This prevents "MergedCell" attribute errors
            cell1 = self.ws.cell(row=row, column=1)
            cell2 = self.ws.cell(row=row, column=2)
            cell3 = self.ws.cell(row=row, column=3)
            
            # Check if cells are merged (MergedCell type)
            from openpyxl.cell.cell import MergedCell
            if not isinstance(cell1, MergedCell):
                cell1.value = it.stok_kodu
            if not isinstance(cell2, MergedCell):
                cell2.value = it.miktar
            if not isinstance(cell3, MergedCell):
                cell3.value = it.grup
            
            row += 1
            written += 1
        return written


class LojistikTemplateWriter(BaseExcelWriter):
    """Template-aware writer for the logistics sheet like in lojistik.png.
    - Columns are branch names across the top (row 1 or 2)
    - Rows below are blank lines where free-form item text is appended under the branch column
    """

    def __init__(self, output_path: str, sheet_name: Optional[str] = None):
        super().__init__(output_path)
        self.sheet_name = sheet_name

    def load(self) -> None:
        super().load()
        assert self.wb is not None and self.ws is not None
        # Eğer belirli bir sayfa istenmişse ve varsa, ona geç
        if self.sheet_name and self.sheet_name in [ws.title for ws in self.wb.worksheets]:
            self.ws = self.wb[self.sheet_name]

    def _find_or_add_branch_col(self, branch_name: str) -> int:
        assert self.ws is not None
        up = TextNormalizer.up(branch_name)
        # search first 2 rows, allow fuzzy contains
        best_c = None
        best_score = 0
        for r in (1, 2):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if not v:
                    continue
                vv = TextNormalizer.up(str(v))
                if vv == up:
                    return c
                # fuzzy score: intersection length
                common = len(set(vv.split()) & set(up.split()))
                if common > best_score:
                    best_score = common
                    best_c = c
        if best_c is not None and best_score > 0:
            return best_c
        # not found: add new column at end of row 1
        col = self.ws.max_column + 1
        self.ws.cell(row=1, column=col, value=branch_name)
        return col

    def append_text_items(self, branch_name: str, items: Iterable[str]) -> int:
        assert self.ws is not None
        col = self._find_or_add_branch_col(self._canonical_branch(branch_name))
        # find first empty row below header, skipping merged regions
        def in_merge(r: int, c: int):
            for mr in self.ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = mr.bounds
                if min_row <= r <= max_row and min_col <= c <= max_col:
                    return (min_row, min_col, max_row, max_col)
            return None
        row = 2
        while True:
            bounds = in_merge(row, col)
            if bounds is not None:
                _, _, max_row, _ = bounds
                row = max_row + 1
                continue
            v = self.ws.cell(row=row, column=col).value
            if v in (None, ""):
                break
            row += 1
        written = 0
        for t in items:
            self.ws.cell(row=row, column=col, value=str(t).strip())
            row += 1
            written += 1
        return written

    def _canonical_branch(self, branch_name: str) -> str:
        up = TextNormalizer.up(branch_name)
        mapping = {
            "GÜZELBAHÇE": "GÜZELBAHÇE", "GUZELBAHCE": "GÜZELBAHÇE",
            "FORUM": "FORUM",
            "URLA": "URLA",
            "ILICA": "ILICA", "ILIÇA": "ILICA", "ILIÇA": "ILICA",
            "SEFERIHISAR": "SEFERIHISAR", "SEFERİHİSAR": "SEFERIHISAR",
            # İzmir listesiyle uyumlu ek varyantlar
            "MAVIBAHCE": "MAVIBAHCE", "MAVIBAHE": "MAVIBAHCE",
            "POINT": "POINT",
            "FOLKART": "FOLKART",
            "EFESUS": "EFESUS",
            "GAZIEMIR": "GAZIEMIR", "GAZİEMİR": "GAZIEMIR",
            "BALCOVA": "BALCOVA", "BALÇOVA": "BALCOVA",
            "HATAY": "HATAY",
            "FOLKART VEGA": "FOLKART VEGA", "VEGA": "FOLKART VEGA",
            "ISTASYON": "ISTASYON", "İSTASYON": "ISTASYON",
        }
        return mapping.get(up, branch_name)


class ImprovedLojistikWriter(LojistikTemplateWriter):
    """Improved lojistik writer with better branch matching and sheet selection."""
    
    def load(self) -> None:
        super().load()
        assert self.wb is not None and self.ws is not None
        
        # If sheet_hint is provided, try to find matching sheet
        if self.sheet_name:
            # Try exact match first
            if self.sheet_name in [ws.title for ws in self.wb.worksheets]:
                self.ws = self.wb[self.sheet_name]
                return
            
            # Try fuzzy matching for sheet names
            best_sheet = None
            best_score = 0
            hint_up = TextNormalizer.up(self.sheet_name)
            
            for ws in self.wb.worksheets:
                ws_up = TextNormalizer.up(ws.title)
                # Check for keyword matches
                hint_words = set(hint_up.split())
                ws_words = set(ws_up.split())
                common = len(hint_words & ws_words)
                
                if common > best_score:
                    best_score = common
                    best_sheet = ws
            
            if best_sheet and best_score > 0:
                self.ws = best_sheet
                return
        
        # If no specific sheet found, use the first available sheet
        if self.wb.worksheets:
            self.ws = self.wb.worksheets[0]
    
    def _find_sheet_for_branch(self, branch_name: str):
        """Find the Excel sheet that contains a column matching branch_name.
        
        Priority logic:
        1. If day/sheet is pre-selected (self.sheet_name exists):
           - First check if branch exists in the selected sheet
           - If found, use that sheet (day selection takes priority)
           - If NOT found, search ALL sheets for branch and use first match
        2. If no day selected: search all sheets for branch match
        """
        assert self.wb is not None
        
        branch_up = TextNormalizer.up(branch_name)
        
        # Helper function to check if branch exists in a sheet
        def branch_exists_in_sheet(ws) -> bool:
            for r in range(1, 4):
                for c in range(1, min(ws.max_column + 1, 30)):
                    val = ws.cell(r, c).value
                    if not val:
                        continue
                    val_up = TextNormalizer.up(str(val))
                    
                    # Check if branch matches (exact or partial)
                    if branch_up == val_up or branch_up in val_up or val_up in branch_up:
                        return True
            return False
        
        # If user selected a specific day/sheet
        if self.sheet_name and self.ws:
            # PRIORITY 1: Check if branch exists in user-selected sheet
            if branch_exists_in_sheet(self.ws):
                # Branch found in selected day - use it (day selection priority)
                return self.ws
            
            # PRIORITY 2: Branch NOT in selected day - search ALL sheets
            # This handles cases where CSV branch is not in the selected day
            for ws in self.wb.worksheets:
                if branch_exists_in_sheet(ws):
                    # Found branch in different sheet - use it
                    return ws
            
            # Branch not found anywhere - return current sheet as fallback
            return self.ws
        
        # No user selection: search all sheets for branch match
        for ws in self.wb.worksheets:
            if branch_exists_in_sheet(ws):
                return ws
        
        # If not found, return current sheet
        return self.ws

    def _find_or_add_branch_col(self, branch_name: str) -> int:
        assert self.ws is not None
        up = TextNormalizer.up(branch_name)
        
        # PASS 1: Exact match only (avoids FOLKART matching FOLKART VEGA)
        for r in range(1, min(4, self.ws.max_row + 1)):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if not v:
                    continue
                vv = TextNormalizer.up(str(v))
                if vv == up:
                    return c
        
        # PASS 2: Partial match with containment (like Tatli/Donuk logic)
        # This helps ELYSIUM find ELAZIG, MEYDAN find MEYDAN AVM etc.
        best_c = None
        best_score = 0
        for r in range(1, min(4, self.ws.max_row + 1)):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if not v:
                    continue
                vv = TextNormalizer.up(str(v))
                
                # Check substring containment in both directions
                if up in vv or vv in up:
                    # Prefer shorter match to avoid FOLKART matching FOLKART VEGA
                    # when both contain each other
                    score = min(len(up), len(vv))
                    if score > best_score:
                        best_score = score
                        best_c = c
        
        if best_c is not None:
            return best_c
        
        # If no match found, add new column
        col = self.ws.max_column + 1
        self.ws.cell(row=1, column=col, value=branch_name)
        return col

    def append_text_items(self, branch_name: str, items: Iterable[str], fallback_branch: Optional[str] = None) -> int:
        assert self.ws is not None and self.wb is not None
        
        # Use canonical branch name for better matching
        canonical_branch = self._canonical_branch(branch_name)
        canonical_fallback = self._canonical_branch(fallback_branch) if fallback_branch else None
        
        # Find the correct sheet for this branch (try primary first, then fallback)
        correct_sheet = self._find_sheet_for_branch(canonical_branch)
        
        # If primary not found in any sheet and fallback exists, try fallback
        if correct_sheet == self.ws and canonical_fallback:
            fallback_sheet = self._find_sheet_for_branch(canonical_fallback)
            if fallback_sheet != self.ws:
                correct_sheet = fallback_sheet
        
        if correct_sheet != self.ws:
            self.ws = correct_sheet
        
        # Try to find branch column (first try primary, then fallback if primary not found)
        col = None
        try_add_new = False
        
        # PASS 1: Try primary branch (exact + partial match)
        for r in range(1, min(4, self.ws.max_row + 1)):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if not v:
                    continue
                vv = TextNormalizer.up(str(v))
                branch_up = TextNormalizer.up(canonical_branch)
                # Exact match
                if vv == branch_up:
                    col = c
                    break
                # Partial match
                if branch_up in vv or vv in branch_up:
                    col = c
                    break
            if col:
                break
        
        # PASS 2: If primary not found and fallback exists, try fallback
        if col is None and canonical_fallback:
            for r in range(1, min(4, self.ws.max_row + 1)):
                for c in range(1, self.ws.max_column + 1):
                    v = self.ws.cell(row=r, column=c).value
                    if not v:
                        continue
                    vv = TextNormalizer.up(str(v))
                    fallback_up = TextNormalizer.up(canonical_fallback)
                    # Exact match
                    if vv == fallback_up:
                        col = c
                        break
                    # Partial match
                    if fallback_up in vv or vv in fallback_up:
                        col = c
                        break
                if col:
                    break
        
        # PASS 3: If neither found, add new column with fallback (or primary if no fallback)
        if col is None:
            col = self.ws.max_column + 1
            header_name = canonical_fallback or canonical_branch
            self.ws.cell(row=1, column=col, value=header_name)
            try_add_new = True
        
        # Find first empty row below header, skipping merged regions
        def in_merge(r: int, c: int):
            for mr in self.ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = mr.bounds
                if min_row <= r <= max_row and min_col <= c <= max_col:
                    return (min_row, min_col, max_row, max_col)
            return None
        
        # Start from row 3 to skip potential headers
        row = 3
        while row <= self.ws.max_row:
            bounds = in_merge(row, col)
            if bounds is not None:
                _, _, max_row, _ = bounds
                row = max_row + 1
                continue
            
            v = self.ws.cell(row=row, column=col).value
            if v in (None, "", " "):
                break
            row += 1
        
        # If we've reached the end, we might need to add more rows
        if row > self.ws.max_row:
            # Ensure we have enough rows
            for _ in range(10):  # Add 10 empty rows
                self.ws.cell(row=row, column=1, value=None)
                row += 1
        
        written = 0
        for t in items:
            if t and str(t).strip():
                cell = self.ws.cell(row=row, column=col)
                cell.value = str(t).strip()
                
                # Enable text wrapping for long product names
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                
                row += 1
                written += 1
        
        return written


# ------------------ Coordinator ------------------
class ShipmentCoordinator:
    """Coordinates parsing and writing into three workbooks following the design document.
    - sevkiyat_tatli.xlsx
    - sevkiyat_donuk.xlsx
    - sevkiyat_lojistik.xlsx
    """

    def __init__(self):
        pass

    def process_tatli(self, csv_path: str, output_path: str = "sevkiyat_tatlı.xlsx", sheet_hint: Optional[str] = None) -> Tuple[int, int]:
        """Delegate to existing advanced matcher if available; fall back to simple list."""
        try:
            from parse_gptfix import process_csv as legacy_tatli
            # legacy returns None or counts; normalize to (matched, unmatched)
            # CRITICAL: Pass sheet_hint to legacy_tatli for multi-day branch support
            res = legacy_tatli(csv_path, output_path=output_path, sheet_name=sheet_hint)
            if isinstance(res, tuple) and len(res) == 2:
                return res
            return (0, 0)
        except Exception:
            # Fallback: append all rows where group includes TATLI
            rdr = CsvOrderReader(csv_path)
            rdr.load()
            items = [r for r in rdr.iter_rows() if "TATLI" in TextNormalizer.up(r.grup)]
            wr = SimpleListWriter(output_path)
            wr.load()
            wr.append_rows(items)
            wr.save()
            return (len(items), 0)

    def process_donuk(self, csv_path: str, output_path: str = "sevkiyat_donuk.xlsx", sheet_hint: Optional[str] = None) -> Tuple[int, int]:
        try:
            from parse_gptfix import process_donuk_csv as legacy_donuk
            res = legacy_donuk(csv_path, output_path=output_path, sheet_name=sheet_hint)
            if isinstance(res, tuple) and len(res) == 2:
                return res
            # Fallback: estimate
            rdr = CsvOrderReader(csv_path)
            rdr.load()
            items = [r for r in rdr.iter_rows() if any(k in TextNormalizer.up(r.grup) for k in ["DONDURMA", "PASTA", "BOREK", "TATLI"])]
            return (len(items), 0)
        except Exception:
            rdr = CsvOrderReader(csv_path)
            rdr.load()
            items = [r for r in rdr.iter_rows() if any(k in TextNormalizer.up(r.grup) for k in ["DONDURMA", "PASTA", "BOREK", "TATLI"])]
            wr = SimpleListWriter(output_path)
            wr.load()
            wr.append_rows(items)
            wr.save()
            return (len(items), 0)

    def process_lojistik(self, csv_path: str, output_path: str = "sevkiyat_lojistik.xlsx", sheet_hint: Optional[str] = None) -> Tuple[int, int]:
        rdr = CsvOrderReader(csv_path)
        rdr.load()

        def read_branch_from_file(path: str) -> Tuple[Optional[str], Optional[str]]:
            """Extract branch name with primary and fallback like TATLI/DONUK logic.
            
            Returns (primary, fallback) where:
            - primary: Inner part from "OUTER(INNER)" format - should be tried first
            - fallback: Outer part - use if primary doesn't match
            
            Example:
            - "MANISA(45 PARK AVM)" -> returns ("45 PARK AVM", "MANISA")
            - "BALÇOVA" -> returns ("BALÇOVA", None)
            """
            try:
                with open(path, encoding="utf-8") as f:
                    for line in f:
                        up = TextNormalizer.up(line)
                        # Match "SUBE KODU", "SUBE KODU-ADI", "SUBE ADI", etc.
                        if "SUBE" in up and ("KODU" in up or "ADI" in up):
                            raw = line.split(":", 1)[-1] if ":" in line else line
                            # e.g. "242 - BALÇOVA" or "241 - MANISA(45 PARK AVM)"
                            part = raw.split("-", 1)[-1] if "-" in raw else raw
                            part = part.strip()
                            # Remove quotes if present
                            part = part.strip('"').strip("'").strip()
                            
                            # If parens exist, return (inner, outer) for priority matching
                            import re
                            m = re.search(r"^([^(]+)\(([^)]+)\)$", part)
                            if m:
                                outer = m.group(1).strip()
                                inner = m.group(2).strip()
                                # PRIMARY: inner (parantez içi) - try this first
                                # FALLBACK: outer (parantez dışı) - try if primary fails
                                return (inner, outer)
                            
                            # No parens - return single value for both
                            if part.upper().endswith(" DEPO"):
                                part = part[:-5].strip()
                            return (part, None)
            except Exception:
                pass
            return (None, None)

        # Get both primary and fallback branch names
        branch_primary_raw, branch_fallback_raw = read_branch_from_file(csv_path)
        if not branch_primary_raw and not branch_fallback_raw:
            branch_primary_raw = rdr.get_branch_name() or "GENEL"
        
        # Apply branch name mapping (e.g., HARMANDALI → EFESUS) for consistency
        branch_primary = BranchDecisionEngine._apply_branch_mapping(branch_primary_raw) if branch_primary_raw else None
        branch_fallback = BranchDecisionEngine._apply_branch_mapping(branch_fallback_raw) if branch_fallback_raw else None
        
        # Use primary if available, otherwise fallback
        branch_name = branch_primary or branch_fallback or "GENEL"
        # Groups mapped as per design doc: SARF MALZEME, KURABIYE, CIKOLATA - HEDIYELIK, ICECEK
        include_keys = ["SARF", "KURABIYE", "CIKOLATA", "HEDIYELIK", "ICECEK", "İCECEK", "İÇECEK"]
        rows = [r for r in rdr.iter_rows() if any(k in TextNormalizer.up(r.grup) for k in include_keys)]

        def clean_display(stok: str) -> str:
            import re
            s = str(stok)
            # remove {...}
            s = re.sub(r"\{[^}]*\}", "", s)
            
            # Clean parentheses content - keep only for specific cases
            # Keep: T-shirt sizes (S, M, L, XL, XXL), colors, and sauce types
            def should_keep_parens(content: str) -> bool:
                content_up = content.upper()
                # Keep if contains t-shirt sizes
                if any(size in content_up for size in ['S)', 'M)', 'L)', 'XL)', 'XXL)', 'BEDEN']):
                    return True
                # Keep if contains color indicators
                if any(color in content_up for color in ['BEYAZ', 'SİYAH', 'MAVİ', 'KIRMIZI', 'YEŞİL', 'SARI', 'GRİ', 'KAHVE', 'MOR', 'TURUNCU', 'PEMBE', 'RENK']):
                    return True
                # Keep if it's sauce type (contains SOS and a product name)
                if 'SOS' in content_up and any(prod in content_up for prod in ['CHOCOLATE', 'CIKOLATA', 'KARAMEL', 'FRAMBUAZ', 'CILEKTE', 'ÇİLEK', 'ORMAN', 'MEYVELI', 'FISTIK', 'ANTEP']):
                    return True
                return False
            
            # Process parentheses: remove unwanted, keep wanted
            def clean_parens(text: str) -> str:
                result = text
                # Find all parentheses content
                pattern = r'\([^)]*\)'
                matches = re.finditer(pattern, text)
                
                for match in reversed(list(matches)):  # Reverse to maintain positions
                    content = match.group(0)
                    inner = content[1:-1]  # Remove outer parens
                    
                    if should_keep_parens(inner):
                        # Keep this parenthesis
                        continue
                    else:
                        # Remove this parenthesis
                        result = result[:match.start()] + result[match.end():]
                
                return result
            
            s = clean_parens(s)
            s = re.sub(r"\s+", " ", s).strip()
            return s

        # Convert to text lines.
        # LOJISTIK-1: Remove static "ADET" suffix for all lojistik products.
        # LOJISTIK-2: Special case for "ACI BIBER RECELI" size-based unit mapping:
        #   ACI BİBER REÇELİ(280 GR) -> "{qty} KAVANOZ"
        #   ACI BİBER REÇELİ(5 KG)   -> "{qty} KOVA"
        lines: List[str] = []
        for r in rows:
            original = r.stok_kodu
            name = clean_display(original)
            qty = int(r.miktar) if float(r.miktar).is_integer() else r.miktar

            up_orig = TextNormalizer.up(original)
            # Detect Acı Biber Reçeli variants BEFORE parentheses removal
            unit_suffix = ""
            if "ACI" in up_orig and "BIBER" in up_orig and "RECELI" in up_orig:
                import re
                m = re.search(r"\([^)]*\)", original)
                parens_content = m.group(0)[1:-1] if m else ""
                up_par = TextNormalizer.up(parens_content)
                # Map size → unit text
                if "280" in up_par:
                    unit_suffix = " KAVANOZ"
                elif ("5" in up_par and "KG" in up_par) or "5KG" in up_par:
                    unit_suffix = " KOVA"
            # Build line (no generic ADET suffix anymore)
            lines.append(f"{name} - {qty}{unit_suffix}")
        
        # Use improved lojistik writer
        wr = ImprovedLojistikWriter(output_path, sheet_name=sheet_hint)
        wr.load()
        # Pass both primary and fallback for branch matching (like TATLI/DONUK)
        # Use primary as main branch name, fallback as backup
        actual_primary = branch_primary or branch_fallback or "GENEL"
        count = wr.append_text_items(actual_primary, lines, fallback_branch=branch_fallback if branch_primary else None)
        wr.save()
        return (count, 0)

    def run_all(self, csv_path: str, izmir_day_sheet: Optional[str] = None) -> Dict[str, Dict[str, int]]:
        out = {}
        matched, unmatched = self.process_tatli(csv_path, output_path="sevkiyat_tatlı.xlsx", sheet_hint=izmir_day_sheet)
        out["tatli"] = {"matched": matched, "unmatched": unmatched, "file": "sevkiyat_tatlı.xlsx"}
        m2, u2 = self.process_donuk(csv_path, output_path="sevkiyat_donuk.xlsx", sheet_hint=izmir_day_sheet)
        out["donuk"] = {"matched": m2, "unmatched": u2, "file": "sevkiyat_donuk.xlsx"}
        m3, u3 = self.process_lojistik(csv_path, output_path="sevkiyat_lojistik.xlsx", sheet_hint=izmir_day_sheet)
        out["lojistik"] = {"matched": m3, "unmatched": u3, "file": "sevkiyat_lojistik.xlsx"}
        return out


# ------------------ Utilities for GUI ------------------

def clear_workbook_values(path: str) -> int:
    """Generic clear - clears all non-formula values. Use specific clear functions instead."""
    if not os.path.exists(path):
        return 0
    wb = openpyxl.load_workbook(path)
    cleared = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
    wb.save(path)
    return cleared


def clear_tatli_values(path: str) -> int:
    """Clear tatlı file data cells and sepet (basket) counts in row 1.
    
    Clears:
    1. Row 1 sepet values (basket counts) for each branch
    2. Data cells (TEPSI, ADET columns) starting from row 3
    
    Preserves:
    - Product names (column 1)
    - Headers (row 2)
    - Formulas
    - Merged cell structures
    """
    if not os.path.exists(path):
        return 0
    
    wb = openpyxl.load_workbook(path)
    cleared = 0
    
    for ws in wb.worksheets:
        # Step 1: Read branch columns from row 2 headers
        subeler = {}
        for cell in ws[2][1:]:  # Skip first column (product names)
            if cell.value:
                sube_ad = str(cell.value).strip()
                # Each branch has 4 columns: TEPSI, TEPSI_2, ADET, ADET_2
                subeler[sube_ad] = {
                    "tepsi": cell.column,
                    "tepsi_2": cell.column + 1,
                    "adet": cell.column + 2,
                    "adet_2": cell.column + 3
                }
        
        # Step 2: Clear row 1 sepet values for each branch
        # Sepet is written to the first column of each branch (TEPSI column)
        for sube in subeler.values():
            sepet_col = sube["tepsi"]
            
            # Find the cell to clear (handle merged cells properly)
            # mr.bounds returns (min_col, min_row, max_col, max_row)
            target_cell = ws.cell(row=1, column=sepet_col)
            
            # Check if this cell is part of a merged range
            for mr in ws.merged_cells.ranges:
                # bounds format: (min_col, min_row, max_col, max_row)
                if (mr.min_row <= 1 <= mr.max_row) and (mr.min_col <= sepet_col <= mr.max_col):
                    # This cell is in a merged range, use the master (top-left)
                    target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
                    break
            
            # Clear value if it exists
            if target_cell.value not in (None, ""):
                target_cell.value = None
                cleared += 1
        
        # Step 3: Clear data cells (rows 3+)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=1):
            ana_cell = row[0]
            if not ana_cell.value:
                continue
            
            # Skip special rows (headers like "SIPARIS TARIHI", etc.)
            ana_ad = str(ana_cell.value).upper()
            skip_keywords = ["SIPARIS TARIHI", "SIPARIS ALAN", "TESLIM TARIHI", "TEYID EDEN"]
            if any(ana_ad.startswith(k) or ana_ad == k for k in skip_keywords):
                continue
            
            # Clear TEPSI, TEPSI_2, ADET, ADET_2 cells for this product
            for sube in subeler.values():
                for col in [sube["tepsi"], sube["tepsi_2"], sube["adet"], sube["adet_2"]]:
                    cell = ws.cell(row=ana_cell.row, column=col)
                    
                    # Skip formulas
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        continue
                    
                    # Clear value
                    if cell.value not in (None, ""):
                        cell.value = None
                        cleared += 1
    
    wb.save(path)
    return cleared


def clear_lojistik_values(path: str) -> int:
    """Clear only data cells in lojistik file, preserving branch headers and yellow-highlighted cells.
    
    Yellow-highlighted cells contain permanent items (demirbaş) that should never be cleared.
    """
    if not os.path.exists(path):
        return 0
    
    wb = openpyxl.load_workbook(path)
    cleared = 0
    
    for ws in wb.worksheets:
        # Find branch columns in first 2 rows
        branch_cols = set()
        for r in range(1, 3):  # rows 1-2
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip():
                    # This is a branch header column
                    branch_cols.add(c)
        
        # Clear data rows (from row 3 onwards) in branch columns only
        for c in branch_cols:
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=c)
                
                # Skip formulas
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                
                # CRITICAL: Skip yellow-highlighted cells (demirbaş items)
                # Check if cell has yellow fill color
                if cell.fill and cell.fill.start_color:
                    # Yellow colors: FFFF00 (pure yellow), FFFFFF00 (yellow), etc.
                    color_rgb = cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else None
                    if color_rgb:
                        # Check if it's a yellow color (starts with FFFF or 00FFFF)
                        color_str = str(color_rgb).upper()
                        # Yellow variations: FFFFFF00, FFFF00, FFFFE0 (light yellow), etc.
                        if 'FFFF' in color_str[:6] or color_str.startswith('00FFFF'):
                            continue  # Skip this cell - it's yellow highlighted
                
                # Clear non-empty cells
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
    
    wb.save(path)
    return cleared


def clear_donuk_values(path: str) -> int:
    """Clear quantity/unit values from donuk file, preserving product names and structure.
    
    Comprehensive cleaning:
    1. Skip first 2 rows (headers)
    2. Clear numeric cells (integers, floats)
    3. Clear text cells that are pure numbers (e.g., "5", "10.5")
    4. Clean text cells containing qty/unit patterns:
       - "KÜNEFE    2 SPT." → "KÜNEFE"
       - "BEYAZ EKMEK    2 KL." → "BEYAZ EKMEK"
       - "5 KL." → "" (clear pure qty cells)
       - "ROKOKO =" → "ROKOKO =" (preserve "=" for ROKOKO/EKLER)
    5. Preserve formulas starting with "="
    """
    if not os.path.exists(path):
        return 0
    
    import re
    wb = openpyxl.load_workbook(path)
    cleared = 0
    
    for ws in wb.worksheets:
        # Process all cells starting from row 3
        for r in range(3, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                
                # Skip formulas
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                
                # Case 1: Pure numeric value - clear it
                if isinstance(cell.value, (int, float)):
                    cell.value = None
                    cleared += 1
                    continue
                
                # Case 2: Process text cells
                if isinstance(cell.value, str) and cell.value.strip():
                    original_value = cell.value
                    
                    # Case 2a: String that can be parsed as pure number - clear it
                    try:
                        float(cell.value.replace(",", "."))
                        cell.value = None
                        cleared += 1
                        continue
                    except:
                        pass
                    
                    # Case 2b: Pure qty/unit cells like "2 KL.", "5 SPT.", "10 TEPSI", "3 KOLİ"
                    # These should be cleared entirely
                    if re.match(r'^\s*\d+(?:[\.\,]\d+)?\s*(?:SPT\.|KL\.|TEPSI|TEPSİ|KOLİ|KOLI)\s*$', cell.value, re.IGNORECASE):
                        cell.value = None
                        cleared += 1
                        continue
                    
                    # Case 2c: Text containing qty/unit patterns (including KOLİ for DOSIDO)
                    # Check if contains quantity markers
                    has_qty_unit = re.search(r'\d+\s*(?:SPT\.|KL\.|TEPSI|TEPSİ|KOLİ|KOLI)', cell.value, re.IGNORECASE)
                    has_trailing_number = re.search(r'\s+\d+\s*$', cell.value)
                    
                    if has_qty_unit or has_trailing_number:
                        val_up = TextNormalizer.up(cell.value)
                        
                        # Special case: ROKOKO and EKLER with "=" should preserve the "="
                        # Pattern: "ROKOKO = 5" → "ROKOKO =", "EKLER = 10" → "EKLER ="
                        if ("ROKOKO" in val_up or "EKLER" in val_up) and "=" in cell.value:
                            # Keep product name and "=", remove only the number after "="
                            cleaned = re.sub(r'(=)\s*\d+(?:[\.\,]\d+)?\s*(?:SPT\.|KL\.|TEPSI|TEPSİ|KOLİ|KOLI)?', r'\1', cell.value, flags=re.IGNORECASE).strip()
                            if cleaned != original_value:
                                cell.value = cleaned
                                cleared += 1
                        else:
                            # General case: Remove trailing qty/unit patterns (including KOLİ for DOSIDO)
                            # Pattern: "KÜNEFE    2 SPT." → "KÜNEFE", "DOSİDO 5 KOLİ" → "DOSİDO"
                            cleaned = re.sub(r'(\s*[0-9]+(?:[\.\,][0-9]+)?\s*(?:SPT\.|KL\.|TEPSI|TEPSİ|KOLİ|KOLI))+$', '', cell.value, flags=re.IGNORECASE).strip()
                            # Also remove bare trailing numbers: "KÜNEFE  5" → "KÜNEFE"
                            cleaned = re.sub(r'\s+\d+\s*$', '', cleaned).strip()
                            
                            if cleaned != original_value:
                                # If cleaned result is empty or just whitespace, clear the cell
                                if not cleaned:
                                    cell.value = None
                                else:
                                    cell.value = cleaned
                                cleared += 1
    
    wb.save(path)
    return cleared


def format_today_in_workbook(path: str) -> None:
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        # Merge-safe write to A2
        target_r, target_c = 2, 1
        cell = ws.cell(row=target_r, column=target_c)
        def master(sheet, r, c):
            for mr in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = mr.bounds
                if min_row <= r <= max_row and min_col <= c <= max_col:
                    return sheet.cell(row=min_row, column=min_col)
            return sheet.cell(row=r, column=c)
        m = master(ws, target_r, target_c)
        m.value = datetime.today().strftime('%d.%m.%Y')
    wb.save(path)
